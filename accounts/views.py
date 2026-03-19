from calendar import month_name
from datetime import date, datetime
from multiprocessing import context
from pprint import pprint
from tkinter.ttk import Entry
from django.http import Http404, JsonResponse
from turtle import left
from urllib import request
from django.http import HttpResponse, HttpResponseForbidden
from django.contrib.auth.decorators import login_required
from django.shortcuts import redirect, render, get_object_or_404
from django.contrib.auth.forms import PasswordChangeForm
from django.contrib.auth import update_session_auth_hash
from django.contrib import messages

from accounts.permissions import ALL_PERMISSIONS
from .models import ULB, UserPermission,User
from .forms import RootCreateUserForm
from django.contrib.auth import get_user_model
from django.views.decorators.http import require_POST
from .permissions import ALL_PERMISSIONS
from django.utils import timezone
from accounts.models import Ledger, ULB
from .models import ULB, Ledger, LedgerGroup, SubGroup, MainGroup, HeadGroupChoices
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook
from django.db.models import Max
from django.db import transaction
from .models import (Transaction, TransactionEntry, ReceiptUCDetails, PaymentVendorDetails, ULB, Ledger, VoucherType,)
from django.db.models import F, Q, Min    
from django.contrib.auth.forms import AuthenticationForm




# ---------- Role decorator ----------

def role_required(allowed_roles):
    """
    allowed_roles: list of role codes, e.g. ['ROOT_DEV', 'DEV']
    """
    def decorator(view_func):
        def _wrapped(request, *args, **kwargs):
            if not request.user.is_authenticated:
                return redirect('login')
            if getattr(request.user, 'role', None) not in allowed_roles:
                return HttpResponseForbidden("Not allowed")
            return view_func(request, *args, **kwargs)
        return _wrapped
    return decorator


# ---------- Helper: check custom permission code ----------

def user_has_code(request, code):
    return UserPermission.objects.filter(user=request.user, code=code).exists()


# ---------- Helper: render with permissions + section ----------

def render_dashboard(request, template_name, context=None, active_section='dashboard'):
    if context is None:
        context = {}
    allowed_codes = set(
        UserPermission.objects.filter(user=request.user)
        .values_list('code', flat=True)
    )
    context.setdefault('allowed_codes', allowed_codes)
    context.setdefault('active_section', active_section)
    return render(request, template_name, context)


# ---------- Login view (common for all roles) ----------

def login_view(request):
    if request.method == "POST":
        form = AuthenticationForm(request, data=request.POST)
        if form.is_valid():
            user = form.get_user()
            login(request, user)
            # Always send to the dashboard router
            return redirect('dashboard')
    else:
        form = AuthenticationForm(request)

    return render(request, "accounts/login.html", {"form": form})


# ---------- Dashboard router ----------

@login_required
def dashboard(request):
    user = request.user
    role = getattr(user, 'role', None)

    if role == 'ROOT_DEV':
        return root_developer_dashboard(request)
    elif role == 'DEV':
        return developer_dashboard(request)
    elif role == 'ADMIN':
        return admin_dashboard(request)
    else:  # USER or anything else
        return user_dashboard(request)


# ---------- Dashboards ----------

@login_required
@role_required(['ROOT_DEV'])   # ROOT only
def root_developer_dashboard(request):
    return render_dashboard(
        request,
        'dashboards/base_root_dashboard.html',
        active_section='dashboard',
    )


@login_required
@role_required(['DEV'])  # DEV only
def developer_dashboard(request):
    return render_dashboard(
        request,
        'dashboards/base_root_dashboard.html',
        active_section='dashboard',
    )


@login_required
@role_required(['ADMIN'])
def admin_dashboard(request):
    return render_dashboard(
        request,
        'dashboards/base_root_dashboard.html',
        active_section='dashboard',
    )


@login_required
@role_required(['USER'])
def user_dashboard(request):
    return render_dashboard(
        request,
        'dashboards/base_root_dashboard.html',
        active_section='dashboard',
    )


# ---------- Change password ----------

@login_required
@role_required(['ROOT_DEV', 'DEV', 'ADMIN', 'USER'])
def change_password(request):
    if request.method == 'POST':
        form = PasswordChangeForm(user=request.user, data=request.POST)
        if form.is_valid():
            user = form.save()
            update_session_auth_hash(request, user)
            messages.success(request, "Password changed successfully.")
            return redirect('dashboard')
    else:
        form = PasswordChangeForm(user=request.user)

    return render_dashboard(
        request,
        'dashboards/change_password.html',
        {'form': form},
        active_section='dashboards',
    )
# ------------------------- ULB: create------------------------------
@login_required
@role_required(['ROOT_DEV', 'DEV'])
def create_ulb(request):
    # enforce custom permission code
    if not user_has_code(request.user, 'MENU_CREATE_ULB'):
        return HttpResponseForbidden("You do not have permission for this action.")

    if request.method == 'POST':
        ulb_name = request.POST.get('ulb_name')
        ulb_type = request.POST.get('ulb_type')
        email = request.POST.get('email')
        pan_no = request.POST.get('pan_no') or None
        tin_no = request.POST.get('tin_no')
        gst_no = request.POST.get('gst_no')
        land_mark = request.POST.get('land_mark')
        district = request.POST.get('district')
        state = request.POST.get('state')
        country = request.POST.get('country')
        date_of_creation = request.POST.get('date_of_creation')
        ulb_code = request.POST.get('ulb_code')  # NEW

        # basic required check
        if not all([
            ulb_name, ulb_type, email, tin_no, gst_no,
            land_mark, district, state, country, date_of_creation,
            ulb_code  # NEW required
        ]):
            messages.error(request, 'Please fill all required fields.')
            return render_dashboard(
                request,
                'create_ulb/create_ulb.html',
                active_section='ulb',
            )

        # duplicate protection
        duplicate = ULB.objects.filter(
            ulb_name__iexact=ulb_name.strip(),
            email__iexact=email.strip(),
            pan_no=pan_no,
            tin_no=tin_no.strip(),
            gst_no=gst_no.strip(),
        ).exists()

        if duplicate:
            messages.error(
                request,
                'A ULB with the same ULB Name, Email ID, PAN No., '
                'TIN No., and GST No. already exists.'
            )
            return render_dashboard(
                request,
                'create_ulb/create_ulb.html',
                active_section='ulb',
            )

        # ensure ULB code is unique
        if ULB.objects.filter(code__iexact=ulb_code.strip()).exists():
            messages.error(
                request,
                'This ULB Code is already in use. Please choose a different code.'
            )
            return render_dashboard(
                request,
                'create_ulb/create_ulb.html',
                active_section='ulb',
            )

        # create record
        ULB.objects.create(
            ulb_name=ulb_name,
            ulb_type=ulb_type,
            email=email,
            pan_no=pan_no,
            tin_no=tin_no,
            gst_no=gst_no,
            land_mark=land_mark,
            district=district,
            state=state,
            country=country,
            date_of_creation=date_of_creation,
            code=ulb_code.strip(),  # NEW
        )

        messages.success(request, 'ULB created successfully.')
        return redirect('create_ulb')

    return render_dashboard(
        request,
        'create_ulb/create_ulb.html',
        active_section='ulb',
    )

#------------------------- ULB: open ------------------------------
@login_required
@role_required(['ROOT_DEV', 'DEV', 'ADMIN', 'USER'])
def open_ulb(request):
    if not user_has_code(request.user, 'MENU_OPEN_ULB'):
        return HttpResponseForbidden("You do not have permission for this action.")

    me = request.user

    if me.role == 'ROOT_DEV':
        ulb_list = ULB.objects.all().order_by('ulb_name')
    else:
        ulb_list = ULB.objects.filter(
            permissions__user=me,
            permissions__code=UserPermission.ULB_ACCESS_CODE,
        ).distinct().order_by('ulb_name')

    if request.method == 'POST':
        selected_id = request.POST.get('selected_ulb_id')  # must match your form field name
        if selected_id:
            ulb = ULB.objects.get(id=selected_id)
            request.session['current_ulb_id'] = ulb.id
            request.session['current_ulb_name'] = ulb.ulb_name
            return redirect('base_accounts_dashboard')

    return render_dashboard(
        request,
        'create_ulb/open_ulb.html',
        {'ulb_list': ulb_list},
        active_section='ulb',
    )

#------------------------- ULB: view / edit / delete ------------------------------
@login_required
@role_required(['ROOT_DEV', 'DEV'])
def view_ulb(request, ulb_id=None):
    # enforce custom permission code
    if not user_has_code(request.user, 'MENU_VIEW_ULB'):
        return HttpResponseForbidden("You do not have permission for this action.")

    # All ULBs for the top dropdown
    ulb_list = ULB.objects.all().order_by('ulb_name')

    reset_after_load = False
    ulb = None

    # 1) Handle dropdown "View" selection
    if request.method == 'POST' and request.POST.get('action') == 'select':
        selected_id = request.POST.get('selected_ulb_id')
        if selected_id:
            # go to /view_ulb/<id>/ so fields show that ULB
            return redirect('view_ulb', ulb_id=int(selected_id))

    # 2) Resolve current ULB to show in form (only when URL has ulb_id)
    if ulb_id is not None:
        ulb = get_object_or_404(ULB, id=ulb_id)

    # 3) Handle edit / delete
    if request.method == 'POST' and request.POST.get('action') in ['edit', 'delete']:
        action = request.POST.get('action')

        if ulb is None:
            messages.error(request, 'No ULB selected to edit or delete.')
            return redirect('view_ulb')  # goes to blank form + --Select ULB--

        if action == 'edit':
            ulb_name = request.POST.get('ulb_name')
            ulb_type = request.POST.get('ulb_type')
            email = request.POST.get('email')
            pan_no = request.POST.get('pan_no') or None
            tin_no = request.POST.get('tin_no')
            gst_no = request.POST.get('gst_no')
            land_mark = request.POST.get('land_mark')
            district = request.POST.get('district')
            state = request.POST.get('state')
            country = request.POST.get('country')
            date_of_creation = request.POST.get('date_of_creation')
            ulb_code = request.POST.get('ulb_code')

            duplicate = ULB.objects.filter(
                ulb_name__iexact=ulb_name.strip(),
                email__iexact=email.strip(),
                pan_no=pan_no,
                tin_no=tin_no.strip(),
                gst_no=gst_no.strip(),
            ).exclude(id=ulb.id).exists()

            if duplicate:
                messages.error(
                    request,
                    'A ULB with the same ULB Name, Email ID, PAN No., '
                    'TIN No., and GST No. already exists.'
                )
                # stay on same ULB; do not clear
                return redirect('view_ulb', ulb_id=ulb.id)

            # check ULB code uniqueness (ignore if unchanged)
            if ulb_code and ULB.objects.filter(
                code__iexact=ulb_code.strip()
            ).exclude(id=ulb.id).exists():
                messages.error(
                    request,
                    'This ULB Code is already in use. Please choose a different code.'
                )
                return redirect('view_ulb', ulb_id=ulb.id)

            ulb.ulb_name = ulb_name
            ulb.ulb_type = ulb_type
            ulb.email = email
            ulb.pan_no = pan_no
            ulb.tin_no = tin_no
            ulb.gst_no = gst_no
            ulb.land_mark = land_mark
            ulb.district = district
            ulb.state = state
            ulb.country = country
            ulb.date_of_creation = date_of_creation
            ulb.code = ulb_code.strip() if ulb_code else None
            ulb.save()
            messages.success(request, 'ULB edited successfully.')
            # after success: redirect to plain /view_ulb/ so refresh also stays clear
            return redirect('view_ulb')

        elif action == 'delete':
            ulb.delete()
            messages.success(request, 'ULB deleted successfully.')
            # after delete: also go to /view_ulb/ for blank form
            return redirect('view_ulb')

    # 4) Render page
    # If ulb_id is None (plain /view_ulb/ or after redirect or browser refresh),
    # we always show empty fields and "-- Select ULB --".
    if ulb_id is None:
        ulb_ctx = None
        reset_after_load = True
    else:
        ulb_ctx = ulb
        reset_after_load = False  # when viewing a specific ULB, don't auto-clear

    return render_dashboard(
        request,
        'create_ulb/view_ulb.html',
        {
            'ulb': ulb_ctx,
            'ulb_list': ulb_list,
            'reset_after_load': reset_after_load,
        },
        active_section='ulb',
    )
# ------------------- ULB-wise report -------------------#
from django.db.models import Q
@login_required
@role_required(['ROOT_DEV', 'DEV', 'ADMIN'])
def ulb_wise_report_view(request):
    # Optional: protect with a menu code
    if not user_has_code(request.user, 'MENU_ULB_WISE_REPORT'):
        return HttpResponseForbidden("You do not have permission for this action.")

    me = request.user

    # base: ULBs that current user is allowed to see
    if me.role == 'ROOT_DEV':
        # ROOT_DEV: see ALL ULBs
        base_ulbs = ULB.objects.all()
    elif me.role == 'DEV':
        # DEV: only ULBs where this dev has ULB_ACCESS
        base_ulbs = UserPermission.ulbs_for_user(me)
    elif me.role == 'ADMIN':
        # ADMIN: ULBs this admin has access to
        base_ulbs = UserPermission.ulbs_for_user(me)
    else:
        base_ulbs = ULB.objects.none()

    # ---- search by ULB name (optional) ----
    q = request.GET.get('q', '').strip()
    if q:
        base_ulbs = base_ulbs.filter(ulb_name__icontains=q)

    base_ulbs = base_ulbs.order_by('ulb_name')

    rows = []

    for ulb in base_ulbs:
        # users that have ULB_ACCESS on this ULB
        perm_users = User.objects.filter(
            permissions__ulb=ulb,
            permissions__code=UserPermission.ULB_ACCESS_CODE,
        ).distinct()

        # For DEV login: only show that dev + dev's users
        if me.role == 'DEV':
            # dev itself, if it has access (it should)
            dev_users = perm_users.filter(id=me.id)

            # dev's admins/users on this ULB (role ADMIN/USER and also have access)
            child_users = perm_users.filter(
                role__in=['ADMIN', 'USER'],
                # later you can add created_by=me here
            ).order_by('role', 'username')

            # if dev has no access at all (shouldn't happen), mark unassigned
            if not dev_users.exists() and not child_users.exists():
                rows.append({
                    'ulb': ulb,
                    'dev': None,
                    'children': [],
                })
            else:
                rows.append({
                    'ulb': ulb,
                    'dev': me,
                    'children': list(child_users),
                })

        # For ROOT_DEV: show ALL users (any role) that have ULB_ACCESS on this ULB
        elif me.role == 'ROOT_DEV':
            all_users = perm_users.order_by('role', 'username')

            if not all_users.exists():
                rows.append({
                    'ulb': ulb,
                    'users': [],
                })
            else:
                rows.append({
                    'ulb': ulb,
                    'users': list(all_users),
                })

        # For ADMIN: show only that admin (no hierarchy)
        elif me.role == 'ADMIN':
            admin_users = perm_users.filter(id=me.id)
            if not admin_users.exists():
                rows.append({'ulb': ulb, 'dev': None, 'children': []})
            else:
                rows.append({'ulb': ulb, 'dev': me, 'children': []})

    return render_dashboard(
        request,
        'create_ulb/ulb_wise_report.html',
        {
            'rows': rows,
            'q': q,
        },
        active_section='ulb',  # or 'reports' depending on your sidebar
    )


# ------------------- Create user -------------------#

User = get_user_model()
from django.contrib.auth.decorators import login_required
from django.http import HttpResponseForbidden
from django.contrib import messages

@login_required
@role_required(['ROOT_DEV', 'DEV', 'ADMIN'])
def create_user_view(request):
    # make sure these roles actually have this permission code assigned
    if not user_has_code(request.user, 'MENU_CREATE_USERS'):
        return HttpResponseForbidden("You do not have permission for this action.")

    me = request.user
    my_role = getattr(me, 'role', None)

    # allowed roles for new user, based on creator role
    if my_role == 'ROOT_DEV':
        allowed_roles = ['ROOT_DEV', 'DEV', 'ADMIN', 'USER']
    elif my_role == 'DEV':
        allowed_roles = ['ADMIN', 'USER']
    elif my_role == 'ADMIN':
        allowed_roles = ['USER']
    else:
        allowed_roles = []

    if request.method == 'POST':
        form = RootCreateUserForm(request.POST)

        if form.is_valid():
            user_obj = form.save(commit=False)

            # extra safety: enforce allowed roles server side
            if user_obj.role not in allowed_roles:
                messages.error(request, "You are not allowed to assign this role.")
                return redirect('create_user')

            user_obj.save()
            messages.success(request, 'User created successfully.')
            return redirect('create_user')
    else:
        form = RootCreateUserForm()

    return render_dashboard(
        request,
        'user_management/create_user.html',
        {
            'form': form,
            'my_role': my_role,  # used in template for role <select> options
        },
        active_section='users',
    )

# ------------------- Manage users -------------------#
User = get_user_model()

from django.db.models import Q

@login_required
@role_required(['ROOT_DEV', 'DEV', 'ADMIN'])
def manage_users(request):
    if not user_has_code(request.user, 'MENU_MANAGE_USERS'):
        return HttpResponseForbidden("You do not have permission for this action.")

    me = request.user
    role = getattr(me, 'role', None)

    if role == 'ROOT_DEV':
        # ROOT_DEV sees only DEV, ADMIN, USER (no ROOT_DEV)
        qs = User.objects.filter(role__in=['DEV', 'ADMIN', 'USER'])
    elif role == 'DEV':
        # DEV sees only ADMIN, USER (no ROOT_DEV, no DEV)
        qs = User.objects.filter(role__in=['ADMIN', 'USER'])
    elif role == 'ADMIN':
        # ADMIN sees only USER
        qs = User.objects.filter(role='USER')
    else:
        qs = User.objects.none()

    # search by username or mobile or email
    q = request.GET.get('q', '').strip()
    if q:
        qs = qs.filter(
            Q(username__icontains=q) |
            Q(mobile_number__icontains=q) |
            Q(email__icontains=q)
        )

    # never show myself
    users = qs.exclude(id=me.id).order_by('username')

    return render_dashboard(
        request,
        'user_management/manage_users.html',
        {
            'users': users,
            'q': q,
        },
        active_section='users',
    )

# ------------------- Manage user action (edit/disable/delete) -------------------#
User = get_user_model()
@login_required
@role_required(['ROOT_DEV', 'DEV', 'ADMIN'])
@require_POST
def manage_user_action(request, user_id):
    # same permission as manage_users
    if not user_has_code(request.user, 'MENU_MANAGE_USERS'):
        return HttpResponseForbidden("You do not have permission for this action.")

    me = request.user
    my_role = getattr(me, 'role', None)

    user = get_object_or_404(User, id=user_id)
    action = request.POST.get('action')

    # prevent self-delete or self-disable
    if user == me and action in ['disable', 'delete']:
        messages.error(request, "You cannot modify your own status this way.")
        return redirect('manage_users')

    # ----- enforce hierarchy on target user -----
    # ROOT_DEV can modify DEV/ADMIN/USER
    # DEV can modify ADMIN/USER
    # ADMIN can modify USER
    if my_role == 'ROOT_DEV':
        allowed_target_roles = ['DEV', 'ADMIN', 'USER']
        allowed_assign_roles = ['DEV', 'ADMIN', 'USER']
    elif my_role == 'DEV':
        allowed_target_roles = ['ADMIN', 'USER']
        allowed_assign_roles = ['ADMIN', 'USER']
    elif my_role == 'ADMIN':
        allowed_target_roles = ['USER']
        allowed_assign_roles = ['USER']
    else:
        allowed_target_roles = []
        allowed_assign_roles = []

    # target must be in allowed roles
    if user.role not in allowed_target_roles:
        messages.error(request, "You are not allowed to modify this user.")
        return redirect('manage_users')

    if action == 'save':
        # update editable fields
        user.username = request.POST.get('username') or user.username
        user.first_name = request.POST.get('first_name') or ''
        user.last_name = request.POST.get('last_name') or ''
        user.email = request.POST.get('email') or ''
        user.mobile_number = request.POST.get('mobile_number') or ''

        # role change must also respect hierarchy
        new_role = request.POST.get('role') or user.role
        if new_role not in allowed_assign_roles:
            messages.error(request, "You are not allowed to assign this role.")
            return redirect('manage_users')
        user.role = new_role

        # password change (keep session if editing self)
        new_password = request.POST.get('new_password') or ''
        if new_password:
            user.set_password(new_password)

        user.save()

        if new_password and user == me:
            # keep current user logged in after password change
            update_session_auth_hash(request, user)

        messages.success(request, f"User {user.username} updated.")

    elif action == 'disable':
        user.is_active = False
        user.save()
        messages.success(request, f"User {user.username} disabled.")

    elif action == 'activate':
        user.is_active = True
        user.save()
        messages.success(request, f"User {user.username} activated.")

    elif action == 'delete':
        username = user.username
        user.delete()
        messages.success(request, f"User {username} deleted.")

    # Always go back to the manage users list
    return redirect('manage_users')

# ------------------- Manage access -------------------#
@login_required
@role_required(['ROOT_DEV', 'DEV', 'ADMIN'])
def manage_access(request):
    # enforce custom permission code
    if not user_has_code(request.user, 'MENU_MANAGE_ACCESS'):
        return HttpResponseForbidden("You do not have permission for this action.")

    me = request.user
    role = getattr(me, 'role', None)

    # ----- filter users dropdown by current role -----
    if role == 'ROOT_DEV':
        # ROOT_DEV can see all roles, including ROOT_DEV itself
        users = User.objects.filter(
            role__in=['ROOT_DEV', 'DEV', 'ADMIN', 'USER']
        ).order_by('username')
    elif role == 'DEV':
        users = User.objects.filter(role__in=['ADMIN', 'USER']).order_by('username')
    elif role == 'ADMIN':
        users = User.objects.filter(role='USER').order_by('username')
    else:
        users = User.objects.none()

    action = request.POST.get('action') if request.method == 'POST' else None

    selected_user_id = (
        request.POST.get('user') if request.method == 'POST' else request.GET.get('user')
    )
    selected_user = get_object_or_404(User, id=selected_user_id) if selected_user_id else None

    # Defaults
    available_ulbs = []
    assigned_ulbs = []
    selected_perm_ulb = None
    perm_available = []
    perm_assigned = []

    # ---------- STEP 1: search user -> show ULB boxes ----------
    if selected_user:
        # ULBs the current operator (me) is allowed to manage
        if role == 'ROOT_DEV':
            my_ulbs_qs = ULB.objects.all()
        else:
            my_ulbs_qs = ULB.objects.filter(
                permissions__user=me,
                permissions__code='ULB_ACCESS'
            ).distinct()

        # Assigned ULBs for selected_user, within my_ulbs_qs
        assigned_ulbs_qs = my_ulbs_qs.filter(
            permissions__user=selected_user,
            permissions__code='ULB_ACCESS'
        ).distinct().order_by('ulb_name')
        assigned_ulbs = list(assigned_ulbs_qs)

        if action == 'save_ulbs':
            # Read new assigned ULB ids from right box
            new_ulb_ids = request.POST.getlist('assigned_ulbs')

            # Remove existing ULB_ACCESS for this user, but only for ULBs I can manage
            UserPermission.objects.filter(
                user=selected_user,
                code='ULB_ACCESS',
                ulb__in=my_ulbs_qs
            ).delete()

            # Create for new set
            bulk_ulb = [
                UserPermission(user=selected_user, ulb_id=ulb_id, code='ULB_ACCESS')
                for ulb_id in new_ulb_ids
            ]
            UserPermission.objects.bulk_create(bulk_ulb)

            messages.success(request, f"ULB access updated for {selected_user.username}.")

            # recompute assigned ulbs
            assigned_ulbs_qs = my_ulbs_qs.filter(
                permissions__user=selected_user,
                permissions__code='ULB_ACCESS'
            ).distinct().order_by('ulb_name')
            assigned_ulbs = list(assigned_ulbs_qs)

        # Available ULBs = my_ulbs_qs minus assigned_ulbs
        if assigned_ulbs:
            available_ulbs = list(
                my_ulbs_qs.exclude(id__in=[u.id for u in assigned_ulbs]).order_by('ulb_name')
            )
        else:
            available_ulbs = list(my_ulbs_qs.order_by('ulb_name'))

    # ---------- STEP 2: permissions per ULB ----------
    if selected_user and assigned_ulbs:
        # ULB selected in step 2
        perm_ulb_id = (
            request.POST.get('perm_ulb') if request.method == 'POST' else request.GET.get('perm_ulb')
        )
        if perm_ulb_id:
            selected_perm_ulb = get_object_or_404(ULB, id=perm_ulb_id)

        if action == 'save_permissions' and selected_perm_ulb:
            new_codes = request.POST.getlist('assigned_permissions')

            parent_children = {
                'MENU_VIEW_ULB': ['BTN_EDIT_ULB', 'BTN_DELETE_ULB'],
                'MENU_MANAGE_USERS': ['BTN_EDIT_USER', 'BTN_DELETE_USER', 'BTN_TOGGLE_ACTIVE_USER'],
                'MENU_MANAGE_ACCESS': ['BTN_SEARCH_ACCESS', 'BTN_SAVE_PERMISSIONS_ACCESS', 'SELECT_ALL_ACCESS'],
                'MENU_ULB_HOME': ['BTN_ULB_RECEIPTS', 'BTN_ULB_PAYMENT', 'BTN_ULB_CONTRA', 'BTN_ULB_JOURNAL'],
            }

            for parent, children in parent_children.items():
                for child in children:
                    if child in new_codes and parent not in new_codes:
                        new_codes.append(parent)

            # Remove old non-ULB_ACCESS codes for this user+ulb
            UserPermission.objects.filter(
                user=selected_user,
                ulb=selected_perm_ulb
            ).exclude(code='ULB_ACCESS').delete()

            bulk = [
                UserPermission(user=selected_user, ulb=selected_perm_ulb, code=code)
                for code in new_codes
            ]
            UserPermission.objects.bulk_create(bulk)

            messages.success(
                request,
                f"Menu/button permissions updated for {selected_user.username} in {selected_perm_ulb.ulb_name}."
            )

        if selected_perm_ulb:
            assigned_codes = set(
                UserPermission.objects.filter(
                    user=selected_user,
                    ulb=selected_perm_ulb
                ).exclude(code='ULB_ACCESS').values_list('code', flat=True)
            )
            perm_available = [
                (code, label) for code, label in ALL_PERMISSIONS if code not in assigned_codes
            ]
            perm_assigned = [
                (code, label) for code, label in ALL_PERMISSIONS if code in assigned_codes
            ]

    context = {
        'users': users,
        'selected_user': selected_user,
        'available_ulbs': available_ulbs,
        'assigned_ulbs': assigned_ulbs,
        'selected_perm_ulb': selected_perm_ulb,
        'perm_available': perm_available,
        'perm_assigned': perm_assigned,
    }
    return render_dashboard(
        request,
        'user_management/manage_access.html',
        context,
        active_section='users',
    )

# ------------------- User-wise report -------------------#
from django.db.models import Prefetch
@login_required
@role_required(['ROOT_DEV', 'DEV', 'ADMIN'])
def user_wise_report_view(request):
    if not user_has_code(request.user, 'MENU_USER_WISE_REPORT'):
        return HttpResponseForbidden("You do not have permission for this action.")

    me = request.user

    # ---------- ROOT_DEV ----------
    if me.role == 'ROOT_DEV':
        base_users = User.objects.filter(role__in=['DEV', 'ADMIN', 'USER'])
        qs_users = base_users.exclude(id=me.id)

        qs_users = qs_users.prefetch_related(
            Prefetch(
                'permissions',
                queryset=UserPermission.objects.filter(
                    code=UserPermission.ULB_ACCESS_CODE
                ).select_related('ulb'),
                to_attr='ulb_access_perms'
            )
        ).order_by('username')

        rows = []
        for u in qs_users:
            ulbs = [perm.ulb for perm in getattr(u, 'ulb_access_perms', [])]
            
            rows.append({'user': u, 'ulbs': ulbs})

    # ---------- DEV ----------
    elif me.role == 'DEV':
        my_ulbs_qs = UserPermission.ulbs_for_user(me)

        if not my_ulbs_qs.exists():
            rows = []
        else:
            my_ulb_ids = set(my_ulbs_qs.values_list('id', flat=True))

            qs_users = (
                User.objects
                .filter(role__in=['ADMIN', 'USER'])
                .exclude(id=me.id)
                .prefetch_related(
                    Prefetch(
                        'permissions',
                        queryset=UserPermission.objects.filter(
                            code=UserPermission.ULB_ACCESS_CODE
                        ).select_related('ulb'),
                        to_attr='ulb_access_perms'
                    )
                )
                .order_by('username')
            )

            rows = []
            for u in qs_users:
                user_ulbs = [perm.ulb for perm in getattr(u, 'ulb_access_perms', [])]
                shared_ulbs = [ulb for ulb in user_ulbs if ulb.id in my_ulb_ids]
                if shared_ulbs:
                    rows.append({'user': u, 'ulbs': shared_ulbs})

    # ---------- ADMIN ----------
    elif me.role == 'ADMIN':
        my_ulbs_qs = UserPermission.ulbs_for_user(me)

        if not my_ulbs_qs.exists():
            rows = []
        else:
            my_ulb_ids = set(my_ulbs_qs.values_list('id', flat=True))

            qs_users = (
                User.objects
                .filter(role='USER')
                .exclude(id=me.id)
                .prefetch_related(
                    Prefetch(
                        'permissions',
                        queryset=UserPermission.objects.filter(
                            code=UserPermission.ULB_ACCESS_CODE
                        ).select_related('ulb'),
                        to_attr='ulb_access_perms'
                    )
                )
                .order_by('username')
            )

            rows = []
            for u in qs_users:
                user_ulbs = [perm.ulb for perm in getattr(u, 'ulb_access_perms', [])]
                shared_ulbs = [ulb for ulb in user_ulbs if ulb.id in my_ulb_ids]
                if shared_ulbs:
                    rows.append({'user': u, 'ulbs': shared_ulbs})
    else:
        rows = []

    # -------- search filter on rows --------
    q = request.GET.get('q', '').strip()
    if q:
        q_lower = q.lower()
        filtered = []
        for row in rows:
            u = row['user']
            if (
                q_lower in (u.username or '').lower()
                or q_lower in (u.role or '').lower()
                or q_lower in (u.mobile_number or '').lower()
                or q_lower in (u.email or '').lower()
            ):
                filtered.append(row)
        rows = filtered

    return render_dashboard(
        request,
        'user_management/user_wise_report.html',
        {
            'rows': rows,
            'q': q,
        },
        active_section='users',
    )

# ---------- ULB home + accounting views ----------




# ---- helpers ----

def get_allowed_codes_for(user):
    """Return a set/list of permission codes for sidebar + checks."""
    if not user.is_authenticated:
        return set()
    return set(
        UserPermission.objects
        .filter(user=user)
        .values_list("code", flat=True)
    )


def user_has_code(user, code: str) -> bool:
    if not user.is_authenticated:
        return False
    return UserPermission.objects.filter(user=user, code=code).exists()


def render_accounts(request, template_name, context=None, active_section=None):
    """
    Small helper similar to your render_dashboard, but for accounts.
    """
    context = context or {}

    # inject allowed_codes for sidebar
    context.setdefault("allowed_codes", get_allowed_codes_for(request.user))

    if active_section is not None:
        context["active_section"] = active_section

    return render(request, template_name, context)


# ---- views ----

@login_required
@role_required(['ROOT_DEV', 'DEV', 'ADMIN'])
def base_accounts_dashboard_view(request):
    current_ulb_name = request.session.get('current_ulb_name', None)

    context = {
        'current_ulb_name': current_ulb_name,
        'allowed_codes': get_allowed_codes_for(request.user),
        'active_section': 'dashboard',
    }
    return render(request, 'accounts/base_accounts_dashboard.html', context)


# NEW: master landing view that shows only header tabs
@login_required
@role_required(['ROOT_DEV', 'DEV', 'ADMIN',])
def accounts_master_tabs(request):
    if not user_has_code(request.user, 'MENU_ACCOUNTS_MASTER'):
        return HttpResponseForbidden("You do not have permission for this action.")

    context = {
        'active_section': 'master',
        'current_ulb_name': request.session.get('current_ulb_name'),
        'allowed_codes': get_allowed_codes_for(request.user),
        'active_master_tab': None,   # no tab active yet
    }
    return render(request, 'accounts/master_tabs.html', context)


@login_required
@role_required(['ROOT_DEV', 'DEV', 'ADMIN',])
def accounts_master_home(request):
    """
    Accounts -> Master tab (Create Ledger & Group screen).
    """

    # enforce permission: cannot open via URL if no access
    if not user_has_code(request.user, 'MENU_ACCOUNTS_MASTER'):
        return HttpResponseForbidden("You do not have permission for this action.")

    # ----- Current ULB resolution -----
    current_ulb_id = request.session.get('current_ulb_id')
    current_ulb = None

    if current_ulb_id:
        current_ulb = get_object_or_404(ULB, id=current_ulb_id)
    else:
        current_ulb_name = request.session.get('current_ulb_name')
        if current_ulb_name:
            current_ulb = ULB.objects.filter(ulb_name=current_ulb_name).first()

    # ----- Handle POST: create Ledger + groups -----
    if request.method == 'POST':
        if not current_ulb:
            messages.error(request, "No ULB selected; cannot create ledger.")
            return redirect('accounts_master_home')  # use your url name

        ledger_name = request.POST.get('ledger_name', '').strip()

        # group chain: group-1, group-2, group-3, ... from the form
        group_names = []
        g1 = request.POST.get('group-1', '').strip()
        g2 = request.POST.get('group-2', '').strip()
        g3 = request.POST.get('group-3', '').strip()
        # add more lines for group-4, group-5 if your form has them
        if g1:
            group_names.append(g1)
        if g2:
            group_names.append(g2)
        if g3:
            group_names.append(g3)

        subgroup_name = request.POST.get('subgroup_name', '').strip()
        main_group_name = request.POST.get('main_group_name', '').strip()
        head_group = request.POST.get('head_group')  # "1","2","3","4" or None
        opening_date = request.POST.get('opening_date') or None
        opening_type = request.POST.get('opening_type') or None
        opening_balance = request.POST.get('opening_balance') or None

        # basic validation
        if not ledger_name:
            messages.error(request, "Ledger name is required.")
            return redirect('accounts_master_home')

        # ----- build/get LedgerGroup chain (REVERSED ORDER) -----
        from .models import LedgerGroup  # ensure imported at top too

        parent_group = None
        deepest_group = None

        # Reverse group_names to save from top to bottom hierarchy
        for name in reversed(group_names):
            group_obj, _ = LedgerGroup.objects.get_or_create(
                name=name,
                parent=parent_group,
            )
            parent_group = group_obj
            deepest_group = group_obj

        # ----- get/create subgroup / main_group -----
        subgroup_obj = None
        if subgroup_name:
            subgroup_obj, _ = SubGroup.objects.get_or_create(name=subgroup_name)

        main_group_obj = None
        if main_group_name:
            main_group_obj, _ = MainGroup.objects.get_or_create(name=main_group_name)

        if opening_balance == '':
            opening_balance = None

        # ----- create Ledger -----
        led = Ledger.objects.create(
            ulb=current_ulb,
            name=ledger_name,
            group=deepest_group,   # deepest group (Group-1)
            subgroup=subgroup_obj,
            main_group=main_group_obj,
            head_group_code=head_group or None,
            opening_date=opening_date or None,
            opening_type=opening_type or None,
            opening_balance=opening_balance,
            created_by=request.user,
        )
        messages.success(request, f"Ledger '{led.name}' created successfully.")

        # after save, redirect to GET to avoid resubmission
        return redirect('accounts_master_home')

    # ----- GET: build hierarchies for template -----
    if current_ulb:
        ledgers_qs = (
            Ledger.objects
            .select_related('group', 'subgroup', 'main_group', 'ulb')
            .filter(ulb=current_ulb)
            .order_by('name')
        )
        ulb_created_date = current_ulb.date_of_creation
    else:
        ledgers_qs = Ledger.objects.none()
        ulb_created_date = timezone.now().date()

    ledger_hierarchies = []
    for led in ledgers_qs:
        # walk up LedgerGroup chain to build all group levels
        groups_chain = []
        grp = led.group
        while grp is not None:
            groups_chain.append(grp.name)
            grp = grp.parent
        groups_chain.reverse()  # so it becomes [group-1, group-2, group-3, ...]

        hierarchy = {
            "id": led.id,
            "label": led.name,
            "ledger_name": led.name,
            "groups": groups_chain,
            "subgroup": led.subgroup.name if led.subgroup else "",
            "main_group": led.main_group.name if led.main_group else "",
            "head_group": str(led.head_group_code) if led.head_group_code is not None else "",
            "opening_date": led.opening_date.isoformat() if led.opening_date else "",
            "opening_type": led.opening_type or "",
            "opening_balance": str(led.opening_balance) if led.opening_balance is not None else "",
        }
        ledger_hierarchies.append(hierarchy)

    # ensure ulb_created_date is ISO string for HTML min attr
    ulb_created_date_iso = ulb_created_date.isoformat() if ulb_created_date else ""

    context = {
        'active_section': 'master',
        'active_master_tab': 'create',
        'current_ulb_name': current_ulb.ulb_name if current_ulb else None,
        'ledger_hierarchies': ledger_hierarchies,
        'ulb_created_date': ulb_created_date_iso,
        'allowed_codes': get_allowed_codes_for(request.user),
    }
    return render(request, 'accounts/create_lg.html', context)

@login_required
@role_required(['ROOT_DEV', 'DEV', 'ADMIN',])
def accounts_master_view_edit(request):
    if not user_has_code(request.user, 'MENU_ACCOUNTS_MASTER_VIEW_EDIT'):
        return HttpResponseForbidden("You do not have permission for this action.")

    current_ulb_id = request.session.get('current_ulb_id')
    current_ulb = get_object_or_404(ULB, id=current_ulb_id) if current_ulb_id else None

    # base queryset for this ULB
    if current_ulb:
        ledgers_qs = (
            Ledger.objects
            .select_related('group', 'subgroup', 'main_group', 'ulb')
            .filter(ulb=current_ulb)
            .order_by('name')
        )
    else:
        ledgers_qs = Ledger.objects.none()

    if request.method == 'POST':
        action = request.POST.get('action')          # "save" or "delete"
        ledger_id = request.POST.get('ledger_id')    # hidden field
        ledger = get_object_or_404(Ledger, id=ledger_id, ulb=current_ulb)

        if action == 'delete':
            ledger_name = ledger.name
            ledger.delete()
            messages.error(request, f"Ledger '{ledger_name}' deleted successfully.")
            return redirect('accounts_master_view_edit')

        # save/edit
        ledger_name = request.POST.get('ledger_name', '').strip()
        g1 = request.POST.get('group-1', '').strip()
        g2 = request.POST.get('group-2', '').strip()
        g3 = request.POST.get('group-3', '').strip()
        subgroup_name = request.POST.get('subgroup_name', '').strip()
        main_group_name = request.POST.get('main_group_name', '').strip()
        head_group = request.POST.get('head_group') or None
        opening_date = request.POST.get('opening_date') or None
        opening_type = request.POST.get('opening_type') or None
        opening_balance = request.POST.get('opening_balance') or None

        # basic validation
        if not ledger_name:
            messages.error(request, "Ledger name is required.")
            return redirect('accounts_master_view_edit')

        from .models import LedgerGroup

        # group-1 = parent, group-2 = child, group-3 = deeper child
        group_names = [x for x in [g1, g2, g3] if x]

        # build chain in REVERSED order (same logic as accounts_master_home)
        parent_group = None
        deepest_group = None
        for name in reversed(group_names):
            group_obj, _ = LedgerGroup.objects.get_or_create(
                name=name,
                parent=parent_group,
            )
            parent_group = group_obj
            deepest_group = group_obj

        subgroup_obj = None
        if subgroup_name:
            subgroup_obj, _ = SubGroup.objects.get_or_create(name=subgroup_name)

        main_group_obj = None
        if main_group_name:
            main_group_obj, _ = MainGroup.objects.get_or_create(name=main_group_name)

        if opening_balance == '':
            opening_balance = None

        # Sirf Assets (4) ya Liabilities (3) ke liye hi opening fields save karenge
        if head_group not in ['3', '4']:
            opening_date = None
            opening_type = None
            opening_balance = None

        # update this Ledger in-place
        ledger.name = ledger_name
        ledger.group = deepest_group
        ledger.subgroup = subgroup_obj
        ledger.main_group = main_group_obj
        ledger.head_group_code = int(head_group) if head_group else None
        ledger.opening_date = opening_date
        ledger.opening_type = opening_type
        ledger.opening_balance = opening_balance
        ledger.save()

        messages.success(request, f"Ledger '{ledger.name}' updated successfully.")
        return redirect('accounts_master_view_edit')

    # GET: build hierarchies + list for select
    ledger_hierarchies = []
    for led in ledgers_qs:
        # walk up LedgerGroup chain from deepest -> parent -> ... -> root
        groups_chain = []
        grp = led.group
        while grp is not None:
            groups_chain.append(grp.name)
            grp = grp.parent
        # IMPORTANT: DO NOT reverse here; groups_chain[0] = deepest, 1 = parent, etc.
        # Your JS now reverses groups for display so Group-1 becomes parent.

        # expose explicit fields if ever needed in template
        group_1 = groups_chain[1] if len(groups_chain) >= 2 else (groups_chain[0] if len(groups_chain) >= 1 else "")
        group_2 = groups_chain[0] if len(groups_chain) >= 1 else ""
        group_3 = groups_chain[2] if len(groups_chain) >= 3 else ""

        hierarchy = {
            "id": led.id,
            "ledger_name": led.name,
            "group_1": group_1,
            "group_2": group_2,
            "group_3": group_3,
            # send raw chain as deepest->...->root; JS will flip for fields
            "groups": groups_chain,
            "subgroup": led.subgroup.name if led.subgroup else "",
            "main_group": led.main_group.name if led.main_group else "",
            "head_group": str(led.head_group_code) if led.head_group_code is not None else "",
            "opening_date": led.opening_date.isoformat() if led.opening_date else "",
            "opening_type": led.opening_type or "",
            "opening_balance": str(led.opening_balance) if led.opening_balance is not None else "",
        }
        ledger_hierarchies.append(hierarchy)

    context = {
        'active_section': 'master',
        'active_master_tab': 'view_edit',
        'current_ulb_name': current_ulb.ulb_name if current_ulb else None,
        'ledger_hierarchies': ledger_hierarchies,
        'allowed_codes': get_allowed_codes_for(request.user),
    }
    return render(request, 'accounts/view_edit_lg.html', context)

@login_required
@role_required(['ROOT_DEV', 'DEV', 'ADMIN', 'USER'])
def accounts_master_export(request):

    current_ulb = get_object_or_404(
        ULB,
        id=request.session.get('current_ulb_id')
    )

    ledgers = (
        Ledger.objects
        .select_related('group', 'subgroup', 'main_group')
        .filter(ulb=current_ulb)
        .order_by(
            'head_group_code',
            'main_group__name',
            'subgroup__name',
            'name'
        )
    )

    # -------- group chain --------
    def get_group_chain(group):
        """
        Returns:
        [Group-1, Group-2, ..., Group-n]
        """
        chain = []
        while group:
            chain.append(group.name)
            group = group.parent
        return chain

    rows = []
    max_depth = 0

    for led in ledgers:
        chain = get_group_chain(led.group) if led.group else []
        max_depth = max(max_depth, len(chain))
        rows.append((led, chain))

    wb = Workbook()
    ws = wb.active
    ws.title = "Ledger Master"

    # -------- headers --------
    headers = ["Head group", "Main group", "Subgroup"]
    for i in range(max_depth, 0, -1):
        headers.append(f"Group-{i}")
    headers.append("Ledger")
    ws.append(headers)

    # -------- styles --------
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1D4ED8", end_color="1D4ED8", fill_type="solid")
    header_alignment = Alignment(horizontal="left", vertical="center", wrap_text=False)
    body_alignment = Alignment(horizontal="left", vertical="center", wrap_text=False)
    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # style header row
    for col in range(1, len(headers) + 1):
        c = ws.cell(row=1, column=col)
        c.font = header_font
        c.fill = header_fill
        c.alignment = header_alignment
        c.border = border

    # -------- fill data --------
    for led, chain in rows:

        head = {
            1: "1 Income",
            2: "2 Expenses",
            3: "3 Liabilities",
            4: "4 Assets"
        }.get(led.head_group_code, "")

        row = [
            head,
            led.main_group.name if led.main_group else "",
            led.subgroup.name if led.subgroup else "",
        ]

        # chain = [Group-1, Group-2, ..., Group-n]
        group_map = {i + 1: name for i, name in enumerate(chain)}

        # Fill Group-n → Group-1 strictly by number
        for i in range(max_depth, 0, -1):
            row.append(group_map.get(i, ""))

        row.append(led.name)
        ws.append(row)

    max_row = ws.max_row

    # -------- merge helper --------
    def merge_column(col):
        if max_row < 3:
            return
        start = 2
        prev = ws.cell(row=2, column=col).value

        for r in range(3, max_row + 1):
            curr = ws.cell(row=r, column=col).value
            if curr != prev:
                if start < r - 1 and prev not in ("", None):
                    ws.merge_cells(start_row=start, start_column=col,
                                   end_row=r - 1, end_column=col)
                start = r
                prev = curr

        if start < max_row and prev not in ("", None):
            ws.merge_cells(start_row=start, start_column=col,
                           end_row=max_row, end_column=col)

    # merge all columns except last (Ledger)
    for col in range(1, len(headers)):
        merge_column(col)

    # style body
    for r in ws.iter_rows(min_row=2, max_row=max_row,
                          min_col=1, max_col=len(headers)):
        for cell in r:
            cell.alignment = body_alignment
            cell.border = border

    # -------- auto-fit column widths --------
    for col_idx in range(1, len(headers) + 1):
        col_letter = get_column_letter(col_idx)
        max_length = 0
        for row in range(1, max_row + 1):
            cell = ws[f"{col_letter}{row}"]
            if cell.value is not None:
                length = len(str(cell.value))
                if length > max_length:
                    max_length = length
        ws.column_dimensions[col_letter].width = max_length

    filename = f"ledger_master_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response

@login_required
@role_required(['ROOT_DEV', 'DEV', 'ADMIN'])
def accounts_master_import(request):
    current_ulb = get_object_or_404(ULB, id=request.session.get('current_ulb_id'))

    if request.method == "POST":
        upload = request.FILES.get("excel_file")
        if not upload:
            messages.error(request, "Please select an Excel file to upload.")
            return redirect("accounts_master_import")

        try:
            wb = load_workbook(upload, data_only=True)
            ws = wb.active

            # ---- validate headers ----
            headers = [cell.value for cell in ws[1] or []]
            if (
                len(headers) < 4
                or headers[0] != "Head group"
                or headers[1] != "Main group"
                or headers[2] != "Subgroup"
            ):
                messages.error(
                    request,
                    "Invalid Excel format. Please use exported Ledger & Group file.",
                )
                return redirect("accounts_master_import")

            # ["Head group","Main group","Subgroup","Group-n",...,"Group-1","Ledger"]
            ledger_col_index = len(headers) - 1
            group_start_index = 3
            group_end_index = ledger_col_index - 1
            group_col_count = group_end_index - group_start_index + 1

            # ---- fill-down trackers ----
            last_head_group_label = ""
            last_main_group_text = ""
            last_subgroup_text = ""
            last_group_names = [""] * group_col_count

            for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                if not any(row):
                    continue

                # ---- Head / Main / Subgroup with fill-down ----
                raw_head = row[0] or ""
                raw_main = row[1] or ""
                raw_sub = row[2] or ""

                main_changed = False
                subgroup_changed = False

                if raw_head:
                    last_head_group_label = str(raw_head).strip()

                if raw_main:
                    new_main = str(raw_main).strip()
                    if new_main != last_main_group_text:
                        main_changed = True
                    last_main_group_text = new_main

                if raw_sub:
                    new_sub = str(raw_sub).strip()
                    if new_sub != last_subgroup_text:
                        subgroup_changed = True
                    last_subgroup_text = new_sub

                # 🔥 RESET GROUP FILL-DOWN WHEN CONTEXT CHANGES
                if main_changed or subgroup_changed:
                    last_group_names = [""] * group_col_count

                head_group_label = last_head_group_label
                main_group_text = last_main_group_text
                subgroup_text = last_subgroup_text

                # ---- ledger name ----
                ledger_cell = row[ledger_col_index]
                ledger_name = str(ledger_cell).strip() if ledger_cell else ""

                # ---- collect group names (LEFT → RIGHT) ----
                group_names = []
                for idx in range(group_start_index, group_end_index + 1):
                    col_offset = idx - group_start_index
                    raw_val = row[idx]

                    cell_val = str(raw_val).strip() if raw_val is not None else ""
                    if cell_val:
                        last_group_names[col_offset] = cell_val
                        group_name = cell_val
                    else:
                        group_name = last_group_names[col_offset]

                    if group_name:
                        group_names.append(group_name)

                # ---- mandatory validation ----
                mandatory_missing = []
                if not head_group_label:
                    mandatory_missing.append("Head group")
                if not main_group_text:
                    mandatory_missing.append("Main group")
                if not subgroup_text:
                    mandatory_missing.append("Subgroup")
                # Group-1 is the last group in the row
                group1_name = last_group_names[-1] if last_group_names else ""
                if not group1_name:
                    mandatory_missing.append("Group-1")
                if not ledger_name:
                    mandatory_missing.append("Ledger")

                if mandatory_missing:
                    raise ValueError(f"Row {row_idx}: Missing mandatory field(s): {', '.join(mandatory_missing)}")

                # ---- head group code ----
                head_group_code = None
                if head_group_label:
                    try:
                        head_group_code = int(str(head_group_label).split()[0])
                    except Exception:
                        pass

                # ---- build LedgerGroup hierarchy (NO reverse) ----
                parent_group = None
                deepest_group = None

                for name in group_names:
                    group_obj, _ = LedgerGroup.objects.get_or_create(
                        name=name,
                        parent=parent_group,
                    )
                    parent_group = group_obj
                    deepest_group = group_obj

                # ---- get/create subgroup & main group ----
                subgroup_obj = None
                if subgroup_text:
                    subgroup_obj, _ = SubGroup.objects.get_or_create(
                        name=subgroup_text
                    )

                main_group_obj = None
                if main_group_text:
                    main_group_obj, _ = MainGroup.objects.get_or_create(
                        name=main_group_text
                    )

                # ---- upsert Ledger ----
                defaults = {
                    "head_group_code": head_group_code,
                    "group": deepest_group,
                    "subgroup": subgroup_obj,
                    "main_group": main_group_obj,
                }

                ledger_obj, created = Ledger.objects.get_or_create(
                    ulb=current_ulb,
                    name=ledger_name,
                    defaults=defaults,
                )

                updated = False

                if (
                    head_group_code is not None
                    and ledger_obj.head_group_code != head_group_code
                ):
                    ledger_obj.head_group_code = head_group_code
                    updated = True

                if deepest_group and ledger_obj.group_id != deepest_group.id:
                    ledger_obj.group = deepest_group
                    updated = True

                if subgroup_obj and ledger_obj.subgroup_id != subgroup_obj.id:
                    ledger_obj.subgroup = subgroup_obj
                    updated = True

                if main_group_obj and ledger_obj.main_group_id != main_group_obj.id:
                    ledger_obj.main_group = main_group_obj
                    updated = True

                if updated:
                    ledger_obj.save()

            messages.success(
                request,
                "Ledger & Group Excel imported successfully. "
                "Existing rows were updated, not duplicated.",
            )
            return redirect("accounts_master_import")

        except Exception as e:
            messages.error(request, f"Error while importing Excel: {e}")
            return redirect("accounts_master_import")

    # ---- GET ----
    context = {
        "current_ulb": current_ulb,
        "current_ulb_name": current_ulb.ulb_name if current_ulb else None,
        "allowed_codes": get_allowed_codes_for(request.user),
        "active_section": "master",
        "active_master_tab": "import",
    }
    return render(request, "accounts/import_lg.html", context)


#---- transaction entry view -------------------------------------
from datetime import date
from django.http import JsonResponse  # add this


def get_financial_year(date):
    """Get FY as YYYY for voucher sequence (April-March)"""
    year = date.year
    if date.month >= 4:  # April onwards
        return year
    return year - 1


def generate_voucher_no(ulb, voucher_type, voucher_date):
    """Generate DNP/RECEIPT/26/01/31/0001 format with separate sequence per type per FY"""
    fy = get_financial_year(voucher_date)

    # Sequence per ULB, voucher_type and financial year
    last_seq = Transaction.objects.filter(
        ulb=ulb,
        voucher_type=voucher_type,
        voucher_date__gte=date(fy, 4, 1),
        voucher_date__lte=date(fy + 1, 3, 31),
    ).aggregate(Max('sequence_no'))['sequence_no__max'] or 0

    next_seq = last_seq + 1

    # Format: DNP/RECEIPT/26/01/31/0001
    yy = voucher_date.strftime('%y')
    mm = voucher_date.strftime('%m')
    dd = voucher_date.strftime('%d')
    seq_str = f"{next_seq:04d}"

    return f"{ulb.code}/{voucher_type}/{yy}/{mm}/{dd}/{seq_str}", next_seq


@login_required
@role_required(['ROOT_DEV', 'DEV', 'ADMIN', 'USER'])
def transaction_entry(request):
    if not user_has_code(request.user, 'MENU_ACCOUNTS_TRANSACTION_ENTRY'):
        return HttpResponseForbidden("You do not have permission for this action.")

    current_ulb_id = request.session.get('current_ulb_id')
    current_ulb = get_object_or_404(ULB, id=current_ulb_id) if current_ulb_id else None
    
    if not current_ulb:
        messages.error(request, "Please select a ULB first.")
        return redirect('ulb_select')

        # All ledgers for this ULB (so every head/main/sub appears)
    ledgers = Ledger.objects.filter(ulb=current_ulb).select_related(
        "main_group", "subgroup", "group"
    )

    rows = []
    for ledger in ledgers:
        # Sum entries in date range for this ledger
        qs = TransactionEntry.objects.select_related("transaction", "ledger").filter(ledger=ledger)
        agg = qs.aggregate(
            dr_amount=Sum("dr_amount"),
            cr_amount=Sum("cr_amount"),
        )
        dr_amount = float(agg["dr_amount"] or 0)
        cr_amount = float(agg["cr_amount"] or 0)

        opening_amount = float(ledger.opening_balance or 0)
        opening_type = ledger.opening_type or ""  # "DR"/"CR"

        rows.append(
            {
                "ledger": ledger,
                "opening_amount": opening_amount,
                "opening_type": opening_type,
                "dr_amount": dr_amount,
                "cr_amount": cr_amount,
            }
        )


    sidebar_context = {
        "current_ulb": current_ulb,
        "current_ulb_name": current_ulb.ulb_name if current_ulb else None,
        "allowed_codes": get_allowed_codes_for(request.user),
        "active_section": "transactions",
    }

    if request.method == 'POST':
        with transaction.atomic():
            voucher_date = request.POST.get('voucher_date')
            voucher_type = request.POST.get('voucher_type')
            narration = request.POST.get('narration', '').strip()

            if not all([voucher_date, voucher_type]):
                messages.error(request, "Date and Voucher Type are required.")
                return render(
                    request,
                    'accounts/transaction_entry.html',
                    {
                        **sidebar_context,
                        "ledgers": ledgers,
                    },
                )

            voucher_date_obj = datetime.strptime(voucher_date, '%Y-%m-%d').date()

            voucher_no = request.POST.get('voucher_no', '').strip()
            if not voucher_no:
                messages.error(request, "Voucher number is missing.")
                return render(
                    request,
                    'accounts/transaction_entry.html',
                    {
                        **sidebar_context,
                        "ledgers": ledgers,
                    },
                )

            try:
                seq_str = voucher_no.rsplit('/', 1)[-1]
                sequence_no = int(seq_str)
            except (ValueError, IndexError):
                messages.error(request, "Invalid voucher number format.")
                return render(
                    request,
                    'accounts/transaction_entry.html',
                    {
                        **sidebar_context,
                        "ledgers": ledgers,
                    },
                )

            if Transaction.objects.filter(voucher_no=voucher_no).exists():
                messages.error(request, f"Voucher number {voucher_no} already exists.")
                return render(
                    request,
                    'accounts/transaction_entry.html',
                    {
                        **sidebar_context,
                        "ledgers": ledgers,
                    },
                )

            # main Transaction row
            txn = Transaction.objects.create(
                ulb=current_ulb,
                voucher_type=voucher_type,
                voucher_date=voucher_date_obj,
                voucher_no=voucher_no,
                sequence_no=sequence_no,
                narration=narration,
            )

            # Accounting entries
            entry_count = 1
            while True:
                type_key = f"type_{entry_count}"
                ledger_key = f"ledger_{entry_count}"
                dr_key = f"dr_amount_{entry_count}"
                cr_key = f"cr_amount_{entry_count}"

                if type_key not in request.POST:
                    break

                entry_type = request.POST.get(type_key)
                ledger_id = request.POST.get(ledger_key)
                dr_amount = request.POST.get(dr_key, "0")
                cr_amount = request.POST.get(cr_key, "0")

                if ledger_id:
                    ledger_obj = get_object_or_404(
                        Ledger, id=ledger_id, ulb=current_ulb
                    )

                    TransactionEntry.objects.create(
                        transaction=txn,
                        entry_type=entry_type,
                        ledger=ledger_obj,
                        dr_amount=float(dr_amount.replace(',', '')) if dr_amount else 0,
                        cr_amount=float(cr_amount.replace(',', '')) if cr_amount else 0,
                    )

                entry_count += 1

            # Receipt specific fields
            if voucher_type == VoucherType.RECEIPT:
                uc_applicable = request.POST.get("uc_applicable") == "yes"
                ReceiptUCDetails.objects.create(
                    transaction=txn,
                    uc_applicable=uc_applicable,
                    major_head=request.POST.get("major_head", ""),
                    treasury_code=request.POST.get("treasury_code", ""),
                    uc_bill_no=request.POST.get("uc_bill_no", ""),
                    uc_bill_date=request.POST.get("uc_bill_date") or None,
                    sub_major_head=request.POST.get("sub_major_head", ""),
                    ddo_code=request.POST.get("ddo_code", ""),
                    letter_no=request.POST.get("letter_no", ""),
                    letter_date=request.POST.get("letter_date") or None,
                    minor_head=request.POST.get("minor_head", ""),
                    bank_code=request.POST.get("bank_code", ""),
                    tv_no=request.POST.get("tv_no", ""),
                    tv_date=request.POST.get("tv_date") or None,
                    sub_head=request.POST.get("sub_head", ""),
                    bill_code=request.POST.get("bill_code", ""),
                    grant_amount=float(request.POST.get("grant_amount", 0) or 0),
                )

            # Payment specific fields
            elif voucher_type == VoucherType.PAYMENT:
                PaymentVendorDetails.objects.create(
                    transaction=txn,
                    vendor_name=request.POST.get("vendor_name", ""),
                    vendor_amount=float(request.POST.get("vendor_amount", 0) or 0),
                    cheque_no=request.POST.get("cheque_no", ""),
                    gst_applicable=request.POST.get("gst_applicable") == "yes",
                    gst_no=request.POST.get("gst_no", ""),
                    gst_type=request.POST.get("gst_type", ""),
                    gst_rate=float(request.POST.get("gst_rate", 0))
                    if request.POST.get("gst_rate")
                    else None,
                    igst_amount=float(request.POST.get("igst_amount", 0) or 0),
                    cgst_amount=float(request.POST.get("cgst_amount", 0) or 0),
                    sgst_amount=float(request.POST.get("sgst_amount", 0) or 0),
                    tds_applicable=request.POST.get("tds_applicable") == "yes",
                    tds_pan_no=request.POST.get("tds_pan_no", ""),
                    tds_section=request.POST.get("tds_section", ""),
                    tds_nature=request.POST.get("tds_nature", ""),
                    tds_type=request.POST.get("tds_type", ""),
                    tds_rate=request.POST.get("tds_rate", ""),
                    tds_amount=float(request.POST.get("tds_amount", 0) or 0),
                )

            messages.success(request, f"Transaction {voucher_no} saved successfully.")
            return redirect("accounts_transaction_entry")

    # GET request
    return render(
        request,
        "accounts/transaction_entry.html",
        {
            **sidebar_context,
            "ledgers": ledgers,
        },
    )


@login_required
@role_required(['ROOT_DEV', 'DEV', 'ADMIN', 'USER'])
def get_next_voucher_no(request):
    """AJAX endpoint to get next voucher number for a given type and date"""
    current_ulb_id = request.session.get('current_ulb_id')
    current_ulb = get_object_or_404(ULB, id=current_ulb_id) if current_ulb_id else None

    voucher_type = request.GET.get('voucher_type')
    voucher_date = request.GET.get('voucher_date')

    if not (current_ulb and voucher_type and voucher_date):
        return JsonResponse({'error': 'Missing data'}, status=400)

    try:
        voucher_date_obj = datetime.strptime(voucher_date, '%Y-%m-%d').date()
    except ValueError:
        return JsonResponse({'error': 'Invalid date'}, status=400)

    voucher_no, sequence_no = generate_voucher_no(current_ulb, voucher_type, voucher_date_obj)
    return JsonResponse({'voucher_no': voucher_no, 'sequence_no': sequence_no})

#---- Trial Balance views -------------------------------------

def _get_tb_date_range(request):
    """
    Read from_date and to_date from GET.
    If not provided, default:
      from_date = 1-Apr of current FY (India style)
      to_date   = today
    """
    today = date.today()

    # Current FY starts on 1 April
    fy_start_year = today.year if today.month >= 4 else today.year - 1
    default_from = date(fy_start_year, 4, 1)
    default_to = today

    from_str = request.GET.get("from_date")
    to_str = request.GET.get("to_date")

    if from_str:
        from_date = datetime.strptime(from_str, "%Y-%m-%d").date()
    else:
        from_date = default_from

    if to_str:
        to_date = datetime.strptime(to_str, "%Y-%m-%d").date()
    else:
        to_date = default_to

    return from_date, to_date


def _signed_from_dc(amount, dc_type):
    """
    dc_type: "Dr"/"Cr".
    Dr = +, Cr = -
    """
    if not amount:
        return 0.0
    amount = float(amount)
    return -abs(amount) if dc_type == "Cr" else abs(amount)


def _get_ledgergroup_chain(group: LedgerGroup):
    """
    Given deepest LedgerGroup (ledger.group), return [root, ..., leaf].
    """
    if not group:
        return []
    chain = []
    node = group
    while node is not None:
        chain.append(node)
        node = node.parent
    chain.reverse()
    return chain


def _build_trial_balance_tree(ledger_rows):
    """
    ledger_rows: iterable of dicts:
      {
        "ledger": <Ledger>,
        "opening_amount": float,
        "opening_type": "DR"/"CR" or "",
        "dr_amount": float,
        "cr_amount": float,
      }

    Structure:
      Head Group (1 Income/Expenses/Liabilities/Assets)
        -> Main Group
          -> Sub Group
            -> LedgerGroup chain (root..leaf)
              -> Ledger
    """
    root = {}

    for row in ledger_rows:
        ledger: Ledger = row["ledger"]

        # ----- detect Suspense ledger -----
        ledger_name = (ledger.name or "").strip()
        name_l = ledger_name.lower()
        is_suspense = "suspense" in name_l

        # ----- Head group (from choices OR Suspense head) -----
        hg_code = getattr(ledger, "head_group_code", None)
        if is_suspense:
            head_name = "Suspense A/c"
            head_key = "HG_SUSPENSE"
        else:
            head_name = ledger.get_head_group_code_display() if hg_code else "Others"
            head_key = f"HG_{hg_code or 0}"

        # ----- Main group -----
        main_name = ledger.main_group.name if getattr(ledger, "main_group", None) else None

        # ----- Sub group -----
        sub_name = ledger.subgroup.name if getattr(ledger, "subgroup", None) else None

        # ----- LedgerGroup chain -----
        group_chain = _get_ledgergroup_chain(getattr(ledger, "group", None))

        opening_amount = row["opening_amount"] or 0
        opening_type_raw = (row["opening_type"] or "").upper()  # "DR"/"CR" or ""
        if opening_type_raw == "DR":
            opening_type_display = "Dr"
        elif opening_type_raw == "CR":
            opening_type_display = "Cr"
        else:
            opening_type_display = ""

        dr_amount = row["dr_amount"] or 0
        cr_amount = row["cr_amount"] or 0

        # signed values: Dr = +, Cr = -
        opening_signed = _signed_from_dc(opening_amount, opening_type_display)
        dr_signed = float(dr_amount)
        cr_signed = float(cr_amount)
        closing_signed = opening_signed + dr_signed - cr_signed

        # leaf ledger node
        leaf = {
            "id": f"L{ledger.id}",
            "parent_id": "",
            "name": ledger_name,
            "opening_amount": float(opening_amount),
            "opening_type": opening_type_display,
            "opening_signed": opening_signed,
            "dr_amount": float(dr_amount),
            "dr_signed": dr_signed,
            "cr_amount": float(cr_amount),
            "cr_signed": cr_signed,
            "closing_amount": abs(closing_signed),
            "closing_type": "Cr" if closing_signed < 0 else "Dr",
            "closing_signed": closing_signed,
            "children": [],
            "is_suspense": is_suspense,
        }

        # ----- traverse/create hierarchy nodes -----
        def get_or_create_node(container, key, display_name, prefix):
            if key not in container:
                node_id = f"{prefix}_{key}"
                container[key] = {
                    "id": node_id,
                    "parent_id": "",
                    "name": display_name,
                    "opening_amount": 0.0,
                    "opening_type": "",
                    "opening_signed": 0.0,
                    "dr_amount": 0.0,
                    "dr_signed": 0.0,
                    "cr_amount": 0.0,
                    "cr_signed": 0.0,
                    "closing_amount": 0.0,
                    "closing_type": "",
                    "closing_signed": 0.0,
                    "children": {},  # start as dict
                }
            return container[key]

        # Head
        head_node = get_or_create_node(root, head_key, head_name, "H")

        # Main group
        if main_name:
            main_key = f"MG_{main_name}"
            main_node = get_or_create_node(head_node["children"], main_key, main_name, "M")
            main_node["parent_id"] = head_node["id"]
        else:
            main_node = None

        # Sub group
        parent_for_sub = main_node or head_node
        if sub_name:
            sub_key = f"SG_{sub_name}"
            sub_node = get_or_create_node(parent_for_sub["children"], sub_key, sub_name, "S")
            sub_node["parent_id"] = parent_for_sub["id"]
        else:
            sub_node = None

        # LedgerGroup chain
        parent_for_groups = sub_node or main_node or head_node
        current_parent = parent_for_groups
        for idx, g in enumerate(group_chain):
            g_key = f"G{idx}_{g.id}"
            g_name = g.name
            current_parent = get_or_create_node(
                current_parent["children"], g_key, g_name, f"G{idx}"
            )
            current_parent["parent_id"] = (
                parent_for_groups["id"] if parent_for_groups else head_node["id"]
            )
            parent_for_groups = current_parent

        # Attach ledger to deepest parent
        if current_parent is None:
            current_parent = head_node
        if "children" not in current_parent:
            current_parent["children"] = {}
        leaf["parent_id"] = current_parent["id"]
        current_parent["children"][leaf["id"]] = leaf

        # roll up amounts
        def rollup(node):
            node["opening_signed"] += opening_signed
            node["dr_signed"] += dr_signed
            node["cr_signed"] += cr_signed
            node["closing_signed"] += closing_signed

        rollup(head_node)
        if main_node:
            rollup(main_node)
        if sub_node:
            rollup(sub_node)
        if (
            current_parent
            and current_parent is not head_node
            and current_parent is not main_node
            and current_parent is not sub_node
        ):
            rollup(current_parent)

    # finalize tree: compute display amounts + normalize children
    def finalize(node):
        node["opening_amount"] = abs(node["opening_signed"])
        if node["opening_signed"]:
            node["opening_type"] = "Cr" if node["opening_signed"] < 0 else "Dr"
        else:
            node["opening_type"] = ""

        node["dr_amount"] = node["dr_signed"]
        node["cr_amount"] = node["cr_signed"]

        node["closing_amount"] = abs(node["closing_signed"])
        if node["closing_signed"]:
            node["closing_type"] = "Cr" if node["closing_signed"] < 0 else "Dr"
        else:
            node["closing_type"] = ""

        children = node.get("children", [])

        # children may be dict (during build) or list (after first finalize)
        if isinstance(children, dict):
            children_iter = list(children.values())
        else:
            children_iter = list(children)

        for child in children_iter:
            finalize(child)

        node["children"] = children_iter

    tree = []
    for h in root.values():
        finalize(h)
        tree.append(h)

    # custom head order: 3 Liabilities, 4 Assets, 1 Income, 2 Expenses, Suspense A/c, Others
    def head_order_key(node):
        name = (node.get("name") or "").lower()
        if name.startswith("3 liabilities"):
            return 10
        if name.startswith("4 assets"):
            return 20
        if name.startswith("1 income"):
            return 30
        if name.startswith("2 expenses"):
            return 40
        if name.startswith("suspense a/c") or name.startswith("suspense"):
            return 50
        return 99

    tree.sort(key=head_order_key)
    return tree


def _compute_opening_suspense_row(current_ulb, rows):
    """
    From existing rows, compute total opening Dr and Cr.
    If there is a difference, return one extra row for Suspense A/c (not saved in DB),
    as its own head.
    Otherwise return None.
    """
    total_opening_dr = 0.0
    total_opening_cr = 0.0

    # sum opening for all real ledgers
    for r in rows:
        amt = float(r["opening_amount"] or 0)
        t = (r["opening_type"] or "").upper()
        if t == "DR":
            total_opening_dr += amt
        elif t == "CR":
            total_opening_cr += amt

    if abs(total_opening_dr - total_opening_cr) < 0.005:
        return None  # no difference, no suspense row

    diff = abs(total_opening_dr - total_opening_cr)

    # minimal fake ledger object only for tree display
    class FakeSuspenseLedger:
        def __init__(self, ulb):
            self.id = 0  # ID zero or negative so it won't clash
            self.ulb = ulb
            self.name = "Suspense A/c"
            self.main_group = None
            self.subgroup = None
            self.group = None
            # IMPORTANT: no head_group_code 1–4; own head instead
            self.head_group_code = None

        def get_head_group_code_display(self):
            return "Suspense A/c"

    fake_ledger = FakeSuspenseLedger(current_ulb)

    # if total Dr > total Cr, Suspense is a Cr balance; else Dr
    if total_opening_dr > total_opening_cr:
        opening_type = "CR"
    else:
        opening_type = "DR"

    suspense_row = {
        "ledger": fake_ledger,
        "opening_amount": diff,
        "opening_type": opening_type,
        "dr_amount": 0.0,
        "cr_amount": 0.0,
    }
    return suspense_row


@login_required
@role_required(['ROOT_DEV', 'DEV', 'ADMIN', 'USER'])
def trial_balance(request):
    if not user_has_code(request.user, "MENU_ACCOUNTS_TRIAL_BALANCE"):
        return HttpResponse("You do not have permission for this action.", status=403)

    current_ulb_id = request.session.get("current_ulb_id")
    current_ulb = get_object_or_404(ULB, id=current_ulb_id) if current_ulb_id else None

    if not current_ulb:
        messages.error(request, "Please select a ULB first.")
        return redirect("ulb_select")

    from_date, to_date = _get_tb_date_range(request)

    # base entries for this ULB
    base_entries = TransactionEntry.objects.select_related("transaction", "ledger").filter(
        transaction__ulb=current_ulb
    )

    # entries before from_date (for opening adjustment)
    entries_before = base_entries
    # entries within period (for movement)
    entries_period = base_entries

    if from_date:
        entries_before = entries_before.filter(transaction__voucher_date__lt=from_date)
        entries_period = entries_period.filter(transaction__voucher_date__gte=from_date)
    if to_date:
        entries_period = entries_period.filter(transaction__voucher_date__lte=to_date)

    # Aggregate per ledger BEFORE period (to adjust opening)
    before_data = (
        entries_before.values("ledger_id")
        .annotate(
            dr_amount=Sum("dr_amount"),
            cr_amount=Sum("cr_amount"),
        )
    )

    # Aggregate per ledger IN period (movement)
    period_data = (
        entries_period.values("ledger_id")
        .annotate(
            dr_amount=Sum("dr_amount"),
            cr_amount=Sum("cr_amount"),
        )
    )

    before_by_ledger = {row["ledger_id"]: row for row in before_data}
    period_by_ledger = {row["ledger_id"]: row for row in period_data}

    # Build rows for ledgers in this ULB
    rows = []
    for ledger in Ledger.objects.filter(ulb=current_ulb):
        before_agg = before_by_ledger.get(ledger.id)
        period_agg = period_by_ledger.get(ledger.id)

        before_dr = float(before_agg["dr_amount"] or 0) if before_agg else 0.0
        before_cr = float(before_agg["cr_amount"] or 0) if before_agg else 0.0
        period_dr = float(period_agg["dr_amount"] or 0) if period_agg else 0.0
        period_cr = float(period_agg["cr_amount"] or 0) if period_agg else 0.0

        # base opening from ledger
        opening_amount = float(ledger.opening_balance or 0)
        opening_type = (ledger.opening_type or "").upper()  # "DR"/"CR"

        # convert to signed, adjust with movement before from_date
        opening_signed = _signed_from_dc(
            opening_amount,
            "Dr" if opening_type == "DR" else "Cr" if opening_type == "CR" else "",
        )
        opening_signed += (before_dr - before_cr)

        # back to amount + type
        if opening_signed > 0:
            adj_opening_amount = abs(opening_signed)
            adj_opening_type = "DR"
        elif opening_signed < 0:
            adj_opening_amount = abs(opening_signed)
            adj_opening_type = "CR"
        else:
            adj_opening_amount = 0.0
            adj_opening_type = ""

        # SKIP ledgers that are completely zero (opening + period)
        if (
            abs(adj_opening_amount) < 0.005
            and abs(period_dr) < 0.005
            and abs(period_cr) < 0.005
        ):
            continue

        rows.append(
            {
                "ledger": ledger,
                "opening_amount": adj_opening_amount,
                "opening_type": adj_opening_type,
                "dr_amount": period_dr,  # only movement within period
                "cr_amount": period_cr,
            }
        )

    # ---------- auto Suspense A/c for opening difference ----------
    suspense_row = _compute_opening_suspense_row(current_ulb, rows)
    if suspense_row:
        rows.append(suspense_row)
    # --------------------------------------------------------------

    groups_tree = _build_trial_balance_tree(rows)

    sidebar_context = {
        "current_ulb": current_ulb,
        "current_ulb_name": current_ulb.ulb_name if current_ulb else None,
        "allowed_codes": get_allowed_codes_for(request.user),
        "active_section": "trial_balance",
    }

    return render(
        request,
        "accounts/trial_balance.html",
        {
            **sidebar_context,
            "groups_tree": groups_tree,
            "from_date": from_date,
            "to_date": to_date,
        },
    )

@login_required
@role_required(['ROOT_DEV', 'DEV', 'ADMIN', 'USER'])
def trial_balance_export_excel(request):
    # Use same logic as trial_balance view but write to Excel instead of HTML
    current_ulb_id = request.session.get("current_ulb_id")
    current_ulb = get_object_or_404(ULB, id=current_ulb_id) if current_ulb_id else None

    if not current_ulb:
        messages.error(request, "Please select a ULB first.")
        return redirect("ulb_select")

    from_date, to_date = _get_tb_date_range(request)

    entries = TransactionEntry.objects.select_related("transaction", "ledger").filter(
        transaction__ulb=current_ulb
    )
    if from_date:
        entries = entries.filter(transaction__voucher_date__gte=from_date)
    if to_date:
        entries = entries.filter(transaction__voucher_date__lte=to_date)

    ledger_data = (
        entries.values("ledger_id")
        .annotate(
            dr_amount=Sum("dr_amount"),
            cr_amount=Sum("cr_amount"),
        )
    )

    # same as trial_balance view
    movement_by_ledger = {row["ledger_id"]: row for row in ledger_data}

    rows = []
    for ledger in Ledger.objects.filter(ulb=current_ulb):
        agg = movement_by_ledger.get(ledger.id)

        dr_amount = float(agg["dr_amount"] or 0) if agg else 0.0
        cr_amount = float(agg["cr_amount"] or 0) if agg else 0.0

        opening_amount = float(ledger.opening_balance or 0)
        opening_type = (ledger.opening_type or "").upper()

        rows.append(
            {
                "ledger": ledger,
                "opening_amount": opening_amount,
                "opening_type": opening_type,
                "dr_amount": dr_amount,
                "cr_amount": cr_amount,
            }
        )

    # auto Suspense A/c row (same as TB)
    suspense_row = _compute_opening_suspense_row(current_ulb, rows)
    if suspense_row:
        rows.append(suspense_row)

    # build same tree as web page
    groups_tree = _build_trial_balance_tree(rows)

    # ---- flatten tree with level for indentation ----
    flat_rows = []

    def walk(node, level):
        flat_rows.append(
            {
                "level": level,
                "name": node["name"],
                "opening_amount": node["opening_amount"],
                "opening_type": node["opening_type"],
                "dr_amount": node["dr_amount"],
                "cr_amount": node["cr_amount"],
                "closing_amount": node["closing_amount"],
                "closing_type": node["closing_type"],
                "is_suspense": node.get("is_suspense", False),
                "type": node["id"].split("_", 1)[0],  # H, M, S, G0.., L..
            }
        )
        for child in node.get("children", []):
            walk(child, level + 1)

    for head in groups_tree:
        walk(head, 0)

    # ---- create Excel with styles ----
    import openpyxl
    from openpyxl.utils import get_column_letter
    from openpyxl.styles import Font, PatternFill, Alignment, numbers
    from io import BytesIO

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Trial Balance"

    # common styles
    title_font = Font(bold=True, size=14, color="FFFFFFFF")
    subtitle_font = Font(bold=True, size=12, color="FFFFFFFF")
    header_font = Font(bold=True, color="FFFFFFFF")
    header_fill = PatternFill(start_color="FF0F172A", end_color="FF0F172A", fill_type="solid")
    header_align = Alignment(horizontal="center")

    normal_font = Font(color="FFFFFFFF")
    num_align = Alignment(horizontal="right")
    text_align = Alignment(horizontal="left")

    # approximate background colors per level (similar to HTML)
    HEAD_FILL = PatternFill(start_color="FF1D4ED8", end_color="FF1D4ED8", fill_type="solid")
    MAIN_FILL = PatternFill(start_color="FF0F766E", end_color="FF0F766E", fill_type="solid")
    SUB_FILL = PatternFill(start_color="FF4C1D95", end_color="FF4C1D95", fill_type="solid")
    GROUP_FILL = PatternFill(start_color="FF374151", end_color="FF374151", fill_type="solid")
    LEDGER_FILL = PatternFill(start_color="FF111827", end_color="FF111827", fill_type="solid")

    # number format with thousands separator and 2 decimals
    amount_number_format = numbers.FORMAT_NUMBER_COMMA_SEPARATED1

    # ---- top merged rows ----
    ws.merge_cells("A1:E1")
    cell = ws["A1"]
    cell.value = "Trial Balance"
    cell.font = title_font
    cell.alignment = Alignment(horizontal="center")
    cell.fill = PatternFill(start_color="FF1F2937", end_color="FF1F2937", fill_type="solid")

    ws.merge_cells("A2:E2")
    cell = ws["A2"]
    cell.value = current_ulb.ulb_name if current_ulb else ""
    cell.font = subtitle_font
    cell.alignment = Alignment(horizontal="center")
    cell.fill = PatternFill(start_color="FF111827", end_color="FF111827", fill_type="solid")

    ws.merge_cells("A3:E3")
    cell = ws["A3"]
    if from_date and to_date:
        cell.value = f"From {from_date.strftime('%d-%m-%Y')} To {to_date.strftime('%d-%m-%Y')}"
    elif from_date:
        cell.value = f"From {from_date.strftime('%d-%m-%Y')}"
    elif to_date:
        cell.value = f"Up to {to_date.strftime('%d-%m-%Y')}"
    else:
        cell.value = "All Dates"
    cell.font = subtitle_font
    cell.alignment = Alignment(horizontal="center")
    cell.fill = PatternFill(start_color="FF020617", end_color="FF020617", fill_type="solid")

    # header row at row 4
    header_row_index = 4
    headers = [
        "Particulars",
        "Opening Balance",
        "Dr Amount",
        "Cr Amount",
        "Closing Balance",
    ]
    ws.append([""] * len(headers))  # ensure row 4 exists
    for col_idx, title in enumerate(headers, start=1):
        cell = ws.cell(row=header_row_index, column=col_idx)
        cell.value = title
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    # data starts at row 5
    excel_row = header_row_index + 1
    first_data_row = excel_row

    for r in flat_rows:
        level = r["level"]
        name = r["name"]
        node_type = r["type"]  # H, M, S, G0.., L..

        if node_type == "H":
            fill = HEAD_FILL
        elif node_type == "M":
            fill = MAIN_FILL
        elif node_type == "S":
            fill = SUB_FILL
        elif node_type.startswith("G"):
            fill = GROUP_FILL
        else:
            fill = LEDGER_FILL

        indent_spaces = " " * (level * 2)
        ws.cell(row=excel_row, column=1, value=f"{indent_spaces}{name}")

        # write None instead of 0 so cells show blank
        opening_val = r["opening_amount"] if r["opening_amount"] else None
        dr_val = r["dr_amount"] if r["dr_amount"] else None
        cr_val = r["cr_amount"] if r["cr_amount"] else None
        closing_val = r["closing_amount"] if r["closing_amount"] else None

        ws.cell(row=excel_row, column=2, value=opening_val)
        ws.cell(row=excel_row, column=3, value=dr_val)
        ws.cell(row=excel_row, column=4, value=cr_val)
        ws.cell(row=excel_row, column=5, value=closing_val)

        for col in range(1, 6):
            cell = ws.cell(row=excel_row, column=col)
            cell.font = normal_font
            cell.fill = fill
            if col == 1:
                cell.alignment = text_align
            else:
                cell.alignment = num_align
                cell.number_format = amount_number_format

        excel_row += 1

    # ---- Grand Total row ----
    total_row = excel_row
    total_font = Font(bold=True, color="FFFFFFFF")
    total_fill = PatternFill(start_color="FF111827", end_color="FF111827", fill_type="solid")

    ws.cell(row=total_row, column=1, value="Grand Total")
    last_data_row = excel_row - 1

    if last_data_row >= first_data_row:
        ws.cell(row=total_row, column=2, value=f"=SUM(B{first_data_row}:B{last_data_row})")
        ws.cell(row=total_row, column=3, value=f"=SUM(C{first_data_row}:C{last_data_row})")
        ws.cell(row=total_row, column=4, value=f"=SUM(D{first_data_row}:D{last_data_row})")
        ws.cell(row=total_row, column=5, value=f"=SUM(E{first_data_row}:E{last_data_row})")

    for col in range(1, 6):
        cell = ws.cell(row=total_row, column=col)
        cell.font = total_font
        cell.fill = total_fill
        if col == 1:
            cell.alignment = text_align
        else:
            cell.alignment = num_align
            cell.number_format = amount_number_format

    # autosize columns
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if cell.value is not None:
                    max_len = max(max_len, len(str(cell.value)))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = max_len + 2

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    filename = "trial_balance.xlsx"
    response = HttpResponse(
        output.read(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response

def _get_tb_date_range(request):
    """
    Same helper as Trial Balance: FY from 1-Apr, default to today.
    """
    today = date.today()
    fy_start_year = today.year if today.month >= 4 else today.year - 1
    default_from = date(fy_start_year, 4, 1)
    default_to = today

    from_str = request.GET.get("from_date")
    to_str = request.GET.get("to_date")

    if from_str:
        from_date = datetime.strptime(from_str, "%Y-%m-%d").date()
    else:
        from_date = default_from

    if to_str:
        to_date = datetime.strptime(to_str, "%Y-%m-%d").date()
    else:
        to_date = default_to

    return from_date, to_date

#----------Income Expenditure views -------------------------------------
def _get_tb_date_range(request):
    today = date.today()
    fy_start_year = today.year if today.month >= 4 else today.year - 1
    default_from = date(fy_start_year, 4, 1)
    default_to = today

    from_str = request.GET.get("from_date")
    to_str = request.GET.get("to_date")

    if from_str:
        from_date = datetime.strptime(from_str, "%Y-%m-%d").date()
    else:
        from_date = default_from

    if to_str:
        to_date = datetime.strptime(to_str, "%Y-%m-%d").date()
    else:
        to_date = default_to

    return from_date, to_date


def _get_ledgergroup_chain(group: LedgerGroup):
    """
    Given deepest LedgerGroup (ledger.group), return [root, ..., leaf].
    Same pattern as trial balance.
    """
    if not group:
        return []
    chain = []
    node = group
    while node is not None:
        chain.append(node)
        node = node.parent
    chain.reverse()
    return chain

# --------------- income expenditure views -------------------------------------
def get_income_expenditure_context(from_date, to_date, current_ulb):
    """
    Reuse the same logic as income_expenditure view,
    but return data instead of rendering template.
    """

    base_entries = TransactionEntry.objects.select_related("transaction", "ledger").filter(
        transaction__ulb=current_ulb
    )
    if from_date:
        base_entries = base_entries.filter(transaction__voucher_date__gte=from_date)
    if to_date:
        base_entries = base_entries.filter(transaction__voucher_date__lte=to_date)

    # only Income (1) and Expenditure (2)
    base_entries = base_entries.filter(ledger__head_group_code__in=[1, 2])

    period_data = (
        base_entries.values("ledger_id")
        .annotate(
            dr_amount=Sum("dr_amount"),
            cr_amount=Sum("cr_amount"),
        )
    )
    movement_by_ledger = {row["ledger_id"]: row for row in period_data}

    total_income = 0.0
    total_expense = 0.0

    income_root = {}
    expense_root = {}

    def get_or_create_node(container, key, display_name, hg_code, parent_id=""):
        if key not in container:
            node_id = key  # HG_*, MG_*, SG_*, G*_, L*
            container[key] = {
                "id": node_id,
                "parent_id": str(parent_id) if parent_id else "",
                "name": display_name,
                "head_group_code": hg_code,
                "main_group_name": None,
                "group_name": None,
                "subgroup_name": None,
                "income": 0.0,
                "expense": 0.0,
                "children": {},
            }
        return container[key]

    # Head -> Main Group -> Subgroup -> LedgerGroup chain -> Ledger
    for ledger in Ledger.objects.filter(
        ulb=current_ulb,
        head_group_code__in=[1, 2],
    ).select_related("main_group", "subgroup", "group"):

        agg = movement_by_ledger.get(ledger.id)
        dr = float(agg["dr_amount"] or 0) if agg else 0.0
        cr = float(agg["cr_amount"] or 0) if agg else 0.0
        net_signed = dr - cr  # Dr = +, Cr = -

        hg_code = ledger.head_group_code

        income_amount = 0.0
        expense_amount = 0.0

        if hg_code == 1:
            income_amount = -net_signed  # Cr - Dr
            total_income += income_amount
        elif hg_code == 2:
            expense_amount = net_signed  # Dr - Cr
            total_expense += expense_amount

        # skip pure-zero lines
        if abs(income_amount) < 0.005 and abs(expense_amount) < 0.005:
            continue

        root_container = income_root if hg_code == 1 else expense_root

        # 1) Head node
        head_name = ledger.get_head_group_code_display()
        head_key = f"HG_{hg_code}"
        head_node = get_or_create_node(
            container=root_container,
            key=head_key,
            display_name=head_name,
            hg_code=hg_code,
            parent_id="",
        )

        # 2) Main group
        main_obj = getattr(ledger, "main_group", None)
        if main_obj:
            main_key = f"MG_{main_obj.id}"
            main_node = get_or_create_node(
                container=head_node["children"],
                key=main_key,
                display_name=main_obj.name,
                hg_code=hg_code,
                parent_id=head_node["id"],
            )
            main_node["main_group_name"] = main_obj.name
        else:
            main_node = None

        # 3) Subgroup
        sub_obj = getattr(ledger, "subgroup", None)
        parent_for_sub = main_node or head_node
        if sub_obj:
            sub_key = f"SG_{sub_obj.id}"
            sub_node = get_or_create_node(
                container=parent_for_sub["children"],
                key=sub_key,
                display_name=sub_obj.name,
                hg_code=hg_code,
                parent_id=parent_for_sub["id"],
            )
            sub_node["subgroup_name"] = sub_obj.name
        else:
            sub_node = None

        # 4) LedgerGroup chain
        group_chain = _get_ledgergroup_chain(getattr(ledger, "group", None))

        parent_for_groups = sub_node or main_node or head_node
        current_parent = parent_for_groups
        for idx, g in enumerate(group_chain):
            g_key = f"G{idx}_{g.id}"
            g_name = g.name
            current_parent = get_or_create_node(
                container=current_parent["children"],
                key=g_key,
                display_name=g_name,
                hg_code=hg_code,
                parent_id=current_parent["id"],
            )
            current_parent["group_name"] = g.name

        # 5) Ledger leaf
        parent_node = current_parent or sub_node or main_node or head_node
        if "children" not in parent_node:
            parent_node["children"] = {}

        leaf_key = f"L{ledger.id}"
        leaf = {
            "id": leaf_key,
            "parent_id": parent_node["id"],
            "name": ledger.name,
            "head_group_code": hg_code,
            "main_group_name": main_obj.name if main_obj else None,
            "group_name": current_parent["name"] if current_parent else None,
            "subgroup_name": sub_obj.name if sub_obj else None,
            "income": income_amount,
            "expense": expense_amount,
            "children": {},
        }
        parent_node["children"][leaf_key] = leaf

        # roll up amounts
        def rollup_ie(node):
            node["income"] += income_amount
            node["expense"] += expense_amount

        rollup_ie(head_node)
        if main_node:
            rollup_ie(main_node)
        if sub_node:
            rollup_ie(sub_node)
        if current_parent and current_parent not in (head_node, main_node, sub_node):
            rollup_ie(current_parent)

    # finalize children dict -> list
    def finalize_ie(node):
        children = node.get("children", {})
        if isinstance(children, dict):
            children_iter = list(children.values())
        else:
            children_iter = list(children)
        for ch in children_iter:
            finalize_ie(ch)
        node["children"] = children_iter

    income_roots = []
    for n in income_root.values():
        finalize_ie(n)
        income_roots.append(n)

    expense_roots = []
    for n in expense_root.values():
        finalize_ie(n)
        expense_roots.append(n)

    surplus = total_income - total_expense
    deficit = total_expense - total_income

    return {
        "expense_roots": expense_roots,
        "income_roots": income_roots,
        "total_expense": total_expense,
        "total_income": total_income,
        "surplus": surplus,
        "deficit": deficit,
        "ulb_name": current_ulb.ulb_name if current_ulb else "",
        "from_date": from_date,
        "to_date": to_date,
    }


@login_required
@role_required(['ROOT_DEV', 'DEV', 'ADMIN', 'USER'])
def income_expenditure(request):
    if not user_has_code(request.user, "MENU_ACCOUNTS_INCOME_EXPENDITURE"):
        return HttpResponse("You do not have permission for this action.", status=403)

    current_ulb_id = request.session.get("current_ulb_id")
    current_ulb = get_object_or_404(ULB, id=current_ulb_id) if current_ulb_id else None
    if not current_ulb:
        messages.error(request, "Please select a ULB first.")
        return redirect("ulb_select")

    from_date, to_date = _get_tb_date_range(request)

    base_entries = TransactionEntry.objects.select_related("transaction", "ledger").filter(
        transaction__ulb=current_ulb
    )
    if from_date:
        base_entries = base_entries.filter(transaction__voucher_date__gte=from_date)
    if to_date:
        base_entries = base_entries.filter(transaction__voucher_date__lte=to_date)

    # only Income (1) and Expenditure (2)
    base_entries = base_entries.filter(ledger__head_group_code__in=[1, 2])

    period_data = (
        base_entries.values("ledger_id")
        .annotate(
            dr_amount=Sum("dr_amount"),
            cr_amount=Sum("cr_amount"),
        )
    )
    movement_by_ledger = {row["ledger_id"]: row for row in period_data}

    total_income = 0.0
    total_expense = 0.0

    # separate roots per side
    income_root = {}
    expense_root = {}

    def get_or_create_node(container, key, display_name, hg_code, parent_id=""):
        if key not in container:
            node_id = key  # key encodes level: HG_*, MG_*, SG_*, G*_, L*
            container[key] = {
                "id": node_id,
                "parent_id": str(parent_id) if parent_id else "",
                "name": display_name,
                "head_group_code": hg_code,
                "main_group_name": None,
                "group_name": None,
                "subgroup_name": None,
                "income": 0.0,
                "expense": 0.0,
                "children": {},
            }
        return container[key]

    # Head -> Main Group -> Subgroup -> LedgerGroup chain -> Ledger
    for ledger in Ledger.objects.filter(
        ulb=current_ulb,
        head_group_code__in=[1, 2],
    ).select_related("main_group", "subgroup", "group"):

        agg = movement_by_ledger.get(ledger.id)
        dr = float(agg["dr_amount"] or 0) if agg else 0.0
        cr = float(agg["cr_amount"] or 0) if agg else 0.0
        net_signed = dr - cr  # Dr = +, Cr = -

        hg_code = ledger.head_group_code

        income_amount = 0.0
        expense_amount = 0.0

        if hg_code == 1:
            income_amount = -net_signed  # Cr - Dr
            total_income += income_amount
        elif hg_code == 2:
            expense_amount = net_signed  # Dr - Cr
            total_expense += expense_amount

        # skip pure-zero lines
        if abs(income_amount) < 0.005 and abs(expense_amount) < 0.005:
            continue

        root_container = income_root if hg_code == 1 else expense_root

        # 1) Head node: e.g. "2 Expenses"
        head_name = ledger.get_head_group_code_display()
        head_key = f"HG_{hg_code}"
        head_node = get_or_create_node(
            container=root_container,
            key=head_key,
            display_name=head_name,
            hg_code=hg_code,
            parent_id="",  # head has no parent
        )

        # 2) Main group: 210 Establishment Expenses
        main_obj = getattr(ledger, "main_group", None)
        if main_obj:
            main_key = f"MG_{main_obj.id}"
            main_node = get_or_create_node(
                container=head_node["children"],
                key=main_key,
                display_name=main_obj.name,
                hg_code=hg_code,
                parent_id=head_node["id"],
            )
            main_node["main_group_name"] = main_obj.name
        else:
            main_node = None

        # 3) Subgroup: 21010 Salaries, Wages and bonus
        sub_obj = getattr(ledger, "subgroup", None)
        parent_for_sub = main_node or head_node
        if sub_obj:
            sub_key = f"SG_{sub_obj.id}"
            sub_node = get_or_create_node(
                container=parent_for_sub["children"],
                key=sub_key,
                display_name=sub_obj.name,
                hg_code=hg_code,
                parent_id=parent_for_sub["id"],
            )
            sub_node["subgroup_name"] = sub_obj.name
        else:
            sub_node = None

        # 4) LedgerGroup chain: 2101001 Salaries & Allowances
        group_chain = _get_ledgergroup_chain(getattr(ledger, "group", None))

        parent_for_groups = sub_node or main_node or head_node
        current_parent = parent_for_groups
        for idx, g in enumerate(group_chain):
            g_key = f"G{idx}_{g.id}"
            g_name = g.name
            current_parent = get_or_create_node(
                container=current_parent["children"],
                key=g_key,
                display_name=g_name,
                hg_code=hg_code,
                parent_id=current_parent["id"],
            )
            current_parent["group_name"] = g.name

        # 5) Ledger leaf: 210100101 / 210100103 etc.
        parent_node = current_parent or sub_node or main_node or head_node
        if "children" not in parent_node:
            parent_node["children"] = {}

        leaf_key = f"L{ledger.id}"
        leaf = {
            "id": leaf_key,
            "parent_id": parent_node["id"],
            "name": ledger.name,
            "head_group_code": hg_code,
            "main_group_name": main_obj.name if main_obj else None,
            "group_name": current_parent["name"] if current_parent else None,
            "subgroup_name": sub_obj.name if sub_obj else None,
            "income": income_amount,
            "expense": expense_amount,
            "children": {},
        }
        parent_node["children"][leaf_key] = leaf

        # roll up amounts
        def rollup_ie(node):
            node["income"] += income_amount
            node["expense"] += expense_amount

        rollup_ie(head_node)
        if main_node:
            rollup_ie(main_node)
        if sub_node:
            rollup_ie(sub_node)
        if current_parent and current_parent not in (head_node, main_node, sub_node):
            rollup_ie(current_parent)

    # finalize: children dict -> list (for template recursion)
    def finalize_ie(node):
        children = node.get("children", {})
        if isinstance(children, dict):
            children_iter = list(children.values())
        else:
            children_iter = list(children)
        for ch in children_iter:
            finalize_ie(ch)
        node["children"] = children_iter

    income_roots = []
    for n in income_root.values():
        finalize_ie(n)
        income_roots.append(n)

    expense_roots = []
    for n in expense_root.values():
        finalize_ie(n)
        expense_roots.append(n)

    surplus = total_income - total_expense
    deficit = total_expense - total_income
    surplus_abs = abs(surplus)

    sidebar_context = {
        "current_ulb": current_ulb,
        "current_ulb_name": current_ulb.ulb_name if current_ulb else None,
        "allowed_codes": get_allowed_codes_for(request.user),
        "active_section": "income_expenditure",
    }

    return render(
        request,
        "accounts/income_expenditure.html",
        {
            **sidebar_context,
            "from_date": from_date,
            "to_date": to_date,
            "income_roots": income_roots,
            "expense_roots": expense_roots,
            "total_income": total_income,
            "total_expense": total_expense,
            "surplus": surplus,
            "deficit": deficit,
            "surplus_abs": surplus_abs,
        },
    )

@login_required
@role_required(['ROOT_DEV', 'DEV', 'ADMIN', 'USER'])
def income_expenditure_export_excel(request):
    # 1) ULB from session
    current_ulb_id = request.session.get("current_ulb_id")
    current_ulb = get_object_or_404(ULB, id=current_ulb_id) if current_ulb_id else None

    if not current_ulb:
        messages.error(request, "Please select a ULB first.")
        return redirect("ulb_select")

    # 2) Date range (reuse TB helper)
    from_date, to_date = _get_tb_date_range(request)

    # 3) Use same context as HTML page
    ctx = get_income_expenditure_context(from_date, to_date, current_ulb)
    expense_roots = ctx["expense_roots"]
    income_roots = ctx["income_roots"]
    total_expense = ctx["total_expense"]
    total_income = ctx["total_income"]
    surplus = ctx["surplus"]
    deficit = ctx["deficit"]
    ulb_name = ctx["ulb_name"]

    # 4) Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Income & Expenditure"

    center_bold = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    header_font = Font(bold=True, color="FFFFFFFF")
    title_font = Font(bold=True, size=14, color="FFFFFFFF")
    yellow_font = Font(bold=True, color="00FCC602")
    normal_font = Font(color="FFFFFFFF")

    # Row fills by level (tune hex codes to match your CSS)
    HEAD_FILL = PatternFill(start_color="FF1D4ED8", end_color="FF1D4ED8", fill_type="solid")
    MAIN_FILL = PatternFill(start_color="FF0F766E", end_color="FF0F766E", fill_type="solid")
    SUB_FILL = PatternFill(start_color="FF4C1D95", end_color="FF4C1D95", fill_type="solid")
    GROUP_FILL = PatternFill(start_color="FF374151", end_color="FF374151", fill_type="solid")
    LEDGER_FILL = PatternFill(start_color="FF111827", end_color="FF111827", fill_type="solid")

    TITLE_FILL = PatternFill(start_color="FF1F2937", end_color="FF1F2937", fill_type="solid")
    SUBTITLE_FILL = PatternFill(start_color="FF111827", end_color="FF111827", fill_type="solid")
    DATE_FILL = PatternFill(start_color="FF020617", end_color="FF020617", fill_type="solid")
    HEADER_FILL = PatternFill(start_color="FF0F172A", end_color="FF0F172A", fill_type="solid")
    TOTAL_FILL = PatternFill(start_color="FF111827", end_color="FF111827", fill_type="solid")

    # Indian number format
    INDIAN_NUM_FMT = "#,##,##0.00"

    # 5) Title row (1)
    ws.merge_cells("A1:D1")
    c = ws["A1"]
    c.value = "Income & Expenditure"
    c.alignment = center_bold
    c.font = title_font
    c.fill = TITLE_FILL

    # 6) ULB row (2)
    ws.merge_cells("A2:D2")
    c = ws["A2"]
    c.value = ulb_name
    c.alignment = center_bold
    c.font = header_font
    c.fill = SUBTITLE_FILL

    # 7) Date row (3) – DD-MM-YYYY
    ws.merge_cells("A3:D3")
    c = ws["A3"]
    if from_date and to_date:
        c.value = f"From {from_date.strftime('%d-%m-%Y')} To {to_date.strftime('%d-%m-%Y')}"
    elif to_date:
        c.value = f"As on {to_date.strftime('%d-%m-%Y')}"
    else:
        c.value = ""
    c.alignment = center_bold
    c.font = header_font
    c.fill = DATE_FILL

    # 8) Header row (4), no blank row
    current_row = 4

    ws[f"A{current_row}"] = "Expenditure"
    ws[f"B{current_row}"] = "Amount"
    ws[f"C{current_row}"] = "Income"
    ws[f"D{current_row}"] = "Amount"

    for col in range(1, 5):
        cell = ws.cell(row=current_row, column=col)
        cell.alignment = center_bold
        cell.font = header_font
        cell.border = thin_border
        cell.fill = HEADER_FILL

    current_row += 1  # data starts at 5

    # 9) Flatten with type info (HG, MG, SG, G*, L*)
    def flatten_tree(nodes, side, depth=0):
        rows = []
        for n in nodes:
            amount = (n["expense"] if side == "expense" else n["income"]) or 0
            node_id = n.get("id", "")
            node_type = node_id.split("_", 1)[0] if "_" in node_id else node_id  # HG, MG, SG, G0.., L..
            rows.append((n["name"], amount, depth, node_type))
            children = n.get("children", [])
            if children:
                rows.extend(flatten_tree(children, side, depth + 1))
        return rows

    expense_rows = flatten_tree(expense_roots, "expense")
    income_rows = flatten_tree(income_roots, "income")
    max_len = max(len(expense_rows), len(income_rows)) if (expense_rows or income_rows) else 0

    # 10) Data rows with colors per level
    for i in range(max_len):
        exp_name = exp_amt = exp_type = ""
        inc_name = inc_amt = inc_type = ""

        if i < len(expense_rows):
            name, amt, depth, node_type = expense_rows[i]
            exp_name = (" " * (depth * 2)) + name
            exp_amt = amt
            exp_type = node_type

        if i < len(income_rows):
            name, amt, depth, node_type = income_rows[i]
            inc_name = (" " * (depth * 2)) + name
            inc_amt = amt
            inc_type = node_type

        ws[f"A{current_row}"] = exp_name
        ws[f"B{current_row}"] = exp_amt if exp_amt != "" else None
        ws[f"C{current_row}"] = inc_name
        ws[f"D{current_row}"] = inc_amt if inc_amt != "" else None

        # decide fill per side based on node_type
        def pick_fill(node_type):
            if node_type == "HG":
                return HEAD_FILL
            if node_type == "MG":
                return MAIN_FILL
            if node_type == "SG":
                return SUB_FILL
            if node_type.startswith("G"):
                return GROUP_FILL
            if node_type.startswith("L"):
                return LEDGER_FILL
            return LEDGER_FILL

        exp_fill = pick_fill(exp_type) if exp_name else LEDGER_FILL
        inc_fill = pick_fill(inc_type) if inc_name else LEDGER_FILL

        for col in range(1, 5):
            cell = ws.cell(row=current_row, column=col)
            cell.border = thin_border
            cell.font = normal_font
            if col in (1, 2):  # expenditure side
                cell.fill = exp_fill
            else:              # income side
                cell.fill = inc_fill

            if col in (2, 4):
                cell.alignment = Alignment(horizontal="right", vertical="center")
                cell.number_format = INDIAN_NUM_FMT
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center")

        current_row += 1

    # 11) Surplus/Deficit row (immediately after data, no blank)
    if surplus > 0 or deficit > 0:
        if surplus > 0:
            ws[f"A{current_row}"] = "Excess of Income over Expenditure"
            ws[f"B{current_row}"] = surplus
        else:
            ws[f"C{current_row}"] = "Excess of Expenditure over Income"
            ws[f"D{current_row}"] = deficit

        for col in range(1, 5):
            cell = ws.cell(row=current_row, column=col)
            cell.font = yellow_font
            cell.border = thin_border
            cell.fill = TOTAL_FILL
            if col in (2, 4):
                cell.alignment = Alignment(horizontal="right", vertical="center")
                cell.number_format = INDIAN_NUM_FMT
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center")

        current_row += 1  # totals come immediately after
    # else: if no surplus/deficit, totals come right after data

    # 12) Totals row (no blank above)
    ws[f"A{current_row}"] = "Total Expenditure"
    ws[f"B{current_row}"] = total_expense
    ws[f"C{current_row}"] = "Total Income"
    ws[f"D{current_row}"] = total_income

    for col in range(1, 5):
        cell = ws.cell(row=current_row, column=col)
        cell.font = yellow_font
        cell.border = thin_border
        cell.fill = TOTAL_FILL
        if col in (2, 4):
            cell.alignment = Alignment(horizontal="right", vertical="center")
            cell.number_format = INDIAN_NUM_FMT
        else:
            cell.alignment = center_bold

    # Column widths
    ws.column_dimensions["A"].width = 40
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 40
    ws.column_dimensions["D"].width = 18

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="income_expenditure.xlsx"'
    wb.save(response)
    return response

#------------------ Balance Sheet views -------------------------------------

def _get_tb_date_range(request):
    today = date.today()
    fy_start_year = today.year if today.month >= 4 else today.year - 1
    default_from = date(fy_start_year, 4, 1)
    default_to = today

    from_str = request.GET.get("from_date")
    to_str = request.GET.get("to_date")

    if from_str:
        from_date = datetime.strptime(from_str, "%Y-%m-%d").date()
    else:
        from_date = default_from

    if to_str:
        to_date = datetime.strptime(to_str, "%Y-%m-%d").date()
    else:
        to_date = default_to

    return from_date, to_date


def get_trial_balance_rows(current_ulb, from_date, to_date):
    base_entries = TransactionEntry.objects.select_related("transaction", "ledger").filter(
        transaction__ulb=current_ulb
    )

    entries_before = base_entries
    entries_period = base_entries

    if from_date:
        entries_before = entries_before.filter(transaction__voucher_date__lt=from_date)
        entries_period = entries_period.filter(transaction__voucher_date__gte=from_date)
    if to_date:
        entries_period = entries_period.filter(transaction__voucher_date__lte=to_date)

    before_data = (
        entries_before.values("ledger_id")
        .annotate(
            dr_amount=Sum("dr_amount"),
            cr_amount=Sum("cr_amount"),
        )
    )
    period_data = (
        entries_period.values("ledger_id")
        .annotate(
            dr_amount=Sum("dr_amount"),
            cr_amount=Sum("cr_amount"),
        )
    )

    before_by_ledger = {row["ledger_id"]: row for row in before_data}
    period_by_ledger = {row["ledger_id"]: row for row in period_data}

    # helper: convert amount + Dr/Cr + head group to natural signed balance
    def natural_signed(head_group_code, amount, dc_type):
        """
        head_group_code: 1 Income, 2 Expenses, 3 Liabilities, 4 Assets
        amount: absolute value (>0)
        dc_type: 'Dr' or 'Cr' (or '')
        """
        if amount == 0 or not dc_type:
            return 0.0

        # Liabilities: Cr positive, Dr negative
        if head_group_code == 3:
            if dc_type == "Cr":
                return amount
            else:
                return -amount

        # Assets: Dr positive, Cr negative
        if head_group_code == 4:
            if dc_type == "Dr":
                return amount
            else:
                return -amount

        # Income / Expenses: keep standard Dr+/Cr-
        if dc_type == "Dr":
            return amount
        else:
            return -amount

    rows = []
    for ledger in Ledger.objects.filter(ulb=current_ulb):
        before_agg = before_by_ledger.get(ledger.id)
        period_agg = period_by_ledger.get(ledger.id)

        before_dr = float(before_agg["dr_amount"] or 0) if before_agg else 0.0
        before_cr = float(before_agg["cr_amount"] or 0) if before_agg else 0.0
        period_dr = float(period_agg["dr_amount"] or 0) if period_agg else 0.0
        period_cr = float(period_agg["cr_amount"] or 0) if period_agg else 0.0

        opening_amount = float(ledger.opening_balance or 0)
        opening_type = (ledger.opening_type or "").upper()  # "DR"/"CR"

        # original signed opening (Dr+, Cr-)
        opening_signed = _signed_from_dc(
            opening_amount,
            "Dr" if opening_type == "DR" else "Cr" if opening_type == "CR" else "",
        )
        opening_signed += (before_dr - before_cr)

        if opening_signed > 0:
            adj_opening_amount = abs(opening_signed)
            adj_opening_type = "DR"
        elif opening_signed < 0:
            adj_opening_amount = abs(opening_signed)
            adj_opening_type = "CR"
        else:
            adj_opening_amount = 0.0
            adj_opening_type = ""

        if (
            abs(adj_opening_amount) < 0.005
            and abs(period_dr) < 0.005
            and abs(period_cr) < 0.005
        ):
            continue

        rows.append(
            {
                "ledger": ledger,
                "opening_amount": adj_opening_amount,
                "opening_type": adj_opening_type,
                "dr_amount": period_dr,
                "cr_amount": period_cr,
                # we do NOT set closing_signed here; that is done in _build_trial_balance_tree
            }
        )

    suspense_row = _compute_opening_suspense_row(current_ulb, rows)
    if suspense_row:
        rows.append(suspense_row)

    return rows


def get_balance_sheet_context(from_date, to_date, current_ulb):
    """
    Build Liabilities & Assets trees from TB closing balances.

    Suspense A/c is included based on its net Dr/Cr calculated from children:
      - Dr Suspense -> Assets side
      - Cr Suspense -> Liabilities side

    Difference based on absolute totals:
        diff = abs(total_asset_signed) - abs(total_liab_signed)

    Posting rule (no fetch from I&E):
      - if diff > 0  -> post into 310900000 Excess of income over Expenditure (surplus)
      - if diff < 0  -> post into 310910000 Excess of Expenditure over income (deficit)

    If the Municipal Fund hierarchy or the ledger is missing from the TB tree
    (because its balance is nil), it is created inside the Liabilities tree so the
    Balance Sheet grouping matches the master.
    """
    rows = get_trial_balance_rows(current_ulb, from_date, to_date)
    tb_tree = _build_trial_balance_tree(rows)

    liability_roots = []
    asset_roots = []

    total_liab_signed = 0.0
    total_asset_signed = 0.0

    suspense_head = None

    # 1) classify all heads except Suspense
    for head in tb_tree:
        name = (head.get("name") or "").lower()
        closing_signed = float(head.get("closing_signed") or 0.0)

        # store Suspense head but don’t decide side yet
        if name.startswith("suspense a/c") or name.startswith("suspense"):
            suspense_head = head
            continue

        if name.startswith("3 liabilities"):
            liability_roots.append(head)
            total_liab_signed += closing_signed
        elif name.startswith("4 assets"):
            asset_roots.append(head)
            total_asset_signed += closing_signed
        # 1 Income, 2 Expenses ignored

    # 2) decide Suspense side using sum of child closing_signed
    if suspense_head is not None:
        net_susp_signed = 0.0
        for child in suspense_head.get("children", []):
            net_susp_signed += float(child.get("closing_signed") or 0.0)

        if abs(net_susp_signed) >= 0.005:
            if net_susp_signed > 0:
                # Dr Suspense -> Assets
                asset_roots.append(suspense_head)
                total_asset_signed += net_susp_signed
            else:
                # Cr Suspense -> Liabilities
                liability_roots.append(suspense_head)
                total_liab_signed += net_susp_signed

    # 3) Compute difference using absolute totals
    asset_total_abs = abs(total_asset_signed)
    liab_total_abs = abs(total_liab_signed)
    diff = asset_total_abs - liab_total_abs

    # Helper: add amount to node totals, preserving sign
    def bump_amount(node, amount):
        current = float(node.get("closing_signed") or 0.0)
        new_val = current + amount
        node["closing_signed"] = new_val
        node["closing_amount"] = abs(new_val)
        node["closing_type"] = "Dr" if new_val > 0 else "Cr" if new_val < 0 else ""
        return current, new_val

    # Helper: find or create a group node by exact name under a parent node
    def ensure_group_child(parent_node, group_name):
        for ch in parent_node.get("children", []):
            if (ch.get("name") or "").strip() == group_name:
                return ch
        new_group = {
            "id": f"AUTO_G_{group_name}",
            "name": group_name,
            "children": [],
            "closing_amount": 0.0,
            "closing_signed": 0.0,
            "closing_type": "",
            "cr_amount": 0.0,
            "cr_signed": 0.0,
            "dr_amount": 0.0,
            "dr_signed": 0.0,
            "opening_amount": 0.0,
            "opening_signed": 0.0,
            "opening_type": "",
            "parent_id": parent_node.get("id") or "",
            "is_suspense": False,
        }
        parent_node.setdefault("children", []).append(new_group)
        return new_group

    # Helper: ensure a ledger node by code+name under a given group node
    def ensure_ledger_child(parent_node, ledger_name):
        for ch in parent_node.get("children", []):
            if (ch.get("name") or "").strip() == ledger_name:
                return ch
        new_ledger = {
            "id": f"AUTO_L_{ledger_name}",
            "name": ledger_name,
            "children": [],
            "closing_amount": 0.0,
            "closing_signed": 0.0,
            "closing_type": "",
            "cr_amount": 0.0,
            "cr_signed": 0.0,
            "dr_amount": 0.0,
            "dr_signed": 0.0,
            "opening_amount": 0.0,
            "opening_signed": 0.0,
            "opening_type": "",
            "parent_id": parent_node.get("id") or "",
            "is_suspense": False,
        }
        parent_node.setdefault("children", []).append(new_ledger)
        return new_ledger

    # Helper: get Liabilities root node ("3 Liabilities")
    def get_liabilities_root(liability_roots_):
        for root in liability_roots_:
            if (root.get("name") or "").strip().startswith("3 Liabilities"):
                return root
        return liability_roots_[0] if liability_roots_ else None

    # 4) Post diff into surplus/deficit ledger under correct Municipal Fund chain
    if abs(diff) >= 0.005:
        liab_root = get_liabilities_root(liability_roots)

        if liab_root is not None:
            mf_name = "310 Municipal (General) Fund"

            if diff > 0:
                # Surplus: must be CREDIT in Municipal Fund (liability)
                level2_name = "31090 Excess of income over Expenditure"
                level3_name = "3109000 Excess of income over Expenditure"
                ledger_name = "310900000 Excess of income over Expenditure"
                post_amount = -abs(diff)  # force Cr sign
            else:
                # Deficit: must be DEBIT in Municipal Fund (liability)
                level2_name = "31091 Excess of Expenditure over income"
                level3_name = "3109100 Excess of Expenditure over income"
                ledger_name = "310910000 Excess of Expenditure over income"
                post_amount = abs(diff)   # force Dr sign

            # build / ensure hierarchy
            mf_node = ensure_group_child(liab_root, mf_name)
            level2_node = ensure_group_child(mf_node, level2_name)
            level3_node = ensure_group_child(level2_node, level3_name)
            target = ensure_ledger_child(level3_node, ledger_name)

            # push post_amount into all levels: ledger, level3, level2, Municipal Fund, 3 Liabilities
            _, new_val = bump_amount(target, post_amount)
            if new_val > 0:
                target["dr_amount"] = abs(new_val)
                target["dr_signed"] = new_val
                target["cr_amount"] = 0.0
                target["cr_signed"] = 0.0
            elif new_val < 0:
                target["cr_amount"] = abs(new_val)
                target["cr_signed"] = new_val
                target["dr_amount"] = 0.0
                target["dr_signed"] = 0.0

            bump_amount(level3_node, post_amount)
            bump_amount(level2_node, post_amount)
            bump_amount(mf_node, post_amount)
            bump_amount(liab_root, post_amount)

            # Update liabilities total so Balance Sheet tallies
            total_liab_signed += post_amount

            # Ensure Municipal Fund group is first under 3 Liabilities
            children = liab_root.get("children", [])
            mf_index = next(
                (i for i, ch in enumerate(children)
                 if (ch.get("name") or "").strip() == mf_name),
                None,
            )
            if mf_index is not None and mf_index != 0:
                mf_node_move = children.pop(mf_index)
                children.insert(0, mf_node_move)
                liab_root["children"] = children

    # 5) Total types from sign (TB convention)
    liab_type = "Cr" if total_liab_signed < 0 else "Dr"
    asset_type = "Dr" if total_asset_signed > 0 else "Cr"

    return {
        "liability_roots": liability_roots,
        "asset_roots": asset_roots,
        "total_liability_signed": total_liab_signed,
        "total_liability_type": liab_type,
        "total_asset_signed": total_asset_signed,
        "total_asset_type": asset_type,
        "ulb_name": current_ulb.ulb_name if current_ulb else "",
        "from_date": from_date,
        "to_date": to_date,
    }


@login_required
@role_required(['ROOT_DEV', 'DEV', 'ADMIN', 'USER'])
def balance_sheet(request):
    if not user_has_code(request.user, 'MENU_ACCOUNTS_BALANCE_SHEET'):
        return HttpResponseForbidden("You do not have permission for this action.")

    current_ulb_id = request.session.get("current_ulb_id")
    current_ulb = get_object_or_404(ULB, id=current_ulb_id) if current_ulb_id else None
    if not current_ulb:
        messages.error(request, "Please select a ULB first.")
        return redirect("ulb_select")

    from_date, to_date = _get_tb_date_range(request)

    ctx = get_balance_sheet_context(from_date, to_date, current_ulb)

    sidebar_context = {
        "current_ulb": current_ulb,
        "current_ulb_name": current_ulb.ulb_name if current_ulb else None,
        "allowed_codes": get_allowed_codes_for(request.user),
        "active_section": "balance_sheet",
    }

    return render(
        request,
        "accounts/balance_sheet.html",
        {
            **sidebar_context,
            "from_date": ctx["from_date"],
            "to_date": ctx["to_date"],
            "liability_roots": ctx["liability_roots"],
            "asset_roots": ctx["asset_roots"],
            "total_liability_signed": ctx["total_liability_signed"],
            "total_liability_type": ctx["total_liability_type"],
            "total_asset_signed": ctx["total_asset_signed"],
            "total_asset_type": ctx["total_asset_type"],
        },
    )

@login_required
@role_required(['ROOT_DEV', 'DEV', 'ADMIN', 'USER'])
def balance_sheet_export_excel(request):
    current_ulb_id = request.session.get("current_ulb_id")
    current_ulb = get_object_or_404(ULB, id=current_ulb_id) if current_ulb_id else None

    if not current_ulb:
        messages.error(request, "Please select a ULB first.")
        return redirect("ulb_select")

    from_date, to_date = _get_tb_date_range(request)

    ctx = get_balance_sheet_context(from_date, to_date, current_ulb)
    liability_roots = ctx["liability_roots"]
    asset_roots = ctx["asset_roots"]
    total_liability = ctx["total_liability_signed"]
    total_asset = ctx["total_asset_signed"]
    ulb_name = ctx["ulb_name"]

    wb = Workbook()
    ws = wb.active
    ws.title = "Balance Sheet"

    center_bold = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    header_font = Font(bold=True, color="FFFFFFFF")
    title_font = Font(bold=True, size=14, color="FFFFFFFF")
    yellow_font = Font(bold=True, color="00FCC602")
    normal_font = Font(color="FFFFFFFF")

    HEAD_FILL = PatternFill(start_color="FF1D4ED8", end_color="FF1D4ED8", fill_type="solid")
    MAIN_FILL = PatternFill(start_color="FF0F766E", end_color="FF0F766E", fill_type="solid")
    SUB_FILL = PatternFill(start_color="FF4C1D95", end_color="FF4C1D95", fill_type="solid")
    GROUP_FILL = PatternFill(start_color="FF374151", end_color="FF374151", fill_type="solid")
    LEDGER_FILL = PatternFill(start_color="FF111827", end_color="FF111827", fill_type="solid")

    TITLE_FILL = PatternFill(start_color="FF1F2937", end_color="FF1F2937", fill_type="solid")
    SUBTITLE_FILL = PatternFill(start_color="FF111827", end_color="FF111827", fill_type="solid")
    DATE_FILL = PatternFill(start_color="FF020617", end_color="FF020617", fill_type="solid")
    HEADER_FILL = PatternFill(start_color="FF0F172A", end_color="FF0F172A", fill_type="solid")
    TOTAL_FILL = PatternFill(start_color="FF111827", end_color="FF111827", fill_type="solid")

    INDIAN_NUM_FMT = "#,##,##0.00"

    # Title
    ws.merge_cells("A1:D1")
    c = ws["A1"]
    c.value = "Balance Sheet"
    c.alignment = center_bold
    c.font = title_font
    c.fill = TITLE_FILL

    # ULB
    ws.merge_cells("A2:D2")
    c = ws["A2"]
    c.value = ulb_name
    c.alignment = center_bold
    c.font = header_font
    c.fill = SUBTITLE_FILL

    # Date
    ws.merge_cells("A3:D3")
    c = ws["A3"]
    if from_date and to_date:
        c.value = f"From {from_date.strftime('%d-%m-%Y')} To {to_date.strftime('%d-%m-%Y')}"
    elif to_date:
        c.value = f"As on {to_date.strftime('%d-%m-%Y')}"
    else:
        c.value = ""
    c.alignment = center_bold
    c.font = header_font
    c.fill = DATE_FILL

    # Header row
    current_row = 4
    ws[f"A{current_row}"] = "Liabilities"
    ws[f"B{current_row}"] = "Amount"
    ws[f"C{current_row}"] = "Assets"
    ws[f"D{current_row}"] = "Amount"

    for col in range(1, 5):
        cell = ws.cell(row=current_row, column=col)
        cell.alignment = center_bold
        cell.font = header_font
        cell.border = thin_border
        cell.fill = HEADER_FILL

    current_row += 1

    # flatten including depth so we can color by level like the web page
    def flatten_bs(nodes, depth=0):
        rows = []
        for n in nodes:
            signed = n.get("closing_signed") or 0.0
            rows.append((n["name"], signed, depth))
            children = n.get("children", [])
            if children:
                rows.extend(flatten_bs(children, depth + 1))
        return rows

    liab_rows = flatten_bs(liability_roots)
    asset_rows = flatten_bs(asset_roots)
    max_len = max(len(liab_rows), len(asset_rows)) if (liab_rows or asset_rows) else 0

    # map level to fill, roughly matching HTML levels
    def pick_fill_by_depth(depth):
        if depth == 0:
            return HEAD_FILL   # main head (3 Liabilities / 4 Assets)
        if depth == 1:
            return MAIN_FILL
        if depth == 2:
            return SUB_FILL
        if depth == 3:
            return GROUP_FILL
        if depth >= 4:
            return LEDGER_FILL
        return LEDGER_FILL

    red_font = Font(color="FFFF0000")

    for i in range(max_len):
        liab_name = ""
        liab_signed = 0.0
        liab_depth = 0
        asset_name = ""
        asset_signed = 0.0
        asset_depth = 0

        if i < len(liab_rows):
            name, signed, depth = liab_rows[i]
            liab_name = (" " * (depth * 2)) + name
            liab_signed = float(signed or 0.0)
            liab_depth = depth

        if i < len(asset_rows):
            name, signed, depth = asset_rows[i]
            asset_name = (" " * (depth * 2)) + name
            asset_signed = float(signed or 0.0)
            asset_depth = depth

        # Liabilities:
        # - normal Cr stored negative -> show positive
        # - abnormal Dr stored positive -> show in brackets
        if liab_name:
            if liab_signed > 0:
                liab_display = f"({abs(liab_signed):,.2f})"
            elif abs(liab_signed) < 0.005:
                liab_display = None
            else:
                liab_display = abs(liab_signed)
        else:
            liab_display = None

        # Assets:
        # - normal Dr stored positive -> show positive
        # - abnormal Cr stored negative -> brackets
        if asset_name:
            if asset_signed < 0:
                asset_display = f"({abs(asset_signed):,.2f})"
            elif abs(asset_signed) < 0.005:
                asset_display = None
            else:
                asset_display = asset_signed
        else:
            asset_display = None

        ws[f"A{current_row}"] = liab_name
        ws[f"B{current_row}"] = liab_display
        ws[f"C{current_row}"] = asset_name
        ws[f"D{current_row}"] = asset_display

        liab_fill = pick_fill_by_depth(liab_depth) if liab_name else LEDGER_FILL
        asset_fill = pick_fill_by_depth(asset_depth) if asset_name else LEDGER_FILL

        for col in range(1, 5):
            cell = ws.cell(row=current_row, column=col)
            cell.border = thin_border
            cell.font = normal_font
            if col in (1, 2):
                cell.fill = liab_fill
            else:
                cell.fill = asset_fill

            if col in (2, 4):
                cell.alignment = Alignment(horizontal="right", vertical="center")
                # only apply numeric format when value is numeric
                if isinstance(cell.value, (int, float)):
                    cell.number_format = INDIAN_NUM_FMT
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center")

        # red font for bracketed negatives
        b_val = ws[f"B{current_row}"].value
        d_val = ws[f"D{current_row}"].value
        if isinstance(b_val, str) and b_val.startswith("("):
            ws[f"B{current_row}"].font = red_font
        if isinstance(d_val, str) and d_val.startswith("("):
            ws[f"D{current_row}"].font = red_font

        current_row += 1

    # totals: show positive figures
    ws[f"A{current_row}"] = "Total Liabilities"
    ws[f"B{current_row}"] = abs(total_liability)
    ws[f"C{current_row}"] = "Total Assets"
    ws[f"D{current_row}"] = abs(total_asset)

    for col in range(1, 5):
        cell = ws.cell(row=current_row, column=col)
        cell.font = yellow_font
        cell.border = thin_border
        cell.fill = TOTAL_FILL
        if col in (2, 4):
            cell.alignment = Alignment(horizontal="right", vertical="center")
            cell.number_format = INDIAN_NUM_FMT
        else:
            cell.alignment = center_bold

    # Auto-fit columns A–D based on content, a bit wider for amount columns
    for col_letter in ["A", "B", "C", "D"]:
        max_len = 0
        for cell in ws[col_letter]:
            value = cell.value
            if value is None:
                continue
            text = str(value)
            if len(text) > max_len:
                max_len = len(text)
        extra = 4 if col_letter in ("B", "D") else 2
        ws.column_dimensions[col_letter].width = max_len + extra

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="balance_sheet.xlsx"'
    wb.save(response)
    return response

@login_required
@role_required(['ROOT_DEV', 'DEV', 'ADMIN', 'USER'])
def day_book(request):
    if not user_has_code(request.user, 'MENU_ACCOUNTS_DAY_BOOK'):
        return HttpResponseForbidden("You do not have permission for this action.")

    # --- DELETE via query param ?delete=ID ---
    delete_id = request.GET.get("delete")
    if delete_id:
        try:
            t = Transaction.objects.get(id=delete_id, ulb_id=request.session.get("current_ulb_id"))
            t.delete()
        except Transaction.DoesNotExist:
            pass
        return redirect("accounts_day_book")

    # current ULB from session
    current_ulb_id = request.session.get("current_ulb_id")
    current_ulb = get_object_or_404(ULB, id=current_ulb_id) if current_ulb_id else None

    if not current_ulb:
        return redirect("ulb_select")

    # base queryset: transactions for this ULB
    qs = (
        Transaction.objects
        .filter(ulb=current_ulb)
        .prefetch_related("entries__ledger")
    )

    # --- Date filters with default Current F.Y. (From 01-04-YYYY To today) ---
    today = date.today()
    fy_start_year = today.year if today.month >= 4 else today.year - 1
    default_from = date(fy_start_year, 4, 1)
    default_to = today

    from_date_str = request.GET.get("from_date") or ""
    to_date_str = request.GET.get("to_date") or ""

    from_date = default_from
    to_date = default_to

    if from_date_str:
        try:
            from_date = datetime.strptime(from_date_str, "%Y-%m-%d").date()
        except ValueError:
            from_date = default_from

    if to_date_str:
        try:
            to_date = datetime.strptime(to_date_str, "%Y-%m-%d").date()
        except ValueError:
            to_date = default_to

    qs = qs.filter(voucher_date__gte=from_date, voucher_date__lte=to_date)

    # --- Text / amount filters ---
    amount_str = (request.GET.get("amount") or "").strip()
    voucher_no = (request.GET.get("voucher_no") or "").strip()
    cheque_no = (request.GET.get("cheque_no") or "").strip()

    if voucher_no:
        qs = qs.filter(voucher_no__icontains=voucher_no)

    # cheque_no lives on PaymentVendorDetails for payments
    if cheque_no:
        qs = qs.filter(paymentvendordetails__cheque_no__icontains=cheque_no)

    # Aggregate Dr / Cr totals from TransactionEntry (entries)
    qs = qs.annotate(
        dr_total=Sum("entries__dr_amount"),
        cr_total=Sum("entries__cr_amount"),
    )

    if amount_str:
        try:
            amt = float(amount_str)
            qs = qs.filter(Q(dr_total=amt) | Q(cr_total=amt))
        except ValueError:
            pass

    # Build context entries for template
    entries = []
    for t in qs.order_by("voucher_date", "id"):
        entry_lines = list(t.entries.all().select_related("ledger"))

        # per-line rows so each ledger appears once, with amount in correct column
        line_rows = []
        for le in entry_lines:
            is_debit = (le.entry_type == "Dr")
            line_rows.append(
                {
                    "ledger_name": le.ledger.name,
                    "is_debit": is_debit,
                    "dr_amount": float(le.dr_amount or 0),
                    "cr_amount": float(le.cr_amount or 0),
                    "entry_type": le.entry_type,
                    "id": le.id,
                }
            )

        # cheque number (only for payments with PaymentVendorDetails)
        pv = getattr(t, "paymentvendordetails", None)
        cheque_number = pv.cheque_no if pv else ""

        entries.append(
            {
                "id": t.id,
                "entry_date": t.voucher_date,
                "voucher_number": t.voucher_no,
                "voucher_type": t.voucher_type,  # RECV/PYMT/CNTR/JRNL
                "cheque_number": cheque_number,
                "narration": t.narration,
                "line_rows": line_rows,
                "dr_amount": float(t.dr_total or 0),
                "cr_amount": float(t.cr_total or 0),
            }
        )

    sidebar_context = {
        "current_ulb": current_ulb,
        "current_ulb_name": current_ulb.ulb_name if current_ulb else None,
        "allowed_codes": get_allowed_codes_for(request.user),
        "active_section": "day_book",
    }

    context = {
        **sidebar_context,
        "entries": entries,
        "from_date": from_date,
        "to_date": to_date,
    }
    return render(request, "accounts/day_book.html", context)


# ---------- HELPER for voucher number generation (used in transaction_edit_entry view) ----------
def get_financial_year(d):
    """Get FY as YYYY for voucher sequence (April-March)"""
    year = d.year
    if d.month >= 4:  # April onwards
        return year
    return year - 1


def generate_voucher_no(ulb, voucher_type, voucher_date):
    """Generate DNP/RECEIPT/26/01/31/0001 format with separate sequence per type per FY"""
    fy = get_financial_year(voucher_date)

    last_seq = Transaction.objects.filter(
        ulb=ulb,
        voucher_type=voucher_type,
        voucher_date__gte=date(fy, 4, 1),
        voucher_date__lte=date(fy + 1, 3, 31),
    ).aggregate(Max("sequence_no"))["sequence_no__max"] or 0

    next_seq = last_seq + 1

    yy = voucher_date.strftime("%y")
    mm = voucher_date.strftime("%m")
    dd = voucher_date.strftime("%d")
    seq_str = f"{next_seq:04d}"

    return f"{ulb.code}/{voucher_type}/{yy}/{mm}/{dd}/{seq_str}", next_seq

# ---------- CREATE: original transaction_entry view ----------

@login_required
@role_required(["ROOT_DEV", "DEV", "ADMIN", "USER"])
def transaction_entry(request):
    if not user_has_code(request.user, "MENU_ACCOUNTS_TRANSACTION_ENTRY"):
        return HttpResponseForbidden("You do not have permission for this action.")

    current_ulb_id = request.session.get("current_ulb_id")
    current_ulb = get_object_or_404(ULB, id=current_ulb_id) if current_ulb_id else None

    if not current_ulb:
        messages.error(request, "Please select a ULB first.")
        return redirect("ulb_select")

    ledgers = Ledger.objects.filter(ulb=current_ulb).select_related(
        "main_group", "subgroup", "group"
    )

    rows = []
    for ledger in ledgers:
        qs = TransactionEntry.objects.select_related("transaction", "ledger").filter(
            ledger=ledger
        )
        agg = qs.aggregate(
            dr_amount=Sum("dr_amount"),
            cr_amount=Sum("cr_amount"),
        )
        dr_amount = float(agg["dr_amount"] or 0)
        cr_amount = float(agg["cr_amount"] or 0)

        opening_amount = float(ledger.opening_balance or 0)
        opening_type = ledger.opening_type or ""

        rows.append(
            {
                "ledger": ledger,
                "opening_amount": opening_amount,
                "opening_type": opening_type,
                "dr_amount": dr_amount,
                "cr_amount": cr_amount,
            }
        )

    sidebar_context = {
        "current_ulb": current_ulb,
        "current_ulb_name": current_ulb.ulb_name if current_ulb else None,
        "allowed_codes": get_allowed_codes_for(request.user),
        "active_section": "transactions",
    }

    if request.method == "POST":
        with transaction.atomic():
            voucher_date = request.POST.get("voucher_date")
            voucher_type = request.POST.get("voucher_type")
            narration = request.POST.get("narration", "").strip()

            if not all([voucher_date, voucher_type]):
                messages.error(request, "Date and Voucher Type are required.")
                return render(
                    request,
                    "accounts/transaction_entry.html",
                    {
                        **sidebar_context,
                        "ledgers": ledgers,
                    },
                )

            voucher_date_obj = datetime.strptime(voucher_date, "%Y-%m-%d").date()

            voucher_no = request.POST.get("voucher_no", "").strip()
            if not voucher_no:
                messages.error(request, "Voucher number is missing.")
                return render(
                    request,
                    "accounts/transaction_entry.html",
                    {
                        **sidebar_context,
                        "ledgers": ledgers,
                    },
                )

            try:
                seq_str = voucher_no.rsplit("/", 1)[-1]
                sequence_no = int(seq_str)
            except (ValueError, IndexError):
                messages.error(request, "Invalid voucher number format.")
                return render(
                    request,
                    "accounts/transaction_entry.html",
                    {
                        **sidebar_context,
                        "ledgers": ledgers,
                    },
                )

            if Transaction.objects.filter(voucher_no=voucher_no).exists():
                messages.error(request, f"Voucher number {voucher_no} already exists.")
                return render(
                    request,
                    "accounts/transaction_entry.html",
                    {
                        **sidebar_context,
                        "ledgers": ledgers,
                    },
                )

            # main transaction row
            txn = Transaction.objects.create(
                ulb=current_ulb,
                voucher_type=voucher_type,
                voucher_date=voucher_date_obj,
                voucher_no=voucher_no,
                sequence_no=sequence_no,
                narration=narration,
            )

            # accounting entries
            entry_count = 1
            while True:
                type_key = f"type_{entry_count}"
                ledger_key = f"ledger_{entry_count}"
                dr_key = f"dr_amount_{entry_count}"
                cr_key = f"cr_amount_{entry_count}"

                if type_key not in request.POST:
                    break

                entry_type = request.POST.get(type_key)
                ledger_id = request.POST.get(ledger_key)
                dr_amount = request.POST.get(dr_key, "0")
                cr_amount = request.POST.get(cr_key, "0")

                if ledger_id:
                    ledger_obj = get_object_or_404(
                        Ledger, id=ledger_id, ulb=current_ulb
                    )

                    TransactionEntry.objects.create(
                        transaction=txn,
                        entry_type=entry_type,
                        ledger=ledger_obj,
                        dr_amount=float(dr_amount.replace(",", "")) if dr_amount else 0,
                        cr_amount=float(cr_amount.replace(",", "")) if cr_amount else 0,
                    )

                entry_count += 1

            # receipt-specific fields
            if voucher_type == VoucherType.RECEIPT:
                uc_applicable = request.POST.get("uc_applicable") == "yes"
                ReceiptUCDetails.objects.create(
                    transaction=txn,
                    uc_applicable=uc_applicable,
                    major_head=request.POST.get("major_head", ""),
                    treasury_code=request.POST.get("treasury_code", ""),
                    uc_bill_no=request.POST.get("uc_bill_no", ""),
                    uc_bill_date=request.POST.get("uc_bill_date") or None,
                    sub_major_head=request.POST.get("sub_major_head", ""),
                    ddo_code=request.POST.get("ddo_code", ""),
                    letter_no=request.POST.get("letter_no", ""),
                    letter_date=request.POST.get("letter_date") or None,
                    minor_head=request.POST.get("minor_head", ""),
                    bank_code=request.POST.get("bank_code", ""),
                    tv_no=request.POST.get("tv_no", ""),
                    tv_date=request.POST.get("tv_date") or None,
                    sub_head=request.POST.get("sub_head", ""),
                    bill_code=request.POST.get("bill_code", ""),
                    grant_amount=float(request.POST.get("grant_amount", 0) or 0),
                )

            # payment-specific fields
            elif voucher_type == VoucherType.PAYMENT:
                PaymentVendorDetails.objects.create(
                    transaction=txn,
                    vendor_name=request.POST.get("vendor_name", ""),
                    vendor_amount=float(request.POST.get("vendor_amount", 0) or 0),
                    cheque_no=request.POST.get("cheque_no", ""),
                    gst_applicable=request.POST.get("gst_applicable") == "yes",
                    gst_no=request.POST.get("gst_no", ""),
                    gst_type=request.POST.get("gst_type", ""),
                    gst_rate=float(request.POST.get("gst_rate", 0))
                    if request.POST.get("gst_rate")
                    else None,
                    igst_amount=float(request.POST.get("igst_amount", 0) or 0),
                    cgst_amount=float(request.POST.get("cgst_amount", 0) or 0),
                    sgst_amount=float(request.POST.get("sgst_amount", 0) or 0),
                    tds_applicable=request.POST.get("tds_applicable") == "yes",
                    tds_pan_no=request.POST.get("tds_pan_no", ""),
                    tds_section=request.POST.get("tds_section", ""),
                    tds_nature=request.POST.get("tds_nature", ""),
                    tds_type=request.POST.get("tds_type", ""),
                    tds_rate=request.POST.get("tds_rate", ""),
                    tds_amount=float(request.POST.get("tds_amount", 0) or 0),
                )

            messages.success(request, f"Transaction {voucher_no} saved successfully.")
            return redirect("accounts_transaction_entry")

    # GET
    return render(
        request,
        "accounts/transaction_entry.html",
        {
            **sidebar_context,
            "ledgers": ledgers,
        },
    )


@login_required
@role_required(["ROOT_DEV", "DEV", "ADMIN", "USER"])
def get_next_voucher_no(request):
    """AJAX endpoint to get next voucher number for a given type and date"""
    current_ulb_id = request.session.get("current_ulb_id")
    current_ulb = get_object_or_404(ULB, id=current_ulb_id) if current_ulb_id else None

    voucher_type = request.GET.get("voucher_type")
    voucher_date = request.GET.get("voucher_date")

    if not (current_ulb and voucher_type and voucher_date):
        from django.http import JsonResponse

        return JsonResponse({"error": "Missing data"}, status=400)

    try:
        voucher_date_obj = datetime.strptime(voucher_date, "%Y-%m-%d").date()
    except ValueError:
        from django.http import JsonResponse

        return JsonResponse({"error": "Invalid date"}, status=400)

    voucher_no, sequence_no = generate_voucher_no(
        current_ulb, voucher_type, voucher_date_obj
    )
    from django.http import JsonResponse

    return JsonResponse({"voucher_no": voucher_no, "sequence_no": sequence_no})


# ---------- EDIT: new separate view using transaction_edit.html ----------

@login_required
@role_required(["ROOT_DEV", "DEV", "ADMIN", "USER"])
def transaction_edit(request, txn_id):
    if not user_has_code(request.user, "MENU_ACCOUNTS_TRANSACTION_ENTRY"):
        return HttpResponseForbidden("You do not have permission for this action.")

    current_ulb_id = request.session.get("current_ulb_id")
    current_ulb = get_object_or_404(ULB, id=current_ulb_id) if current_ulb_id else None

    if not current_ulb:
        messages.error(request, "Please select a ULB first.")
        return redirect("ulb_select")

    txn = get_object_or_404(Transaction, id=txn_id, ulb=current_ulb)

    ledgers = Ledger.objects.filter(ulb=current_ulb).select_related(
        "main_group", "subgroup", "group"
    )

    entries = TransactionEntry.objects.filter(transaction=txn).select_related(
        "ledger"
    ).order_by("id")

    receipt_details = None
    payment_details = None
    if txn.voucher_type == VoucherType.RECEIPT:
        receipt_details = ReceiptUCDetails.objects.filter(transaction=txn).first()
    elif txn.voucher_type == VoucherType.PAYMENT:
        payment_details = PaymentVendorDetails.objects.filter(transaction=txn).first()

    sidebar_context = {
        "current_ulb": current_ulb,
        "current_ulb_name": current_ulb.ulb_name if current_ulb else None,
        "allowed_codes": get_allowed_codes_for(request.user),
        "active_section": "transactions",
    }

    if request.method == "POST":
        with transaction.atomic():
            voucher_date = request.POST.get("voucher_date")
            voucher_type = request.POST.get("voucher_type")
            narration = request.POST.get("narration", "").strip()

            if not all([voucher_date, voucher_type]):
                messages.error(request, "Date and Voucher Type are required.")
                return render(
                    request,
                    "accounts/transaction_edit.html",
                    {
                        **sidebar_context,
                        "txn": txn,
                        "ledgers": ledgers,
                        "entries": entries,
                        "receipt_details": receipt_details,
                        "payment_details": payment_details,
                    },
                )

            voucher_date_obj = datetime.strptime(voucher_date, "%Y-%m-%d").date()

            voucher_no = request.POST.get("voucher_no", "").strip()
            if not voucher_no:
                messages.error(request, "Voucher number is missing.")
                return render(
                    request,
                    "accounts/transaction_edit.html",
                    {
                        **sidebar_context,
                        "txn": txn,
                        "ledgers": ledgers,
                        "entries": entries,
                        "receipt_details": receipt_details,
                        "payment_details": payment_details,
                    },
                )

            # if voucher_no changed, enforce uniqueness
            if voucher_no != txn.voucher_no and Transaction.objects.filter(
                voucher_no=voucher_no
            ).exists():
                messages.error(request, f"Voucher number {voucher_no} already exists.")
                return render(
                    request,
                    "accounts/transaction_edit.html",
                    {
                        **sidebar_context,
                        "txn": txn,
                        "ledgers": ledgers,
                        "entries": entries,
                        "receipt_details": receipt_details,
                        "payment_details": payment_details,
                    },
                )

            # update main transaction (keep same sequence_no)
            txn.voucher_date = voucher_date_obj
            txn.voucher_type = voucher_type
            txn.voucher_no = voucher_no
            txn.narration = narration
            txn.save()

            # rebuild accounting entries
            TransactionEntry.objects.filter(transaction=txn).delete()

            entry_count = 1
            while True:
                type_key = f"type_{entry_count}"
                ledger_key = f"ledger_{entry_count}"
                dr_key = f"dr_amount_{entry_count}"
                cr_key = f"cr_amount_{entry_count}"

                if type_key not in request.POST:
                    break

                entry_type = request.POST.get(type_key)
                ledger_id = request.POST.get(ledger_key)
                dr_amount = request.POST.get(dr_key, "0")
                cr_amount = request.POST.get(cr_key, "0")

                if ledger_id:
                    ledger_obj = get_object_or_404(
                        Ledger, id=ledger_id, ulb=current_ulb
                    )
                    TransactionEntry.objects.create(
                        transaction=txn,
                        entry_type=entry_type,
                        ledger=ledger_obj,
                        dr_amount=float(dr_amount.replace(",", "")) if dr_amount else 0,
                        cr_amount=float(cr_amount.replace(",", "")) if cr_amount else 0,
                    )

                entry_count += 1

            # update receipt/payment details
            if voucher_type == VoucherType.RECEIPT:
                ReceiptUCDetails.objects.filter(transaction=txn).delete()
                uc_applicable = request.POST.get("uc_applicable") == "yes"
                ReceiptUCDetails.objects.create(
                    transaction=txn,
                    uc_applicable=uc_applicable,
                    major_head=request.POST.get("major_head", ""),
                    treasury_code=request.POST.get("treasury_code", ""),
                    uc_bill_no=request.POST.get("uc_bill_no", ""),
                    uc_bill_date=request.POST.get("uc_bill_date") or None,
                    sub_major_head=request.POST.get("sub_major_head", ""),
                    ddo_code=request.POST.get("ddo_code", ""),
                    letter_no=request.POST.get("letter_no", ""),
                    letter_date=request.POST.get("letter_date") or None,
                    minor_head=request.POST.get("minor_head", ""),
                    bank_code=request.POST.get("bank_code", ""),
                    tv_no=request.POST.get("tv_no", ""),
                    tv_date=request.POST.get("tv_date") or None,
                    sub_head=request.POST.get("sub_head", ""),
                    bill_code=request.POST.get("bill_code", ""),
                    grant_amount=float(request.POST.get("grant_amount", 0) or 0),
                )
            elif voucher_type == VoucherType.PAYMENT:
                PaymentVendorDetails.objects.filter(transaction=txn).delete()
                PaymentVendorDetails.objects.create(
                    transaction=txn,
                    vendor_name=request.POST.get("vendor_name", ""),
                    vendor_amount=float(request.POST.get("vendor_amount", 0) or 0),
                    cheque_no=request.POST.get("cheque_no", ""),
                    gst_applicable=request.POST.get("gst_applicable") == "yes",
                    gst_no=request.POST.get("gst_no", ""),
                    gst_type=request.POST.get("gst_type", ""),
                    gst_rate=float(request.POST.get("gst_rate", 0))
                    if request.POST.get("gst_rate")
                    else None,
                    igst_amount=float(request.POST.get("igst_amount", 0) or 0),
                    cgst_amount=float(request.POST.get("cgst_amount", 0) or 0),
                    sgst_amount=float(request.POST.get("sgst_amount", 0) or 0),
                    tds_applicable=request.POST.get("tds_applicable") == "yes",
                    tds_pan_no=request.POST.get("tds_pan_no", ""),
                    tds_section=request.POST.get("tds_section", ""),
                    tds_nature=request.POST.get("tds_nature", ""),
                    tds_type=request.POST.get("tds_type", ""),
                    tds_rate=request.POST.get("tds_rate", ""),
                    tds_amount=float(request.POST.get("tds_amount", 0) or 0),
                )

            messages.success(request, f"Transaction {voucher_no} updated successfully.")
            return redirect("accounts_day_book")

    # GET -> render edit form
    return render(
        request,
        "accounts/transaction_edit.html",
        {
            **sidebar_context,
            "txn": txn,
            "ledgers": ledgers,
            "entries": entries,
            "receipt_details": receipt_details,
            "payment_details": payment_details,
        },
    )
# ---------- DUPLICATE: new view using transaction_edit.html but pre-filled with existing data ----------
def get_financial_year(d):
    year = d.year
    if d.month >= 4:
        return year
    return year - 1


def generate_voucher_no(ulb, voucher_type, voucher_date):
    fy = get_financial_year(voucher_date)

    last_seq = Transaction.objects.filter(
        ulb=ulb,
        voucher_type=voucher_type,
        voucher_date__gte=date(fy, 4, 1),
        voucher_date__lte=date(fy + 1, 3, 31),
    ).aggregate(Max("sequence_no"))["sequence_no__max"] or 0

    next_seq = last_seq + 1

    yy = voucher_date.strftime("%y")
    mm = voucher_date.strftime("%m")
    dd = voucher_date.strftime("%d")
    seq_str = f"{next_seq:04d}"

    return f"{ulb.code}/{voucher_type}/{yy}/{mm}/{dd}/{seq_str}", next_seq


@login_required
@role_required(["ROOT_DEV", "DEV", "ADMIN", "USER"])
def transaction_duplicate(request, txn_id):
    if not user_has_code(request.user, "MENU_ACCOUNTS_TRANSACTION_ENTRY"):
        return HttpResponseForbidden("You do not have permission for this action.")

    current_ulb_id = request.session.get("current_ulb_id")
    current_ulb = get_object_or_404(ULB, id=current_ulb_id) if current_ulb_id else None

    if not current_ulb:
        messages.error(request, "Please select a ULB first.")
        return redirect("ulb_select")

    # original transaction to duplicate
    original_txn = get_object_or_404(Transaction, id=txn_id, ulb=current_ulb)

    ledgers = Ledger.objects.filter(ulb=current_ulb).select_related(
        "main_group", "subgroup", "group"
    )
    entries = (
        TransactionEntry.objects.filter(transaction=original_txn)
        .select_related("ledger")
        .order_by("id")
    )

    receipt_details = None
    payment_details = None
    if original_txn.voucher_type == VoucherType.RECEIPT:
        receipt_details = ReceiptUCDetails.objects.filter(
            transaction=original_txn
        ).first()
    elif original_txn.voucher_type == VoucherType.PAYMENT:
        payment_details = PaymentVendorDetails.objects.filter(
            transaction=original_txn
        ).first()

    sidebar_context = {
        "current_ulb": current_ulb,
        "current_ulb_name": current_ulb.ulb_name if current_ulb else None,
        "allowed_codes": get_allowed_codes_for(request.user),
        "active_section": "transactions",
    }

    # base date for suggested voucher number (use original date by default)
    base_voucher_date = original_txn.voucher_date or date.today()

    if request.method == "POST":
        with transaction.atomic():
            voucher_date_str = request.POST.get("voucher_date")
            voucher_type = request.POST.get("voucher_type") or original_txn.voucher_type
            narration = request.POST.get("narration", "").strip()

            if not voucher_date_str:
                messages.error(request, "Date is required.")
                return render(
                    request,
                    "accounts/transaction_duplicate.html",
                    {
                        **sidebar_context,
                        "txn": original_txn,
                        "entries": entries,
                        "ledgers": ledgers,
                        "receipt_details": receipt_details,
                        "payment_details": payment_details,
                        "voucher_date": voucher_date_str,
                        "suggested_voucher_no": request.POST.get("voucher_no", ""),
                    },
                )

            voucher_date = datetime.strptime(voucher_date_str, "%Y-%m-%d").date()

            # always compute next voucher_no & sequence at save time to ensure "next to last"
            voucher_no, sequence_no = generate_voucher_no(
                current_ulb, voucher_type, voucher_date
            )

            # create new transaction (duplicate)
            new_txn = Transaction.objects.create(
                ulb=current_ulb,
                voucher_type=voucher_type,
                voucher_date=voucher_date,
                voucher_no=voucher_no,
                sequence_no=sequence_no,
                narration=narration,
            )

            # accounting entries
            entry_count = 1
            while True:
                type_key = f"type_{entry_count}"
                ledger_key = f"ledger_{entry_count}"
                dr_key = f"dr_amount_{entry_count}"
                cr_key = f"cr_amount_{entry_count}"

                if type_key not in request.POST:
                    break

                entry_type = request.POST.get(type_key)
                ledger_id = request.POST.get(ledger_key)
                dr_amount = request.POST.get(dr_key, "0")
                cr_amount = request.POST.get(cr_key, "0")

                if ledger_id:
                    ledger_obj = get_object_or_404(
                        Ledger, id=ledger_id, ulb=current_ulb
                    )
                    TransactionEntry.objects.create(
                        transaction=new_txn,
                        entry_type=entry_type,
                        ledger=ledger_obj,
                        dr_amount=float(dr_amount.replace(",", "")) if dr_amount else 0,
                        cr_amount=float(cr_amount.replace(",", "")) if cr_amount else 0,
                    )

                entry_count += 1

            # receipt/payment details duplicate from form
            if voucher_type == VoucherType.RECEIPT:
                uc_applicable = request.POST.get("uc_applicable") == "yes"
                ReceiptUCDetails.objects.create(
                    transaction=new_txn,
                    uc_applicable=uc_applicable,
                    major_head=request.POST.get("major_head", ""),
                    treasury_code=request.POST.get("treasury_code", ""),
                    uc_bill_no=request.POST.get("uc_bill_no", ""),
                    uc_bill_date=request.POST.get("uc_bill_date") or None,
                    sub_major_head=request.POST.get("sub_major_head", ""),
                    ddo_code=request.POST.get("ddo_code", ""),
                    letter_no=request.POST.get("letter_no", ""),
                    letter_date=request.POST.get("letter_date") or None,
                    minor_head=request.POST.get("minor_head", ""),
                    bank_code=request.POST.get("bank_code", ""),
                    tv_no=request.POST.get("tv_no", ""),
                    tv_date=request.POST.get("tv_date") or None,
                    sub_head=request.POST.get("sub_head", ""),
                    bill_code=request.POST.get("bill_code", ""),
                    grant_amount=float(request.POST.get("grant_amount", 0) or 0),
                )
            elif voucher_type == VoucherType.PAYMENT:
                PaymentVendorDetails.objects.create(
                    transaction=new_txn,
                    vendor_name=request.POST.get("vendor_name", ""),
                    vendor_amount=float(request.POST.get("vendor_amount", 0) or 0),
                    cheque_no=request.POST.get("cheque_no", ""),
                    gst_applicable=request.POST.get("gst_applicable") == "yes",
                    gst_no=request.POST.get("gst_no", ""),
                    gst_type=request.POST.get("gst_type", ""),
                    gst_rate=float(request.POST.get("gst_rate", 0))
                    if request.POST.get("gst_rate")
                    else None,
                    igst_amount=float(request.POST.get("igst_amount", 0) or 0),
                    cgst_amount=float(request.POST.get("cgst_amount", 0) or 0),
                    sgst_amount=float(request.POST.get("sgst_amount", 0) or 0),
                    tds_applicable=request.POST.get("tds_applicable") == "yes",
                    tds_pan_no=request.POST.get("tds_pan_no", ""),
                    tds_section=request.POST.get("tds_section", ""),
                    tds_nature=request.POST.get("tds_nature", ""),
                    tds_type=request.POST.get("tds_type", ""),
                    tds_rate=request.POST.get("tds_rate", ""),
                    tds_amount=float(request.POST.get("tds_amount", 0) or 0),
                )

            messages.success(
                request,
                f"Duplicate transaction created with voucher number {voucher_no}.",
            )
            return redirect("accounts_day_book")

    # GET: suggest next voucher number using original date/type
    suggested_voucher_no, _ = generate_voucher_no(
        current_ulb, original_txn.voucher_type, base_voucher_date
    )

    return render(
        request,
        "accounts/transaction_duplicate.html",
        {
            **sidebar_context,
            "txn": original_txn,
            "entries": entries,
            "ledgers": ledgers,
            "receipt_details": receipt_details,
            "payment_details": payment_details,
            "voucher_date": base_voucher_date.strftime("%Y-%m-%d"),
            "suggested_voucher_no": suggested_voucher_no,
        },
    )
# ---------- CASH BOOK: new view with left/right split and filters ----------
from decimal import Decimal
from datetime import date, datetime
def _get_cashbook_date_range(request):
    """Current FY (April–March) default, override via GET ?from_date=&to_date= (YYYY-MM-DD)."""
    today = date.today()
    fy_start_year = today.year if today.month >= 4 else today.year - 1
    default_from = date(fy_start_year, 4, 1)
    default_to = today

    from_str = request.GET.get("from_date") or ""
    to_str = request.GET.get("to_date") or ""

    from_date = default_from
    to_date = default_to

    if from_str:
        try:
            from_date = datetime.strptime(from_str, "%Y-%m-%d").date()
        except ValueError:
            from_date = default_from

    if to_str:
        try:
            to_date = datetime.strptime(to_str, "%Y-%m-%d").date()
        except ValueError:
            to_date = default_to

    return from_date, to_date


def _get_ledger_opening_signed(ledger, as_on_date):
    """
    Return signed opening for one ledger up to as_on_date.
    DR = positive, CR = negative.
    If opening_date is after as_on_date, ignore that opening.
    """
    if not ledger.opening_balance or not ledger.opening_type or not ledger.opening_date:
        return Decimal("0")

    if ledger.opening_date > as_on_date:
        return Decimal("0")

    bal = Decimal(ledger.opening_balance or 0)
    if ledger.opening_type == "DR":
        return bal
    else:  # "CR"
        return -bal


@login_required
@role_required(["ROOT_DEV", "DEV", "ADMIN", "USER"])
def cash_book(request):
    """
    Cash book (Main / Subsidiary) double-page:
    - Left: Receipts + Contra
    - Right: Payments + Contra
    - Contra placement:
      * No ledger selected: both sides.
      * Bank ledger selected: Receipts only.
      * Cash ledger selected: Payments only.
    """
    if not user_has_code(request.user, "MENU_ACCOUNTS_CASH_BOOK"):
        return HttpResponseForbidden("You do not have permission for this action.")

    current_ulb_id = request.session.get("current_ulb_id")
    current_ulb = get_object_or_404(ULB, id=current_ulb_id) if current_ulb_id else None
    if not current_ulb:
        return redirect("ulb_select")

    # Date range (default current FY)
    from_date, to_date = _get_cashbook_date_range(request)

    # All 450* ledgers (cash + bank module)
    cash_ledgers = (
        Ledger.objects.filter(
            ulb=current_ulb,
            name__startswith="450",
        ).order_by("name")
    )
    for l in cash_ledgers:
        name = l.name or ""
        first, sep, rest = name.partition(" ")
        l.display_name = rest if rest else name

    selected_ledger_id = request.GET.get("ledger_id") or ""
    selected_ledger = None
    if selected_ledger_id:
        try:
            selected_ledger = cash_ledgers.get(id=selected_ledger_id)
        except Ledger.DoesNotExist:
            selected_ledger = None

    # Base queryset: all transactions in date range for this ULB
    qs = (
        Transaction.objects.filter(
            ulb=current_ulb,
            voucher_date__gte=from_date,
            voucher_date__lte=to_date,
        )
        .prefetch_related("entries__ledger")
    )

    amount_str = (request.GET.get("amount") or "").strip()
    voucher_no = (request.GET.get("voucher_no") or "").strip()
    cheque_no = (request.GET.get("cheque_no") or "").strip()
    export = request.GET.get("export")  # "excel" to export

    if voucher_no:
        qs = qs.filter(voucher_no__icontains=voucher_no)

    if cheque_no:
        qs = qs.filter(paymentvendordetails__cheque_no__icontains=cheque_no)

    # Subsidiary: only transactions touching the selected ledger
    if selected_ledger:
        qs = qs.filter(entries__ledger=selected_ledger)

    # Annotate totals for amount filter (whole voucher)
    qs = qs.annotate(
        dr_total=Sum("entries__dr_amount"),
        cr_total=Sum("entries__cr_amount"),
    )

    if amount_str:
        try:
            amt = float(amount_str)
            qs = qs.filter(Q(dr_total=amt) | Q(cr_total=amt))
        except ValueError:
            pass

    receipts_entries = []
    payments_entries = []
    total_receipts = 0.0
    total_payments = 0.0

    def build_transaction_rows(t):
        """
        Build lines, LF markers, details for non-cash ledgers,
        and list of cash-module entries (ledgers in cash_ledgers).
        Also track cash Dr / Cr totals separately so we don't
        misuse overall voucher totals for cash book amounts.
        """
        lines = []
        lf_lines = []
        non_cash_details = []
        cash_entries = []  # list of (ledger, dr, cr)

        cash_dr_total = 0.0
        cash_cr_total = 0.0

        for le in t.entries.all():
            full_name = le.ledger.name or ""
            first, sep, rest = full_name.partition(" ")
            display_name = rest if rest else full_name

            lines.append(display_name)

            dr = float(le.dr_amount or 0)
            cr = float(le.cr_amount or 0)
            if dr:
                lf_lines.append((display_name, "Dr"))
            if cr:
                lf_lines.append((display_name, "Cr"))

            if le.ledger in cash_ledgers:
                cash_entries.append((le.ledger, dr, cr))
                cash_dr_total += dr
                cash_cr_total += cr
            else:
                if dr:
                    non_cash_details.append(dr)
                if cr:
                    non_cash_details.append(cr)

        # cheque for payments
        pv = getattr(t, "paymentvendordetails", None)
        cheque_number = pv.cheque_no if pv else ""

        return {
            "date": t.voucher_date,
            "voucher_no": t.voucher_no,
            "cheque_no": cheque_number,
            "lines": lines,
            "lf_lines": lf_lines,
            "non_cash_details": non_cash_details,
            "cash_entries": cash_entries,
            "narration": t.narration or "",
            "voucher_type": t.voucher_type,
            "cash_dr_total": cash_dr_total,
            "cash_cr_total": cash_cr_total,
        }

    for t in qs.order_by("voucher_date", "id"):
        info = build_transaction_rows(t)
        vt = info["voucher_type"]
        cash_entries = info["cash_entries"]

        left_amount = 0.0
        right_amount = 0.0

        if vt == "RECV":
            # Receipts: cash Dr only (not full dr_total)
            left_amount = info["cash_dr_total"]
        elif vt == "PYMT":
            # Payments: cash Cr only
            right_amount = info["cash_cr_total"]
        elif vt == "CNTR":
            # For contra we expect two cash-module legs: one Dr, one Cr
            dr_leg = next((ce for ce in cash_entries if ce[1] > 0), None)
            cr_leg = next((ce for ce in cash_entries if ce[2] > 0), None)

            if dr_leg:
                left_amount = dr_leg[1]   # Dr side -> Receipts
            if cr_leg:
                right_amount = cr_leg[2]  # Cr side -> Payments

        show_on_left = False
        show_on_right = False

        if vt in ("RECV", "PYMT"):
            show_on_left = (vt == "RECV")
            show_on_right = (vt == "PYMT")
        elif vt == "CNTR":
            dr_leg = next((ce for ce in cash_entries if ce[1] > 0), None)
            cr_leg = next((ce for ce in cash_entries if ce[2] > 0), None)

            if not selected_ledger:
                # Main cash book: show contra on both sides
                show_on_left = left_amount > 0
                show_on_right = right_amount > 0
            else:
                # Subsidiary: decide side based on which leg is the selected ledger
                if dr_leg and selected_ledger == dr_leg[0]:
                    # Selected bank ledger -> Receipts only
                    show_on_left = left_amount > 0
                    show_on_right = False
                elif cr_leg and selected_ledger == cr_leg[0]:
                    # Selected cash ledger -> Payments only
                    show_on_left = False
                    show_on_right = right_amount > 0
                else:
                    show_on_left = False
                    show_on_right = False

        # Build left row
        if show_on_left and left_amount:
            details_left = list(info["non_cash_details"])
            # IMPORTANT: do not append left_amount again for contra,
            # otherwise the cash leg gets double-counted in details.
            # if vt == "CNTR" and left_amount:
            #     details_left.append(left_amount)

            receipts_entries.append({
                "date": info["date"],
                "voucher_no": info["voucher_no"],
                "cheque_no": info["cheque_no"],
                "lines": info["lines"],
                "lf_lines": info["lf_lines"],
                "details": details_left,
                "amount": left_amount,
                "narration": info["narration"],
            })
            total_receipts += left_amount

        # Build right row
        if show_on_right and right_amount:
            details_right = list(info["non_cash_details"])
            # Same for payments side
            # if vt == "CNTR" and right_amount:
            #     details_right.append(right_amount)

            payments_entries.append({
                "date": info["date"],
                "voucher_no": info["voucher_no"],
                "cheque_no": info["cheque_no"],
                "lines": info["lines"],
                "lf_lines": info["lf_lines"],
                "details": details_right,
                "amount": right_amount,
                "narration": info["narration"],
            })
            total_payments += right_amount

    # Opening balance: master opening + movement before from_date
    cash_ledger_q = cash_ledgers
    if selected_ledger:
        cash_ledger_q = cash_ledgers.filter(id=selected_ledger.id)

    opening_from_master = Decimal("0")
    for l in cash_ledger_q:
        opening_from_master += _get_ledger_opening_signed(l, from_date)

    entries_before = TransactionEntry.objects.filter(
        transaction__ulb=current_ulb,
        transaction__voucher_date__lt=from_date,
        ledger__in=cash_ledger_q,
    ).aggregate(
        dr=Sum("dr_amount"),
        cr=Sum("cr_amount"),
    )
    opening_dr_txn = Decimal(entries_before.get("dr") or 0)
    opening_cr_txn = Decimal(entries_before.get("cr") or 0)
    opening_from_txn = opening_dr_txn - opening_cr_txn

    opening_balance = opening_from_master + opening_from_txn

    opening_side = "Dr" if opening_balance >= 0 else "Cr"
    opening_amount_abs = abs(opening_balance)

    closing_balance = opening_balance + Decimal(str(total_receipts)) - Decimal(str(total_payments))

    grand_total_left = opening_balance + Decimal(str(total_receipts))
    grand_total_right = Decimal(str(total_payments)) + closing_balance

    # -------- Excel Export --------
    if export == "excel":
        wb = Workbook()
        ws = wb.active
        ws.title = "Cash Book"

        blue_fill = PatternFill(fill_type="solid", fgColor="1E90FF")
        bold_center = Alignment(horizontal="center", vertical="center")

        def fmt_date(d):
            if not d:
                return ""
            return d.strftime("%d-%m-%Y")

        # Dynamic title same as web:
        # - No ledger selected => "Main Cash Book"
        # - Ledger selected    => "Subsidiary Book Of <Ledger Name>"
        if selected_ledger:
            name = selected_ledger.name or ""
            first, sep, rest = name.partition(" ")
            disp_name = rest if rest else name
            title_text = f"Subsidiary Book Of {disp_name}"
        else:
            title_text = "Main Cash Book"

        # Row 1: Title
        ws.merge_cells("A1:N1")
        ws["A1"].value = title_text
        ws["A1"].font = Font(size=14, bold=True)
        ws["A1"].alignment = bold_center

        # Row 2: ULB
        ws.merge_cells("A2:N2")
        ws["A2"].value = current_ulb.ulb_name if current_ulb else ""
        ws["A2"].font = Font(size=12, bold=True)
        ws["A2"].alignment = bold_center

        # Row 3: Period
        ws.merge_cells("A3:N3")
        ws["A3"].value = f"Period: {fmt_date(from_date)} to {fmt_date(to_date)}"
        ws["A3"].font = Font(size=11, bold=False)
        ws["A3"].alignment = bold_center

        # Row 4: Receipts / Payments labels
        ws.merge_cells("A4:G4")
        ws["A4"].value = "RECEIPTS"
        ws["A4"].font = Font(bold=True, color="FFFFFF")
        ws["A4"].fill = blue_fill
        ws["A4"].alignment = bold_center

        ws.merge_cells("H4:N4")
        ws["H4"].value = "PAYMENTS"
        ws["H4"].font = Font(bold=True, color="FFFFFF")
        ws["H4"].fill = blue_fill
        ws["H4"].alignment = bold_center

        # Row 5: column headers
        ws.append([
            "Date", "Voucher No.", "Cheque No.", "Particulars", "L/F", "Details", "Amount",
            "Date", "Voucher No.", "Cheque No.", "Particulars", "L/F", "Details", "Amount",
        ])

        # Row 6: Opening balance row (left side)
        ws.append([
            fmt_date(from_date),
            "Opening Balance",
            opening_side,
            "", "", "",
            float(opening_amount_abs),
            "", "", "", "", "", "", "",
        ])

        # Data rows with narration rows
        max_len = max(len(receipts_entries), len(payments_entries))

        for i in range(max_len):
            recv = receipts_entries[i] if i < len(receipts_entries) else None
            pay = payments_entries[i] if i < len(payments_entries) else None

            recv_parts = "\n".join(recv["lines"]) if recv else ""
            recv_lfs = "\n".join(lf[1] for lf in recv["lf_lines"]) if recv else ""
            recv_details = "\n".join(str(d) for d in recv["details"]) if recv else ""

            pay_parts = "\n".join(pay["lines"]) if pay else ""
            pay_lfs = "\n".join(lf[1] for lf in pay["lf_lines"]) if pay else ""
            pay_details = "\n".join(str(d) for d in pay["details"]) if pay else ""

            ws.append([
                fmt_date(recv["date"]) if recv else "",
                recv["voucher_no"] if recv else "",
                recv["cheque_no"] if recv else "",
                recv_parts,
                recv_lfs,
                recv_details,
                float(recv["amount"]) if recv else "",
                fmt_date(pay["date"]) if pay else "",
                pay["voucher_no"] if pay else "",
                pay["cheque_no"] if pay else "",
                pay_parts,
                pay_lfs,
                pay_details,
                float(pay["amount"]) if pay else "",
            ])

            narr_left = f"Narration: {recv['narration']}" if recv and recv["narration"] else ""
            narr_right = f"Narration: {pay['narration']}" if pay and pay["narration"] else ""

            ws.append([
                "", "", "", narr_left, "", "", "",
                "", "", "", narr_right, "", "", "",
            ])

        # Blank row
        ws.append([])
        footer_start = ws.max_row + 1

        # Total Receipts (left)
        ws.merge_cells(start_row=footer_start, start_column=1, end_row=footer_start, end_column=6)
        ws.cell(row=footer_start, column=1).value = "Total Receipts"
        ws.cell(row=footer_start, column=1).font = Font(bold=True)
        ws.cell(row=footer_start, column=7).value = float(total_receipts)

        # Total Payments (right)
        ws.merge_cells(start_row=footer_start, start_column=8, end_row=footer_start, end_column=13)
        ws.cell(row=footer_start, column=8).value = "Total Payments"
        ws.cell(row=footer_start, column=8).font = Font(bold=True)
        ws.cell(row=footer_start, column=14).value = float(total_payments)

        # Closing Balance (right)
        footer_row_cb = footer_start + 1
        ws.merge_cells(start_row=footer_row_cb, start_column=8, end_row=footer_row_cb, end_column=13)
        ws.cell(row=footer_row_cb, column=8).value = "Closing Balance"
        ws.cell(row=footer_row_cb, column=8).font = Font(bold=True)
        ws.cell(row=footer_row_cb, column=14).value = float(closing_balance)

        # Grand Totals
        footer_row_gt = footer_start + 2
        ws.merge_cells(start_row=footer_row_gt, start_column=1, end_row=footer_row_gt, end_column=6)
        ws.cell(row=footer_row_gt, column=1).value = "Grand Total"
        ws.cell(row=footer_row_gt, column=1).font = Font(bold=True)
        ws.cell(row=footer_row_gt, column=7).value = float(grand_total_left)

        ws.merge_cells(start_row=footer_row_gt, start_column=8, end_row=footer_row_gt, end_column=13)
        ws.cell(row=footer_row_gt, column=8).value = "Grand Total"
        ws.cell(row=footer_row_gt, column=8).font = Font(bold=True)
        ws.cell(row=footer_row_gt, column=14).value = float(grand_total_right)

        # Auto-width for all columns
        for col_idx in range(1, 15):  # A..N
            col_letter = get_column_letter(col_idx)
            max_length = 0
            for r in range(1, ws.max_row + 1):
                cell = ws.cell(row=r, column=col_idx)
                if cell.value is not None:
                    length = len(str(cell.value))
                    if length > max_length:
                        max_length = length
            ws.column_dimensions[col_letter].width = (max_length + 2) * 1.2

        # Fixed width for D and K
        ws.column_dimensions["D"].width = 57
        ws.column_dimensions["K"].width = 57

        # Wrap text for particulars / details / narration columns
        wrap_columns = [4, 5, 6, 11, 12, 13]  # D,E,F and K,L,M
        for row in ws.iter_rows(min_row=6, max_row=ws.max_row, min_col=1, max_col=14):
            for cell in row:
                if cell.column in wrap_columns:
                    cell.alignment = Alignment(wrapText=True, vertical="top")

        # Borders
        thin = Side(style="thin", color="000000")
        border = Border(top=thin, left=thin, right=thin, bottom=thin)
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=14):
            for cell in row:
                cell.border = border

        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = 'attachment; filename="cash_book.xlsx"'
        wb.save(response)
        return response

    # ---------- normal HTML render ----------
    sidebar_context = {
        "current_ulb": current_ulb,
        "current_ulb_name": current_ulb.ulb_name if current_ulb else None,
        "allowed_codes": get_allowed_codes_for(request.user),
        "active_section": "cash_book",
    }

    context = {
        **sidebar_context,
        "cash_ledgers": cash_ledgers,
        "selected_ledger": selected_ledger,
        "from_date": from_date,
        "to_date": to_date,
        "opening_from_date": from_date,
        "opening_balance": opening_balance,
        "opening_side": opening_side,
        "opening_amount_abs": opening_amount_abs,
        "receipts_entries": receipts_entries,
        "payments_entries": payments_entries,
        "total_receipts": total_receipts,
        "total_payments": total_payments,
        "closing_balance": closing_balance,
        "grand_total_left": grand_total_left,
        "grand_total_right": grand_total_right,
    }
    return render(request, "accounts/cash_book.html", context)

# ---------- Ledger view ----------
def _get_cashbook_date_range(request):
    """Current FY (April–March) default, override via GET ?from_date=&to_date= (YYYY-MM-DD)."""
    today = date.today()
    fy_start_year = today.year if today.month >= 4 else today.year - 1
    default_from = date(fy_start_year, 4, 1)
    default_to = today

    from_str = request.GET.get("from_date") or ""
    to_str = request.GET.get("to_date") or ""

    from_date = default_from
    to_date = default_to

    if from_str:
        try:
            from_date = datetime.strptime(from_str, "%Y-%m-%d").date()
        except ValueError:
            from_date = default_from

    if to_str:
        try:
            to_date = datetime.strptime(to_str, "%Y-%m-%d").date()
        except ValueError:
            to_date = default_to

    return from_date, to_date

def _get_ledger_opening_signed(ledger, as_on_date):
    """
    Return signed opening for one ledger up to as_on_date.
    DR = positive, CR = negative.
    If opening_date is after as_on_date, ignore that opening.
    """
    if not ledger.opening_balance or not ledger.opening_type or not ledger.opening_date:
        return Decimal("0")

    if ledger.opening_date > as_on_date:
        return Decimal("0")

    bal = Decimal(ledger.opening_balance or 0)
    if ledger.opening_type == "DR":
        return bal
    else:  # "CR"
        return -bal


@login_required
@role_required(["ROOT_DEV", "DEV", "ADMIN", "USER"])
def ledger(request):
    """
    Ledger book (Dr/Cr two-page) for a single ledger:
    - Left: Debit side (where this ledger is Dr)
    - Right: Credit side (where this ledger is Cr)
    - Filter by date range, voucher no, amount, cheque no, ledger (excluding 450*).
    """
    if not user_has_code(request.user, "MENU_ACCOUNTS_LEDGER"):
        return HttpResponseForbidden("You do not have permission for this action.")

    # current ULB from session
    current_ulb_id = request.session.get("current_ulb_id")
    current_ulb = get_object_or_404(ULB, id=current_ulb_id) if current_ulb_id else None
    if not current_ulb:
        return redirect("ulb_select")

    # Date range (default current FY)
    from_date, to_date = _get_cashbook_date_range(request)

    # All ledgers for this ULB except those starting with 450 (cash/bank)
    all_ledgers = (
        Ledger.objects.filter(ulb=current_ulb)
        .exclude(name__startswith="450")  # exclude cash/bank ledgers
        .order_by("name")
    )

    # Attach display_name (strip code) if you want in dropdown/body
    for l in all_ledgers:
        name = l.name or ""
        first, sep, rest = name.partition(" ")
        l.display_name = rest if rest else name

    selected_ledger_id = request.GET.get("ledger_id") or ""
    selected_ledger = None
    if selected_ledger_id:
        try:
            selected_ledger = all_ledgers.get(id=selected_ledger_id)
        except Ledger.DoesNotExist:
            selected_ledger = None

    # Base queryset: all transactions in date range for this ULB
    qs = (
        Transaction.objects.filter(
            ulb=current_ulb,
            voucher_date__gte=from_date,
            voucher_date__lte=to_date,
        )
        .prefetch_related("entries__ledger")
    )

    # Additional filters
    amount_str = (request.GET.get("amount") or "").strip()
    voucher_no = (request.GET.get("voucher_no") or "").strip()
    cheque_no = (request.GET.get("cheque_no") or "").strip()
    export = (request.GET.get("export") or "").strip().lower()  # "excel" to export

    if voucher_no:
        qs = qs.filter(voucher_no__icontains=voucher_no)

    if cheque_no:
        qs = qs.filter(paymentvendordetails__cheque_no__icontains=cheque_no)

    # If a specific ledger is selected, restrict to transactions that touch that ledger
    if selected_ledger:
        qs = qs.filter(entries__ledger=selected_ledger)

    # Annotate totals for amount filter, if any (dr_total/cr_total on whole voucher)
    qs = qs.annotate(
        dr_total=Sum("entries__dr_amount"),
        cr_total=Sum("entries__cr_amount"),
    )

    if amount_str:
        try:
            amt = float(amount_str)
            qs = qs.filter(Q(dr_total=amt) | Q(cr_total=amt))
        except ValueError:
            pass

    # two-page: left = this ledger Dr, right = this ledger Cr
    receipts_entries = []  # Dr
    payments_entries = []  # Cr
    total_receipts = 0.0
    total_payments = 0.0

    def build_transaction_rows_for_ledger(t, ledger_obj):
        """
        For this transaction, build rows focusing on one ledger (ledger_obj).
        Left: when this ledger is Dr.
        Right: when this ledger is Cr.
        lines    = other ledgers' names.
        lf_lines = Dr/Cr for other ledgers.
        details  = amounts of other ledgers.
        """
        lines = []
        lf_lines = []
        details = []

        this_ledger_dr = 0.0
        this_ledger_cr = 0.0

        for le in t.entries.all():
            full_name = le.ledger.name or ""
            first, sep, rest = full_name.partition(" ")
            display_name = rest if rest else full_name

            dr = float(le.dr_amount or 0)
            cr = float(le.cr_amount or 0)

            if le.ledger == ledger_obj:
                # this is the main ledger we are viewing
                this_ledger_dr += dr
                this_ledger_cr += cr
            else:
                # counter ledgers: show in lines/LF/details
                lines.append(display_name)
                if dr:
                    lf_lines.append((display_name, "Dr"))
                    details.append(dr)
                if cr:
                    lf_lines.append((display_name, "Cr"))
                    details.append(cr)

        pv = getattr(t, "paymentvendordetails", None)
        cheque_number = pv.cheque_no if pv else ""

        return {
            "date": t.voucher_date,
            "voucher_no": t.voucher_no,
            "cheque_no": cheque_number,
            "lines": lines,
            "lf_lines": lf_lines,
            "details": details,
            "this_dr": this_ledger_dr,
            "this_cr": this_ledger_cr,
            "narration": getattr(t, "narration", ""),
        }

    if selected_ledger:
        for t in qs.order_by("voucher_date", "id"):
            info = build_transaction_rows_for_ledger(t, selected_ledger)

            # left side (Dr)
            if info["this_dr"]:
                row = {
                    "date": info["date"],
                    "voucher_no": info["voucher_no"],
                    "cheque_no": info["cheque_no"],
                    "lines": info["lines"],
                    "lf_lines": info["lf_lines"],
                    "details": info["details"],
                    "amount": info["this_dr"],
                    "narration": info["narration"],
                }
                receipts_entries.append(row)
                total_receipts += info["this_dr"]

            # right side (Cr)
            if info["this_cr"]:
                row = {
                    "date": info["date"],
                    "voucher_no": info["voucher_no"],
                    "cheque_no": info["cheque_no"],
                    "lines": info["lines"],
                    "lf_lines": info["lf_lines"],
                    "details": info["details"],
                    "amount": info["this_cr"],
                    "narration": info["narration"],
                }
                payments_entries.append(row)
                total_payments += info["this_cr"]

    # Opening balance and closing only when a specific ledger is selected
    if selected_ledger:
        ledger_q = all_ledgers.filter(id=selected_ledger.id)

        opening_from_master = Decimal("0")
        for l in ledger_q:
            opening_from_master += _get_ledger_opening_signed(l, from_date)

        entries_before = TransactionEntry.objects.filter(
            transaction__ulb=current_ulb,
            transaction__voucher_date__lt=from_date,
            ledger__in=ledger_q,
        ).aggregate(
            dr=Sum("dr_amount"),
            cr=Sum("cr_amount"),
        )
        opening_dr_txn = Decimal(entries_before.get("dr") or 0)
        opening_cr_txn = Decimal(entries_before.get("cr") or 0)
        opening_from_txn = opening_dr_txn - opening_cr_txn

        opening_balance = opening_from_master + opening_from_txn
        opening_side = "Dr" if opening_balance >= 0 else "Cr"
        opening_amount_abs = abs(opening_balance)

        closing_balance = (
            opening_balance
            + Decimal(str(total_receipts))
            - Decimal(str(total_payments))
        )
        grand_total_left = opening_balance + Decimal(str(total_receipts))
        grand_total_right = Decimal(str(total_payments)) + closing_balance
    else:
        # no ledger selected: no opening/closing totals
        opening_balance = None
        opening_side = None
        opening_amount_abs = None
        closing_balance = None
        grand_total_left = None
        grand_total_right = None

    # ---------- Excel export for ledger ----------
    if export == "excel" and selected_ledger:
        wb = Workbook()
        ws = wb.active
        ws.title = "Ledger"

        blue_fill = PatternFill(fill_type="solid", fgColor="1E90FF")
        bold_center = Alignment(horizontal="center", vertical="center")

        def fmt_date(d):
            if not d:
                return ""
            return d.strftime("%d-%m-%Y")

        # Row 1: Title
        ws.merge_cells("A1:N1")
        ws["A1"].value = f"LEDGER: {selected_ledger.name}"
        ws["A1"].font = Font(size=14, bold=True)
        ws["A1"].alignment = bold_center

        # Row 2: ULB
        ws.merge_cells("A2:N2")
        ws["A2"].value = current_ulb.ulb_name if current_ulb else ""
        ws["A2"].font = Font(size=12, bold=True)
        ws["A2"].alignment = bold_center

        # Row 3: Period
        ws.merge_cells("A3:N3")
        ws["A3"].value = f"Period: {fmt_date(from_date)} to {fmt_date(to_date)}"
        ws["A3"].font = Font(size=11, bold=False)
        ws["A3"].alignment = bold_center

        # Row 4: Dr / Cr labels
        ws.merge_cells("A4:G4")
        ws["A4"].value = "DEBIT"
        ws["A4"].font = Font(bold=True, color="FFFFFF")
        ws["A4"].fill = blue_fill
        ws["A4"].alignment = bold_center

        ws.merge_cells("H4:N4")
        ws["H4"].value = "CREDIT"
        ws["H4"].font = Font(bold=True, color="FFFFFF")
        ws["H4"].fill = blue_fill
        ws["H4"].alignment = bold_center

        # Row 5: headers
        ws.append([
            "Date", "Voucher No.", "Cheque No.", "Particulars", "L/F", "Details", "Amount",
            "Date", "Voucher No.", "Cheque No.", "Particulars", "L/F", "Details", "Amount",
        ])

        # Row 6: Opening balance, placed based on side
        if opening_balance is not None:
            if opening_side == "Dr":
                ws.append([
                    fmt_date(from_date),
                    "Opening Balance",
                    "",
                    "",
                    "",
                    "",
                    float(opening_amount_abs),
                    "", "", "", "", "", "", "",
                ])
            else:
                # Cr opening: negative amount on left-grand and right-grand logic,
                # but here we follow your web style: show side on Cr, amount positive
                ws.append([
                    "", "", "", "", "", "", "",
                    fmt_date(from_date),
                    "Opening Balance",
                    "",
                    "",
                    "",
                    "",
                    float(opening_amount_abs),
                ])
        else:
            ws.append([""] * 14)

        # Data rows + narration
        max_len = max(len(receipts_entries), len(payments_entries))

        for i in range(max_len):
            recv = receipts_entries[i] if i < len(receipts_entries) else None
            pay = payments_entries[i] if i < len(payments_entries) else None

            recv_parts = "\n".join(recv["lines"]) if recv else ""
            recv_lfs = "\n".join(lf[1] for lf in recv["lf_lines"]) if recv else ""
            recv_details = "\n".join(str(d) for d in recv["details"]) if recv else ""

            pay_parts = "\n".join(pay["lines"]) if pay else ""
            pay_lfs = "\n".join(lf[1] for lf in pay["lf_lines"]) if pay else ""
            pay_details = "\n".join(str(d) for d in pay["details"]) if pay else ""

            ws.append([
                fmt_date(recv["date"]) if recv else "",
                recv["voucher_no"] if recv else "",
                recv["cheque_no"] if recv else "",
                recv_parts,
                recv_lfs,
                recv_details,
                float(recv["amount"]) if recv else "",
                fmt_date(pay["date"]) if pay else "",
                pay["voucher_no"] if pay else "",
                pay["cheque_no"] if pay else "",
                pay_parts,
                pay_lfs,
                pay_details,
                float(pay["amount"]) if pay else "",
            ])

            narr_left = f"Narration: {recv['narration']}" if recv and recv["narration"] else ""
            narr_right = f"Narration: {pay['narration']}" if pay and pay["narration"] else ""

            ws.append([
                "", "", "", narr_left, "", "", "",
                "", "", "", narr_right, "", "", "",
            ])

        # Blank row
        ws.append([])
        footer_start = ws.max_row + 1

        # Totals + Closing + Grand Totals
        if opening_balance is not None:
            # Total Dr
            ws.merge_cells(start_row=footer_start, start_column=1, end_row=footer_start, end_column=6)
            ws.cell(row=footer_start, column=1).value = "Total Debit"
            ws.cell(row=footer_start, column=1).font = Font(bold=True)
            ws.cell(row=footer_start, column=7).value = float(total_receipts)

            # Total Cr
            ws.merge_cells(start_row=footer_start, start_column=8, end_row=footer_start, end_column=13)
            ws.cell(row=footer_start, column=8).value = "Total Credit"
            ws.cell(row=footer_start, column=8).font = Font(bold=True)
            ws.cell(row=footer_start, column=14).value = float(total_payments)

            # Closing balance
            footer_row_cb = footer_start + 1
            if closing_balance >= 0:
                # Closing Dr (positive)
                ws.merge_cells(start_row=footer_row_cb, start_column=1, end_row=footer_row_cb, end_column=6)
                ws.cell(row=footer_row_cb, column=1).value = "Closing Balance (Dr)"
                ws.cell(row=footer_row_cb, column=1).font = Font(bold=True)
                ws.cell(row=footer_row_cb, column=7).value = float(closing_balance)
            else:
                # Closing Cr – export as negative amount, as on web page
                ws.merge_cells(start_row=footer_row_cb, start_column=8, end_row=footer_row_cb, end_column=13)
                ws.cell(row=footer_row_cb, column=8).value = "Closing Balance (Cr)"
                ws.cell(row=footer_row_cb, column=8).font = Font(bold=True)
                ws.cell(row=footer_row_cb, column=14).value = float(closing_balance)  # negative

            # Grand Totals rows – same as web page (left and right)
            footer_row_gt = footer_row_cb + 1

            ws.merge_cells(start_row=footer_row_gt, start_column=1, end_row=footer_row_gt, end_column=6)
            ws.cell(row=footer_row_gt, column=1).value = "Grand Total"
            ws.cell(row=footer_row_gt, column=1).font = Font(bold=True)
            ws.cell(row=footer_row_gt, column=7).value = float(grand_total_left)

            ws.merge_cells(start_row=footer_row_gt, start_column=8, end_row=footer_row_gt, end_column=13)
            ws.cell(row=footer_row_gt, column=8).value = "Grand Total"
            ws.cell(row=footer_row_gt, column=8).font = Font(bold=True)
            ws.cell(row=footer_row_gt, column=14).value = float(grand_total_right)

        # Auto-width
        for col_idx in range(1, 15):  # A..N
            col_letter = get_column_letter(col_idx)
            max_length = 0
            for r in range(1, ws.max_row + 1):
                cell = ws.cell(row=r, column=col_idx)
                if cell.value is not None:
                    length = len(str(cell.value))
                    if length > max_length:
                        max_length = length
            ws.column_dimensions[col_letter].width = (max_length + 2) * 1.2

        # Fixed width for D and K (Particulars columns)
        ws.column_dimensions["D"].width = 57
        ws.column_dimensions["K"].width = 57

        # Wrap text for particulars / details / narration columns
        wrap_columns = [4, 5, 6, 11, 12, 13]  # D,E,F and K,L,M
        for row in ws.iter_rows(min_row=6, max_row=ws.max_row, min_col=1, max_col=14):
            for cell in row:
                if cell.column in wrap_columns:
                    cell.alignment = Alignment(wrapText=True, vertical="top")

        # Borders
        thin = Side(style="thin", color="000000")
        border = Border(top=thin, left=thin, right=thin, bottom=thin)
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=14):
            for cell in row:
                cell.border = border

        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = 'attachment; filename="ledger.xlsx"'
        wb.save(response)
        return response

    # ---------- normal HTML render ----------
    sidebar_context = {
        "current_ulb": current_ulb,
        "current_ulb_name": current_ulb.ulb_name if current_ulb else None,
        "allowed_codes": get_allowed_codes_for(request.user),
        "active_section": "ledger",
    }

    context = {
        **sidebar_context,
        "all_ledgers": all_ledgers,
        "selected_ledger": selected_ledger,
        "from_date": from_date,
        "to_date": to_date,
        "opening_from_date": from_date,
        "opening_balance": opening_balance,
        "opening_side": opening_side,
        "opening_amount_abs": opening_amount_abs,
        "receipts_entries": receipts_entries,
        "payments_entries": payments_entries,
        "total_receipts": total_receipts,
        "total_payments": total_payments,
        "closing_balance": closing_balance,
        "grand_total_left": grand_total_left,
        "grand_total_right": grand_total_right,
    }
    return render(request, "accounts/ledger.html", context)
# ---------- GST report ----------
from accounts.models import ULB, PaymentVendorDetails, TransactionEntry
from datetime import date
from calendar import month_name
@login_required
@role_required(["ROOT_DEV", "DEV", "ADMIN", "USER"])
def gst_report(request):
    if not user_has_code(request.user, "MENU_ACCOUNTS_GST_REPORT"):
        return HttpResponseForbidden("You do not have permission for this action.")

    current_ulb_id = request.session.get("current_ulb_id")
    current_ulb = get_object_or_404(ULB, id=current_ulb_id) if current_ulb_id else None
    if not current_ulb:
        from django.shortcuts import redirect
        return redirect("ulb_select")

    # filters
    month = (request.GET.get("month") or "").strip()
    year = (request.GET.get("year") or "").strip()
    gst_number = (request.GET.get("gst_number") or "").strip()
    voucher_no = (request.GET.get("voucher_no") or "").strip()
    vendor_name = (request.GET.get("vendor_name") or "").strip()
    amount_str = (request.GET.get("amount") or "").strip()

    # user must select both month and year
    has_period = bool(month and year)

    # base queryset
    qs = PaymentVendorDetails.objects.none()

    if has_period:
        qs = PaymentVendorDetails.objects.filter(
            gst_applicable=True,
            transaction__ulb=current_ulb,
            transaction__voucher_date__year=int(year),
            transaction__voucher_date__month=int(month),
        ).select_related("transaction")

        if gst_number:
            qs = qs.filter(gst_no__icontains=gst_number)

        if voucher_no:
            qs = qs.filter(transaction__voucher_no__icontains=voucher_no)

        if vendor_name:
            qs = qs.filter(vendor_name__icontains=vendor_name)

        if amount_str:
            try:
                amt = Decimal(amount_str)
                qs = qs.filter(
                    Q(igst_amount=amt) | Q(cgst_amount=amt) | Q(sgst_amount=amt)
                )
            except Exception:
                pass

    # ledger 350200202 GST Deducted Payable
    gst_ledger_name = "350200202 GST Deducted Payable"

    if has_period:
        tx_ids = list(qs.values_list("transaction_id", flat=True))
    else:
        tx_ids = []

    if tx_ids:
        ledger_sums_by_tx = (
            TransactionEntry.objects.filter(
                transaction_id__in=tx_ids,
                ledger__name=gst_ledger_name,
            )
            .values("transaction_id")
            .annotate(ledger_gst_amount=Sum("cr_amount"))
        )
        ledger_map = {
            row["transaction_id"]: row["ledger_gst_amount"] or Decimal("0")
            for row in ledger_sums_by_tx
        }
    else:
        ledger_map = {}

    rows = list(qs.order_by("transaction__voucher_date", "transaction__id")) if has_period else []

    for pvd in rows:
        igst = pvd.igst_amount or Decimal("0")
        cgst = pvd.cgst_amount or Decimal("0")
        sgst = pvd.sgst_amount or Decimal("0")
        total_amt = igst + cgst + sgst

        ledger_amt = ledger_map.get(pvd.transaction_id, Decimal("0"))
        pvd.transaction.ledger_gst_amount = ledger_amt
        pvd.total_gst_amount = total_amt
        pvd.diff_amount = total_amt - ledger_amt

    if has_period:
        totals = qs.aggregate(
            total_igst=Sum("igst_amount"),
            total_cgst=Sum("cgst_amount"),
            total_sgst=Sum("sgst_amount"),
        )
        total_igst = totals.get("total_igst") or Decimal("0")
        total_cgst = totals.get("total_cgst") or Decimal("0")
        total_sgst = totals.get("total_sgst") or Decimal("0")
        total_ledger_gst = sum(
            (pvd.transaction.ledger_gst_amount or Decimal("0")) for pvd in rows
        )
        total_gst = total_igst + total_cgst + total_sgst
        total_diff = total_gst - total_ledger_gst
    else:
        total_igst = total_cgst = total_sgst = Decimal("0")
        total_ledger_gst = total_gst = total_diff = Decimal("0")

    # Excel export only when month & year selected
    if has_period and request.GET.get("export") == "1":
        wb = Workbook()
        ws = wb.active
        ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
        ws.merge_cells("A1:L1")
        ws.title = "GSTR-7"

        ws["A1"] = "GSTR-7"
        ws.merge_cells("A2:L2")
        ws["A2"].alignment = Alignment(horizontal="center", vertical="center")
        ws["A2"] = current_ulb.ulb_name if current_ulb else ""

        try:
            month_label = f"{month_name[int(month)]} {year}"
        except Exception:
            month_label = str(year) if year else ""
        ws.merge_cells("A3:L3")
        ws["A3"] = month_label
        ws["A3"].alignment = Alignment(horizontal="center", vertical="center")

        headers = [
            "Date",
            "Voucher no.",
            "Vendor name",
            "GST number",
            "GST type",
            "GST rate",
            "IGST amount",
            "CGST amount",
            "SGST amount",
            "Total amount",
            "As per ledger amount",
            "Difference",
        ]
        header_row = 4
        for col, title in enumerate(headers, start=1):
            ws.cell(row=header_row, column=col, value=title)

        data_row = header_row + 1
        for row in rows:
            ws.cell(row=data_row, column=1, value=row.transaction.voucher_date.strftime("%d-%m-%Y"))
            ws.cell(row=data_row, column=2, value=row.transaction.voucher_no)
            ws.cell(row=data_row, column=3, value=row.vendor_name)
            ws.cell(row=data_row, column=4, value=row.gst_no)
            ws.cell(
                row=data_row,
                column=5,
                value=(
                    "Inter State"
                    if row.gst_type == "inter"
                    else "Intra State"
                    if row.gst_type == "intra"
                    else ""
                ),
            )
            ws.cell(row=data_row, column=6, value=float(row.gst_rate or 0))
            ws.cell(row=data_row, column=7, value=float(row.igst_amount or 0))
            ws.cell(row=data_row, column=8, value=float(row.cgst_amount or 0))
            ws.cell(row=data_row, column=9, value=float(row.sgst_amount or 0))
            ws.cell(row=data_row, column=10, value=float(row.total_gst_amount or 0))
            ws.cell(
                row=data_row,
                column=11,
                value=float(row.transaction.ledger_gst_amount or 0),
            )
            ws.cell(row=data_row, column=12, value=float(row.diff_amount or 0))
            data_row += 1

        ws.cell(row=data_row, column=1, value="Total amount")
        ws.merge_cells(start_row=data_row, start_column=1, end_row=data_row, end_column=6)
        ws.cell(row=data_row, column=7, value=float(total_igst))
        ws.cell(row=data_row, column=8, value=float(total_cgst))
        ws.cell(row=data_row, column=9, value=float(total_sgst))
        ws.cell(row=data_row, column=10, value=float(total_gst))
        ws.cell(row=data_row, column=11, value=float(total_ledger_gst))
        ws.cell(row=data_row, column=12, value=float(total_diff))

        for col in range(1, len(headers) + 1):
            letter = get_column_letter(col)
            ws.column_dimensions[letter].auto_size = True

        filename = f"GSTR-7_{date.today().strftime('%Y%m%d')}.xlsx"
        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = f'attachment; filename=\"{filename}\"'
        wb.save(response)
        return response

    sidebar_context = {
        "current_ulb": current_ulb,
        "current_ulb_name": current_ulb.ulb_name if current_ulb else None,
        "allowed_codes": get_allowed_codes_for(request.user),
        "active_section": "gst_report",
    }

    context = {
        **sidebar_context,
        "gst_rows": rows,
        "total_igst": total_igst,
        "total_cgst": total_cgst,
        "total_sgst": total_sgst,
        "total_ledger_gst": total_ledger_gst,
        "total_gst": total_gst,
        "total_diff": total_diff,
        "has_period": has_period,
    }
    return render(request, "accounts/gst_report.html", context)

# --------- TDS REPORT ----------
@login_required
@role_required(["ROOT_DEV", "DEV", "ADMIN", "USER"])
def tds_report(request):
    if not user_has_code(request.user, "MENU_ACCOUNTS_TDS_REPORT"):
        return HttpResponseForbidden("You do not have permission for this action.")

    current_ulb_id = request.session.get("current_ulb_id")
    current_ulb = get_object_or_404(ULB, id=current_ulb_id) if current_ulb_id else None
    if not current_ulb:
        from django.shortcuts import redirect
        return redirect("ulb_select")

    # filters
    month = (request.GET.get("month") or "").strip()
    year = (request.GET.get("year") or "").strip()
    pan_number = (request.GET.get("pan_number") or "").strip()
    voucher_no = (request.GET.get("voucher_no") or "").strip()
    vendor_name = (request.GET.get("vendor_name") or "").strip()
    amount_str = (request.GET.get("amount") or "").strip()

    has_period = bool(month and year)

    # base queryset
    qs = PaymentVendorDetails.objects.none()

    if has_period:
        qs = PaymentVendorDetails.objects.filter(
            tds_applicable=True,
            transaction__ulb=current_ulb,
            transaction__voucher_type="PYMT",
            transaction__voucher_date__year=int(year),
            transaction__voucher_date__month=int(month),
        ).select_related("transaction")

        if pan_number:
            qs = qs.filter(tds_pan_no__icontains=pan_number)

        if voucher_no:
            qs = qs.filter(transaction__voucher_no__icontains=voucher_no)

        if vendor_name:
            qs = qs.filter(vendor_name__icontains=vendor_name)

        if amount_str:
            try:
                amt = Decimal(amount_str)
                qs = qs.filter(tds_amount=amt)
            except Exception:
                pass

    # ledger 350200201 Income Tax Deducted from Contractors/ Suppliers
    tds_ledger_name = "350200201 Income Tax Deducted from Contractors/ Suppliers"

    if has_period:
        tx_ids = list(qs.values_list("transaction_id", flat=True))
    else:
        tx_ids = []

    if tx_ids:
        ledger_sums_by_tx = (
            TransactionEntry.objects.filter(
                transaction_id__in=tx_ids,
                ledger__name=tds_ledger_name,
            )
            .values("transaction_id")
            .annotate(ledger_tds_amount=Sum("cr_amount"))
        )
        ledger_map = {
            row["transaction_id"]: row["ledger_tds_amount"] or Decimal("0")
            for row in ledger_sums_by_tx
        }
    else:
        ledger_map = {}

    # build rows as dicts (for template + excel)
    rows = []
    if has_period:
        for pvd in qs.order_by("transaction__voucher_date", "transaction__id"):
            tds_amount = pvd.tds_amount or Decimal("0")
            ledger_amt = ledger_map.get(pvd.transaction_id, Decimal("0"))
            diff_amount = tds_amount - ledger_amt

            rows.append({
                "transaction": pvd.transaction,
                "vendor_name": pvd.vendor_name,
                "pan_number": pvd.tds_pan_no,
                "tds_section": pvd.tds_section,
                "nature": pvd.tds_nature,
                "type": pvd.tds_type,
                "rate": pvd.tds_rate,
                "tds_amount": tds_amount,
                "ledger_amount": ledger_amt,
                "diff_amount": diff_amount,
            })

    if has_period:
        totals = qs.aggregate(
            total_tds=Sum("tds_amount"),
        )
        total_tds = totals.get("total_tds") or Decimal("0")
        total_ledger_tds = sum((row["ledger_amount"] or Decimal("0")) for row in rows)
        total_diff = total_tds - total_ledger_tds
    else:
        total_tds = total_ledger_tds = total_diff = Decimal("0")

    # Excel export only when month & year selected
    if has_period and request.GET.get("export") == "1":
        wb = Workbook()
        ws = wb.active
        ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
        ws.merge_cells("A1:K1")
        ws.title = "TDS Report"

        ws["A1"] = "TDS REPORT"
        ws.merge_cells("A2:K2")
        ws["A2"].alignment = Alignment(horizontal="center", vertical="center")
        ws["A2"] = current_ulb.ulb_name if current_ulb else ""

        try:
            month_label = f"{month_name[int(month)]} {year}"
        except Exception:
            month_label = str(year) if year else ""
        ws.merge_cells("A3:K3")
        ws["A3"] = month_label
        ws["A3"].alignment = Alignment(horizontal="center", vertical="center")

        headers = [
            "Date",
            "Voucher number",
            "Vendor name",
            "PAN number",
            "TDS section",
            "Nature",
            "Type",
            "Rate",
            "TDS amount",
            "As per ledger amount (To 350200201 Income Tax Deducted from Contractors/ Suppliers)",
            "Difference",
        ]
        header_row = 4
        for col, title in enumerate(headers, start=1):
            ws.cell(row=header_row, column=col, value=title)

        data_row = header_row + 1
        for row in rows:
            trx = row["transaction"]
            ws.cell(
                row=data_row,
                column=1,
                value=trx.voucher_date.strftime("%d-%m-%Y"),
            )
            ws.cell(row=data_row, column=2, value=trx.voucher_no)
            ws.cell(row=data_row, column=3, value=row["vendor_name"])
            ws.cell(row=data_row, column=4, value=row["pan_number"])
            ws.cell(row=data_row, column=5, value=row["tds_section"])
            ws.cell(row=data_row, column=6, value=row["nature"])
            ws.cell(row=data_row, column=7, value=row["type"])
            ws.cell(row=data_row, column=8, value=row["rate"])
            ws.cell(row=data_row, column=9, value=float(row["tds_amount"] or 0))
            ws.cell(
                row=data_row,
                column=10,
                value=float(row["ledger_amount"] or 0),
            )
            ws.cell(row=data_row, column=11, value=float(row["diff_amount"] or 0))
            data_row += 1

        ws.cell(row=data_row, column=1, value="Total amount")
        ws.merge_cells(start_row=data_row, start_column=1, end_row=data_row, end_column=8)
        ws.cell(row=data_row, column=9, value=float(total_tds))
        ws.cell(row=data_row, column=10, value=float(total_ledger_tds))
        ws.cell(row=data_row, column=11, value=float(total_diff))

        for col in range(1, len(headers) + 1):
            letter = get_column_letter(col)
            ws.column_dimensions[letter].auto_size = True

        filename = f"TDS_Report_{date.today().strftime('%Y%m%d')}.xlsx"
        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        wb.save(response)
        return response

    sidebar_context = {
        "current_ulb": current_ulb,
        "current_ulb_name": current_ulb.ulb_name if current_ulb else None,
        "allowed_codes": get_allowed_codes_for(request.user),
        "active_section": "tds_report",
    }

    context = {
        **sidebar_context,
        "tds_rows": rows,
        "total_tds": total_tds,
        "total_ledger_tds": total_ledger_tds,
        "total_diff": total_diff,
        "has_period": has_period,
    }
    return render(request, "accounts/tds_report.html", context)

@login_required
@role_required(["ROOT_DEV", "DEV", "ADMIN", "USER"])
def accounts_brs(request):
    """
    BRS main screen: only checks MENU_ACCOUNTS_BRS.
    Shows header with three tabs, no Bank Entry logic here.
    """
    allowed_codes = get_allowed_codes_for(request.user)

    # IMPORTANT: only this code is checked here
    if "MENU_ACCOUNTS_BRS" not in allowed_codes:
        return HttpResponseForbidden("You do not have permission for this action.")

    current_ulb_id = request.session.get("current_ulb_id")
    current_ulb = get_object_or_404(ULB, id=current_ulb_id) if current_ulb_id else None
    if not current_ulb:
        return redirect("ulb_select")

    sidebar_context = {
        "current_ulb": current_ulb,
        "current_ulb_name": current_ulb.ulb_name if current_ulb else None,
        "allowed_codes": allowed_codes,
        "active_section": "brs",
    }

    context = {
        **sidebar_context,
        # no form data, just header + empty content
    }
    return render(request, "accounts/base_brs.html", context)



@login_required
@role_required(["ROOT_DEV", "DEV", "ADMIN", "USER"])
def bank_entry(request):
    """
    BRS screen:
    - GET: show ledger dropdown (only ledgers starting with 450) and, if selected, the entry form
    - POST: save multiple BRS lines from rows[*] into BankReconciliationEntry.
            If brs_id is present, update that entry instead of creating new.
    """

    # Strict per‑tab permission for Bank Entry
    if not user_has_code(request.user, "MENU_ACCOUNTS_BRS_BANK_ENTRY"):
        return HttpResponseForbidden("You do not have permission for this action.")

    current_ulb_id = request.session.get("current_ulb_id")
    current_ulb = get_object_or_404(ULB, id=current_ulb_id) if current_ulb_id else None
    if not current_ulb:
        return redirect("ulb_select")

    # ledgers that start with 450 for this ULB
    ledgers = (
        Ledger.objects.filter(ulb=current_ulb, name__startswith="450")
        .order_by("name")
    )

    selected_ledger = None
    editing_entry = None  # entry being edited (if any)

    if request.method == "POST":
        ledger_id = request.POST.get("ledger_id") or ""
        brs_id = request.POST.get("brs_id") or ""

        if ledger_id:
            selected_ledger = get_object_or_404(Ledger, id=ledger_id, ulb=current_ulb)
        else:
            return redirect(request.path)

        # if brs_id present => single-row update
        if brs_id:
            editing_entry = get_object_or_404(
                BankReconciliationEntry,
                id=brs_id,
                ulb=current_ulb,
                ledger=selected_ledger,
            )

            etype = (request.POST.get("entry_type") or "").strip()
            date_str = (request.POST.get("entry_date") or "").strip()
            cheque_number = (request.POST.get("cheque_number") or "").strip()
            narration = (request.POST.get("bank_narration") or "").strip()
            dr_str = (request.POST.get("dr_amount") or "").replace(",", "").strip()
            cr_str = (request.POST.get("cr_amount") or "").replace(",", "").strip()

            # basic validation
            try:
                entry_date = date.fromisoformat(date_str)
            except Exception:
                entry_date = None

            try:
                dr_amount = Decimal(dr_str) if dr_str else Decimal("0")
            except Exception:
                dr_amount = Decimal("0")

            try:
                cr_amount = Decimal(cr_str) if cr_str else Decimal("0")
            except Exception:
                cr_amount = Decimal("0")

            if etype == "Dr":
                cr_amount = Decimal("0")
            elif etype == "Cr":
                dr_amount = Decimal("0")

            if entry_date and etype in ("Dr", "Cr"):
                editing_entry.entry_type = etype
                editing_entry.entry_date = entry_date
                editing_entry.cheque_number = cheque_number
                editing_entry.bank_narration = narration
                editing_entry.dr_amount = dr_amount
                editing_entry.cr_amount = cr_amount
                editing_entry.save()

            # after update, go back with ledger selected
            return redirect(f"{request.path}?ledger_id={ledger_id}")

        # else: normal multi-row create
        rows_idx = []
        prefix = "rows["
        for key in request.POST.keys():
            if key.startswith(prefix) and key.endswith("][entry_type]"):
                idx = key[len(prefix):].split("]")[0]
                rows_idx.append(idx)

        rows_idx = sorted(set(rows_idx), key=int)

        entries_to_create = []
        for idx in rows_idx:
            etype = (request.POST.get(f"rows[{idx}][entry_type]") or "").strip()
            date_str = (request.POST.get(f"rows[{idx}][entry_date]") or "").strip()
            cheque_number = (request.POST.get(f"rows[{idx}][cheque_number]") or "").strip()
            narration = (request.POST.get(f"rows[{idx}][bank_narration]") or "").strip()
            dr_str = (request.POST.get(f"rows[{idx}][dr_amount]") or "").replace(",", "").strip()
            cr_str = (request.POST.get(f"rows[{idx}][cr_amount]") or "").replace(",", "").strip()

            if not etype or not date_str:
                continue

            try:
                entry_date = date.fromisoformat(date_str)
            except Exception:
                continue

            try:
                dr_amount = Decimal(dr_str) if dr_str else Decimal("0")
            except Exception:
                dr_amount = Decimal("0")

            try:
                cr_amount = Decimal(cr_str) if cr_str else Decimal("0")
            except Exception:
                cr_amount = Decimal("0")

            if etype == "Dr":
                cr_amount = Decimal("0")
            elif etype == "Cr":
                dr_amount = Decimal("0")

            entries_to_create.append(
                BankReconciliationEntry(
                    ulb=current_ulb,
                    ledger=selected_ledger,
                    entry_type=etype,
                    entry_date=entry_date,
                    cheque_number=cheque_number,
                    bank_narration=narration,
                    dr_amount=dr_amount,
                    cr_amount=cr_amount,
                    created_by=request.user,
                )
            )

        if entries_to_create:
            BankReconciliationEntry.objects.bulk_create(entries_to_create)

        return redirect(f"{request.path}?ledger_id={ledger_id}")

    else:
        # GET
        ledger_id = request.GET.get("ledger_id") or ""
        brs_id = request.GET.get("brs_id") or ""
        if ledger_id:
            selected_ledger = get_object_or_404(Ledger, id=ledger_id, ulb=current_ulb)

        if brs_id and selected_ledger:
            editing_entry = get_object_or_404(
                BankReconciliationEntry,
                id=brs_id,
                ulb=current_ulb,
                ledger=selected_ledger,
            )

    allowed_codes = get_allowed_codes_for(request.user)

    sidebar_context = {
        "current_ulb": current_ulb,
        "current_ulb_name": current_ulb.ulb_name if current_ulb else None,
        "allowed_codes": allowed_codes,
        "active_section": "brs",  # highlight BRS in sidebar
    }

    context = {
        **sidebar_context,
        "ledgers": ledgers,
        "selected_ledger": selected_ledger,
        "editing_entry": editing_entry,
    }
    return render(request, "accounts/bank_entry.html", context)

# ------------------Show Bank Entry with running balance------------------
from django.db.models import Sum, Value
from accounts.models import ULB, Ledger, BankReconciliationEntry
from django.db.models import Sum, F, Value, DecimalField
from django.db.models.functions import Coalesce
from django.urls import reverse
def _get_current_ulb(request):
    current_ulb_id = request.session.get("current_ulb_id")
    if not current_ulb_id:
        return None
    return get_object_or_404(ULB, id=current_ulb_id)


def _get_default_fy_dates():
    """Indian FY: 1 April to 31 March."""
    today = date.today()
    year = today.year
    if today.month < 4:
        start_year = year - 1
    else:
        start_year = year
    fy_start = date(start_year, 4, 1)
    fy_end = today
    return fy_start, fy_end


@login_required
@role_required(["ROOT_DEV", "DEV", "ADMIN", "USER"])
def show_bank_entry_view(request):
    # Access check for this tab only
    if not user_has_code(request.user, "MENU_ACCOUNTS_BRS_SHOW_BANK_ENTRY"):
        return HttpResponseForbidden("You do not have permission for this action.")

    current_ulb = _get_current_ulb(request)
    if not current_ulb:
        return redirect("ulb_select")

    ledgers = (
        Ledger.objects.filter(ulb=current_ulb, name__startswith="450")
        .order_by("name")
    )

    # default FY dates helper you already have
    default_from, default_to = _get_default_fy_dates()

    # filters
    if request.method == "POST":
        ledger_id = request.POST.get("ledger_id") or ""
        from_date_str = request.POST.get("from_date") or ""
        to_date_str = request.POST.get("to_date") or ""
    else:
        ledger_id = request.GET.get("ledger_id") or ""
        from_date_str = request.GET.get("from_date") or ""
        to_date_str = request.GET.get("to_date") or ""

    def parse_date_or_default(value, default):
        if not value:
            return default
        try:
            return datetime.strptime(value, "%Y-%m-%d").date()
        except ValueError:
            return default

    from_date = parse_date_or_default(from_date_str, default_from)
    to_date = parse_date_or_default(to_date_str, default_to)

    selected_ledger = None
    entries = []
    opening_balance = Decimal("0.00")

    # delete action inside same view
    if request.method == "POST":
        action = request.POST.get("action") or ""
        if action == "delete":
            entry_id = request.POST.get("entry_id")
            if entry_id:
                obj = get_object_or_404(
                    BankReconciliationEntry,
                    id=entry_id,
                    ulb=current_ulb,
                )
                obj.delete()

            url = reverse("accounts_brs_show_bank_entry")
            params = []
            if ledger_id:
                params.append(f"ledger_id={ledger_id}")
            if from_date:
                params.append(f"from_date={from_date.strftime('%Y-%m-%d')}")
            if to_date:
                params.append(f"to_date={to_date.strftime('%Y-%m-%d')}")
            if params:
                url = f"{url}?{'&'.join(params)}"
            return redirect(url)

    if ledger_id:
        selected_ledger = get_object_or_404(Ledger, id=ledger_id, ulb=current_ulb)

        # opening balance before from_date
        opening_qs = BankReconciliationEntry.objects.filter(
            ulb=current_ulb,
            ledger=selected_ledger,
            entry_date__lt=from_date,
        )
        agg = opening_qs.aggregate(
            total_dr=Sum("dr_amount"),
            total_cr=Sum("cr_amount"),
        )
        sum_dr = agg["total_dr"] or Decimal("0.00")
        sum_cr = agg["total_cr"] or Decimal("0.00")
        opening_balance = sum_dr - sum_cr

        qs = (
            BankReconciliationEntry.objects
            .filter(
                ulb=current_ulb,
                ledger=selected_ledger,
                entry_date__gte=from_date,
                entry_date__lte=to_date,
            )
            .order_by("entry_date", "id")
        )

        running = opening_balance
        for obj in qs:
            running = running + obj.dr_amount - obj.cr_amount
            obj.running_balance = running
            entries.append(obj)

    allowed_codes = get_allowed_codes_for(request.user)

    sidebar_context = {
        "current_ulb": current_ulb,
        "current_ulb_name": current_ulb.ulb_name if current_ulb else None,
        "allowed_codes": allowed_codes,
        "active_section": "brs",
    }

    context = {
        **sidebar_context,
        "ledgers": ledgers,
        "selected_ledger": selected_ledger,
        "entries": entries,
        "from_date": from_date,
        "to_date": to_date,
        "from_date_display": from_date.strftime("%d-%m-%Y"),
        "to_date_display": to_date.strftime("%d-%m-%Y"),
        "opening_balance": opening_balance,
    }
    return render(request, "accounts/show_bank_entry.html", context)

from .models import BankReconciliationEntry, TransactionEntry, ULB, Ledger
@login_required
@role_required(["ROOT_DEV", "DEV", "ADMIN", "USER"])
def brs_adjustment(request):
    """
    BRS Adjustment – Cash Receipts / Payments.

    Left:
      - Cash Book entries (both receipts and payments) for selected bank ledger.
      - Bank statement entries (both receipts and payments) for selected bank ledger.

    Right:
      - Selected rows from both sides.
      - Total Dr - Total Cr must be zero to allow save.
    """

    # --------- access code check ----------
    if not user_has_code(request.user, "MENU_ACCOUNTS_BRS_ADJUSTMENT"):
        return HttpResponseForbidden("You do not have permission for this action.")

    # --------- current ULB check ----------
    current_ulb_id = request.session.get("current_ulb_id")
    current_ulb = get_object_or_404(ULB, id=current_ulb_id) if current_ulb_id else None
    if not current_ulb:
        return redirect("ulb_select")

    # 1) Ledgers list: bank ledgers (450...) for this ULB
    ledgers = (
        Ledger.objects.filter(
            ulb=current_ulb,
            name__startswith="450",  # adjust if you have a separate ledger code
        )
        .order_by("name")
    )

    selected_ledger = None

    # we will expose both receipts + payments; template decides using request.GET.entry_type
    cash_entries_all = TransactionEntry.objects.none()
    bank_entries_all = BankReconciliationEntry.objects.none()

    ledger_id = request.GET.get("ledger_id") or request.POST.get("ledger_id")

    if ledger_id:
        selected_ledger = get_object_or_404(Ledger, pk=ledger_id, ulb=current_ulb)

        # 2) Cash Book entries for this ledger:
        #    - Include ALL Dr and Cr entries for this bank ledger, regardless of voucher_type
        #      (Receipts tab will show Dr, Payments tab will show Cr, including Contra).
        cash_entries_all = (
            TransactionEntry.objects.select_related("transaction", "ledger")
            .filter(
                transaction__ulb=current_ulb,
                ledger=selected_ledger,
                is_brs_reconciled=False,
            )
            .filter(
                Q(entry_type="Dr") | Q(entry_type="Cr")
            )
            .annotate(
                cheque_number=Coalesce(
                    F("transaction__paymentvendordetails__cheque_no"),
                    Value(""),
                )
            )
            .order_by("transaction__voucher_date", "id")
        )

        # 3) Bank statement entries (both receipts and payments) for this ledger
        bank_entries_all = (
            BankReconciliationEntry.objects.filter(
                ulb=current_ulb,
                ledger=selected_ledger,
                is_reconciled=False,
            )
            .filter(Q(entry_type="Cr") | Q(entry_type="Dr"))
            .order_by("entry_date", "id")
        )

    # ---------- POST: save reconciliation ----------
    if request.method == "POST" and selected_ledger:
        selections = []
        idx = 0
        while True:
            src = request.POST.get(f"selected[{idx}][source]")
            if src is None:
                break

            entry_id = request.POST.get(f"selected[{idx}][id]")
            amount_str = request.POST.get(f"selected[{idx}][amount]", "0")
            sign = request.POST.get(f"selected[{idx}][sign]")  # 'dr' or 'cr'

            try:
                amount = Decimal(str(amount_str).replace(",", ""))
            except Exception:
                amount = Decimal("0")

            selections.append(
                {
                    "source": src,   # 'cash' or 'bank'
                    "id": entry_id,
                    "amount": amount,
                    "sign": sign,    # 'dr' or 'cr'
                }
            )
            idx += 1

        if not selections:
            messages.error(request, "No entries selected for reconciliation.")
            return redirect(
                f"{reverse('accounts_brs_adjustment')}?ledger_id={selected_ledger.id}"
            )

        total_dr = sum(s["amount"] for s in selections if s["sign"] == "dr")
        total_cr = sum(s["amount"] for s in selections if s["sign"] == "cr")
        diff = total_dr - total_cr

        if diff != Decimal("0"):
            messages.error(
                request,
                f"Cannot save: Dr ({total_dr}) − Cr ({total_cr}) = {diff}, must be 0.",
            )
            return redirect(
                f"{reverse('accounts_brs_adjustment')}?ledger_id={selected_ledger.id}"
            )

        # Save reconciliation: mark entries as reconciled and assign group id.
        with transaction.atomic():
            # 1) Generate a new group id for this adjustment
            last_adj_id = (
                TransactionEntry.objects.aggregate(
                    m=Max("brs_adjustment_id")
                )["m"]
                or 0
            )
            new_adj_id = last_adj_id + 1

            # 2) Mark and tag entries
            for s in selections:
                if s["source"] == "cash":
                    cash_obj = get_object_or_404(
                        TransactionEntry,
                        pk=s["id"],
                        transaction__ulb=current_ulb,
                        ledger=selected_ledger,
                        is_brs_reconciled=False,
                    )
                    cash_obj.is_brs_reconciled = True
                    cash_obj.brs_adjustment_id = new_adj_id
                    cash_obj.save(
                        update_fields=["is_brs_reconciled", "brs_adjustment_id"]
                    )

                elif s["source"] == "bank":
                    bank_obj = get_object_or_404(
                        BankReconciliationEntry,
                        pk=s["id"],
                        ulb=current_ulb,
                        ledger=selected_ledger,
                        is_reconciled=False,
                    )
                    bank_obj.is_reconciled = True
                    bank_obj.brs_adjustment_id = new_adj_id
                    bank_obj.save(
                        update_fields=["is_reconciled", "brs_adjustment_id"]
                    )

        messages.success(
            request,
            f"BRS reconciliation saved for ledger {selected_ledger.name}.",
        )
        return redirect(f"{reverse('accounts_brs_adjustment')}?ledger_id={selected_ledger.id}")

    # ---------- GET / initial render ----------
    context = {
        "current_ulb": current_ulb,
        "current_ulb_name": current_ulb.ulb_name if current_ulb else None,
        "allowed_codes": get_allowed_codes_for(request.user),
        "ledgers": ledgers,
        "selected_ledger": selected_ledger,
        "cash_entries_all": cash_entries_all,
        "bank_entries_all": bank_entries_all,
        "active_section": "brs",
    }
    return render(request, "accounts/brs_adjustment.html", context)
from collections import defaultdict
@login_required
@role_required(["ROOT_DEV", "DEV", "ADMIN", "USER"])
def show_brs_adjustment(request):
    """
    Report page: show grouped BRS adjustments (cash-book + bank) for a selected
    bank ledger and type. Each adjustment shows all cash lines, then all bank
    lines, then a difference row (cash - bank).
    """

    if not user_has_code(request.user, "MENU_ACCOUNTS_SHOW_BRS_ADJUSTMENT"):
        return HttpResponseForbidden("You do not have permission for this action.")

    current_ulb_id = request.session.get("current_ulb_id")
    current_ulb = get_object_or_404(ULB, id=current_ulb_id) if current_ulb_id else None
    if not current_ulb:
        return redirect("ulb_select")

    ledgers = (
        Ledger.objects.filter(
            ulb=current_ulb,
            name__startswith="450",
        )
        .order_by("name")
    )

    selected_ledger = None
    adjustments = []  # list of grouped adjustments to show
    changed_adj_ids = set()  # to track auto‑undone groups

    ledger_id = request.GET.get("ledger_id") or None
    entry_type_param = request.GET.get("entry_type") or None  # 'receipts' / 'payments'

    if ledger_id and entry_type_param:
        selected_ledger = get_object_or_404(Ledger, pk=ledger_id, ulb=current_ulb)

        # Map UI type -> Dr/Cr sign for each side
        if entry_type_param == "receipts":
            cash_sign = "dr"   # cash-book receipts (Dr)
            bank_sign = "cr"   # bank receipts (Cr)
        else:  # "payments"
            cash_sign = "cr"   # cash-book payments (Cr)
            bank_sign = "dr"   # bank payments (Dr)

        # ---------------- AUTO-UNDO: detect changed records ----------------
        # Base cash queryset (only reconciled + with adjustment id)
        cash_qs_base = (
            TransactionEntry.objects.select_related("transaction", "ledger")
            .filter(
                transaction__ulb=current_ulb,
                ledger=selected_ledger,
                is_brs_reconciled=True,
                brs_adjustment_id__isnull=False,
            )
        )
        if cash_sign == "dr":
            cash_qs_base = cash_qs_base.filter(dr_amount__gt=0)
        else:
            cash_qs_base = cash_qs_base.filter(cr_amount__gt=0)

        changed_cash_adj_ids = set()
        for e in cash_qs_base:
            # snapshot of current state
            if cash_sign == "dr":
                amt = e.dr_amount
            else:
                amt = e.cr_amount
            current_snapshot = f"{e.transaction.voucher_date}|{e.transaction.narration or ''}|{amt}"

            # compare with stored snapshot (if any)
            if e.brs_cash_particulars and e.brs_cash_particulars != current_snapshot:
                changed_cash_adj_ids.add(e.brs_adjustment_id)

        # Base bank queryset
        bank_qs_base = BankReconciliationEntry.objects.filter(
            ulb=current_ulb,
            ledger=selected_ledger,
            is_reconciled=True,
            brs_adjustment_id__isnull=False,
        )
        if bank_sign == "dr":
            bank_qs_base = bank_qs_base.filter(dr_amount__gt=0)
        else:
            bank_qs_base = bank_qs_base.filter(cr_amount__gt=0)

        changed_bank_adj_ids = set()
        for b in bank_qs_base:
            if bank_sign == "dr":
                amt = b.dr_amount
            else:
                amt = b.cr_amount
            current_snapshot = f"{b.entry_date}|{b.bank_narration or ''}|{b.cheque_number or ''}|{amt}"

            if b.brs_bank_particulars and b.brs_bank_particulars != current_snapshot:
                changed_bank_adj_ids.add(b.brs_adjustment_id)

        # Any adjustment ids that changed on either side
        changed_adj_ids = changed_cash_adj_ids.union(changed_bank_adj_ids)

        # Auto‑undo these groups for both cash and bank
        if changed_adj_ids:
            TransactionEntry.objects.filter(
                brs_adjustment_id__in=changed_adj_ids
            ).update(
                is_brs_reconciled=False,
                brs_adjustment_id=None,
                brs_cash_particulars="",
            )
            BankReconciliationEntry.objects.filter(
                brs_adjustment_id__in=changed_adj_ids
            ).update(
                is_reconciled=False,
                brs_adjustment_id=None,
                brs_bank_particulars="",
            )

        # ---------------- Cash-book side (clean after auto-undo) ----------------
        cash_qs = (
            TransactionEntry.objects.select_related("transaction", "ledger")
            .filter(
                transaction__ulb=current_ulb,
                ledger=selected_ledger,
                is_brs_reconciled=True,
                brs_adjustment_id__isnull=False,
            )
            .annotate(
                cheque_number=Coalesce(
                    F("transaction__paymentvendordetails__cheque_no"),
                    Value("")
                )
            )
        )

        if cash_sign == "dr":
            cash_qs = cash_qs.filter(dr_amount__gt=0)
        else:
            cash_qs = cash_qs.filter(cr_amount__gt=0)

        # Group by brs_adjustment_id (group id created in brs_adjustment)
        groups = defaultdict(lambda: {"cash": [], "bank": []})

        for e in cash_qs.order_by("transaction__voucher_date", "id"):
            if cash_sign == "dr":
                amt = e.dr_amount
                etype = "Cash Dr"  # Display label
            else:
                amt = e.cr_amount
                etype = "Cash Cr"

            # store fresh snapshot for next time
            snapshot = f"{e.transaction.voucher_date}|{e.transaction.narration or ''}|{amt}"
            if e.brs_cash_particulars != snapshot:
                e.brs_cash_particulars = snapshot
                e.save(update_fields=["brs_cash_particulars"])

            adj_id = e.brs_adjustment_id
            groups[adj_id]["cash"].append(
                {
                    "entry_date": e.transaction.voucher_date,
                    "entry_type": etype,
                    "cheque_no": e.cheque_number or "",  # from PaymentVendorDetails
                    "narration": e.transaction.narration or "",
                    "amount": amt,
                    "source": "Cash Book",
                    "has_action": False,  # will be set later
                }
            )

        # ---------------- Bank side (clean after auto-undo) ----------------
        bank_qs = BankReconciliationEntry.objects.filter(
            ulb=current_ulb,
            ledger=selected_ledger,
            is_reconciled=True,
            brs_adjustment_id__isnull=False,
        )

        if bank_sign == "dr":
            bank_qs = bank_qs.filter(dr_amount__gt=0)
        else:
            bank_qs = bank_qs.filter(cr_amount__gt=0)

        for b in bank_qs.order_by("entry_date", "id"):
            if bank_sign == "dr":
                amt = b.dr_amount
                etype = "Bank Dr"
            else:
                amt = b.cr_amount
                etype = "Bank Cr"

            snapshot = f"{b.entry_date}|{b.bank_narration or ''}|{b.cheque_number or ''}|{amt}"
            if b.brs_bank_particulars != snapshot:
                b.brs_bank_particulars = snapshot
                b.save(update_fields=["brs_bank_particulars"])

            adj_id = b.brs_adjustment_id
            groups[adj_id]["bank"].append(
                {
                    "entry_date": b.entry_date,
                    "entry_type": etype,
                    "cheque_no": b.cheque_number or "",
                    "narration": b.bank_narration or "",
                    "amount": amt,
                    "source": "Bank Entry",
                    "has_action": False,  # will be set later
                }
            )

        # ---------------- Build adjustments list with totals & diff ----------------
        for adj_id, data in groups.items():
            cash_rows = data["cash"]
            bank_rows = data["bank"]

            if not cash_rows and not bank_rows:
                continue

            total_cash = sum(r["amount"] for r in cash_rows)
            total_bank = sum(r["amount"] for r in bank_rows)
            diff = total_cash - total_bank

            # Sort rows inside each group by date then narration for stability
            cash_rows.sort(key=lambda r: (r["entry_date"], r["narration"]))
            bank_rows.sort(key=lambda r: (r["entry_date"], r["narration"]))

            # Decide which row gets the Action buttons (middle of all rows, excluding balance)
            all_rows = cash_rows + bank_rows
            n = len(all_rows)
            if n > 0:
                middle_index = (n - 1) // 2  # 0-based middle
                for r in all_rows:
                    r["has_action"] = False
                all_rows[middle_index]["has_action"] = True

            adjustments.append(
                {
                    "adjustment_id": adj_id,   # used by template for Edit/Undo URLs
                    "cash_rows": cash_rows,
                    "bank_rows": bank_rows,
                    "total_cash": total_cash,
                    "total_bank": total_bank,
                    "diff": diff,
                }
            )

        # Overall order of adjustments: by earliest entry_date inside adjustment
        def group_key(adj):
            dates = [r["entry_date"] for r in adj["cash_rows"] + adj["bank_rows"]]
            return min(dates) if dates else datetime.max.date()

        adjustments.sort(key=group_key)

    context = {
        "current_ulb": current_ulb,
        "current_ulb_name": current_ulb.ulb_name if current_ulb else None,
        "allowed_codes": get_allowed_codes_for(request.user),
        "ledgers": ledgers,
        "selected_ledger": selected_ledger,
        "adjustments": adjustments,
        "active_section": "brs",
    }
    return render(request, "accounts/show_brs_adjustment.html", context)

@login_required
@role_required(["ROOT_DEV", "DEV", "ADMIN", "USER"])
def brs_adjustment_undo(request, adjustment_id):
    if request.method != "POST":
        return HttpResponseForbidden("Invalid method")

    # 1) Undo cash side
    TransactionEntry.objects.filter(
        brs_adjustment_id=adjustment_id
    ).update(is_brs_reconciled=False, brs_adjustment_id=None)

    # 2) Undo bank side
    BankReconciliationEntry.objects.filter(
        brs_adjustment_id=adjustment_id
    ).update(is_reconciled=False, brs_adjustment_id=None)

    # 3) Redirect back to show_brs_adjustment list
    url = reverse("accounts_show_brs_adjustment")
    params = (
        f"?ledger_id={request.GET.get('ledger_id','')}"
        f"&entry_type={request.GET.get('entry_type','')}"
    )
    return redirect(url + params)

import calendar
from datetime import date, timedelta
def format_in_indian(number):
    """
    Format a number as Indian comma style, e.g. 1234567.89 -> '12,34,567.89'.
    Handles negatives correctly: -1234567.89 -> '-12,34,567.89'.
    Returns empty string for None or invalid.
    """
    if number is None:
        return ""

    try:
        num = float(number)
    except (TypeError, ValueError):
        return ""

    # Separate sign
    is_negative = num < 0
    num = abs(num)

    s = f"{num:.2f}"  # '3470000.00'
    if "." in s:
        whole, dec = s.split(".")
    else:
        whole, dec = s, ""

    # Apply Indian grouping to whole part
    if len(whole) > 3:
        last3 = whole[-3:]
        rest = whole[:-3]
        parts = []
        while len(rest) > 2:
            parts.insert(0, rest[-2:])
            rest = rest[:-2]
        if rest:
            parts.insert(0, rest)
        whole = ",".join(parts) + "," + last3

    result = whole + (("." + dec) if dec else "")

    if is_negative:
        result = "-" + result

    return result

from django.http import HttpResponse, HttpResponseForbidden
from django.template.loader import render_to_string
from django.conf import settings
import pdfkit

@login_required
@role_required(["ROOT_DEV", "DEV", "ADMIN", "USER"])
def brs_statement(request):
    if not user_has_code(request.user, "MENU_ACCOUNTS_BRS_STATEMENT"):
        return HttpResponseForbidden("You do not have permission for this action.")

    current_ulb_id = request.session.get("current_ulb_id")
    current_ulb = get_object_or_404(ULB, id=current_ulb_id)

    ledgers = Ledger.objects.filter(
        ulb=current_ulb,
        name__startswith="450",
    ).order_by("name")

    today = date.today()
    default_month = today.month
    default_year = today.year

    # from GET for first load, from POST when saving
    if request.method == "POST":
        month = int(request.POST.get("month") or today.month)
        year = int(request.POST.get("year") or today.year)
        ledger_id = request.POST.get("ledger_id")
    else:
        month = int(request.GET.get("month") or default_month)
        year = int(request.GET.get("year") or default_year)
        ledger_id = request.GET.get("ledger_id")

    selected_ledger = None
    month_name = calendar.month_name[month]

    # first & last date of selected month
    month_start = date(year, month, 1)
    if month == 12:
        month_end = date(year, 12, 31)
    else:
        month_end = date(year, month + 1, 1) - timedelta(days=1)

    opening_cashbook_balance = None
    closing_cashbook_balance = None
    opening_passbook_balance = None
    closing_passbook_balance = None

    opening_cashbook_amount_display = ""
    closing_cashbook_amount_display = ""
    opening_passbook_amount_display = ""
    closing_passbook_amount_display = ""

    less_cb_not_pb = []
    add_cb_not_pb = []
    less_pb_not_cb = []
    add_pb_not_cb = []

    if ledger_id:
        selected_ledger = get_object_or_404(
            Ledger,
            pk=ledger_id,
            ulb=current_ulb,
            name__startswith="450",
        )

        # --------- CASH BOOK BALANCES (ledger view style) ----------
        ledger_q = Ledger.objects.filter(id=selected_ledger.id, ulb=current_ulb)

        # opening as of month_start (same as before)
        opening_from_master = Decimal("0")
        for l in ledger_q:
            opening_from_master += _get_ledger_opening_signed(l, month_start)

        entries_before = TransactionEntry.objects.filter(
            transaction__ulb=current_ulb,
            transaction__voucher_date__lt=month_start,
            ledger__in=ledger_q,
        ).aggregate(
            dr=Sum("dr_amount"),
            cr=Sum("cr_amount"),
        )
        opening_dr_txn = Decimal(entries_before.get("dr") or 0)
        opening_cr_txn = Decimal(entries_before.get("cr") or 0)
        opening_from_txn = opening_dr_txn - opening_cr_txn

        opening_cashbook_balance = opening_from_master + opening_from_txn

        # closing as of month_end = opening + all transactions in selected month
        month_txn = TransactionEntry.objects.filter(
            transaction__ulb=current_ulb,
            transaction__voucher_date__gte=month_start,
            transaction__voucher_date__lte=month_end,
            ledger__in=ledger_q,
        ).aggregate(
            dr=Sum("dr_amount"),
            cr=Sum("cr_amount"),
        )
        month_dr = Decimal(month_txn.get("dr") or 0)
        month_cr = Decimal(month_txn.get("cr") or 0)
        delta_month = month_dr - month_cr

        closing_cashbook_balance = opening_cashbook_balance + delta_month

        opening_cashbook_amount_display = format_in_indian(opening_cashbook_balance)
        closing_cashbook_amount_display = format_in_indian(closing_cashbook_balance)

        # --------- PASS BOOK BALANCES (bank reconciliation entries) ----------
        # opening pass-book as at month_start (same as before)
        bank_open_qs = BankReconciliationEntry.objects.filter(
            ulb=current_ulb,
            ledger=selected_ledger,
            entry_date__lt=month_start,
        ).aggregate(
            dr=Sum("dr_amount"),
            cr=Sum("cr_amount"),
        )
        b_dr_total_open = bank_open_qs["dr"] or 0
        b_cr_total_open = bank_open_qs["cr"] or 0
        opening_passbook_balance = Decimal(b_cr_total_open) - Decimal(b_dr_total_open)

        # closing pass-book as at month_end = sum up to <= month_end
        bank_close_qs = BankReconciliationEntry.objects.filter(
            ulb=current_ulb,
            ledger=selected_ledger,
            entry_date__lte=month_end,
        ).aggregate(
            dr=Sum("dr_amount"),
            cr=Sum("cr_amount"),
        )
        b_dr_total_close = bank_close_qs["dr"] or 0
        b_cr_total_close = bank_close_qs["cr"] or 0
        closing_passbook_balance = Decimal(b_cr_total_close) - Decimal(b_dr_total_close)

        opening_passbook_amount_display = format_in_indian(opening_passbook_balance)
        closing_passbook_amount_display = format_in_indian(closing_passbook_balance)

        # ---------- Period range (for unreconciled lines = selected month) ----------
        period_start = month_start
        period_end = month_end + timedelta(days=1)  # < next day, for < filter

        # ---------- Unreconciled cash-book lines (TransactionEntry) ----------
        cash_qs = TransactionEntry.objects.select_related("transaction").filter(
            transaction__ulb=current_ulb,
            ledger=selected_ledger,
            transaction__voucher_date__gte=period_start,
            transaction__voucher_date__lt=period_end,
            is_brs_reconciled=False,
        )

        # 1) Less: Amount Debited in Cash Book but not Credited in Pass Book
        #    Include ALL Dr entries (any voucher_type, including Contra).
        less_cb_qs = cash_qs.filter(
            dr_amount__gt=0,
        ).order_by("transaction__voucher_date", "id")
        for e in less_cb_qs:
            if not e.dr_amount:
                continue
            less_cb_not_pb.append(
                {
                    "entry_id": e.id,
                    "date": e.transaction.voucher_date,
                    "particulars": e.brs_cash_particulars or "",
                    "voucher_no": e.transaction.voucher_no,
                    "cheque_no": getattr(
                        getattr(e.transaction, "paymentvendordetails", None),
                        "cheque_no",
                        "",
                    ),
                    "amount": e.dr_amount,
                    "amount_display": format_in_indian(e.dr_amount),
                }
            )

        # 2) Add: Amount Credited in Cash Book but not Debited in Pass Book
        #    Include ALL Cr entries (any voucher_type, including Contra).
        add_cb_qs = cash_qs.filter(
            cr_amount__gt=0,
        ).order_by("transaction__voucher_date", "id")
        for e in add_cb_qs:
            if not e.cr_amount:
                continue
            add_cb_not_pb.append(
                {
                    "entry_id": e.id,
                    "date": e.transaction.voucher_date,
                    "particulars": e.brs_cash_particulars or "",
                    "voucher_no": e.transaction.voucher_no,
                    "cheque_no": getattr(
                        getattr(e.transaction, "paymentvendordetails", None),
                        "cheque_no",
                        "",
                    ),
                    "amount": e.cr_amount,
                    "amount_display": format_in_indian(e.cr_amount),
                }
            )

        # ---------- Unreconciled bank lines (pass book) ----------
        bank_qs = BankReconciliationEntry.objects.filter(
            ulb=current_ulb,
            ledger=selected_ledger,
            entry_date__gte=period_start,
            entry_date__lt=period_end,
            is_reconciled=False,
        ).order_by("entry_date", "id")

        # 3) Less: Amount Debited in Pass Book but not Credited in Cash Book
        less_pb_qs = bank_qs.filter(dr_amount__gt=0)
        for b in less_pb_qs:
            if not b.dr_amount:
                continue
            less_pb_not_cb.append(
                {
                    "entry_id": b.id,
                    "date": b.entry_date,
                    "particulars": b.brs_bank_particulars or "",
                    "voucher_no": "",
                    "cheque_no": b.cheque_number or "",
                    "amount": b.dr_amount,
                    "amount_display": format_in_indian(b.dr_amount),
                }
            )

        # 4) Add: Amount Credited in Pass Book but not Debited in Cash Book
        add_pb_qs = bank_qs.filter(cr_amount__gt=0)
        for b in add_pb_qs:
            if not b.cr_amount:
                continue
            add_pb_not_cb.append(
                {
                    "entry_id": b.id,
                    "date": b.entry_date,
                    "particulars": b.brs_bank_particulars or "",
                    "voucher_no": "",
                    "cheque_no": b.cheque_number or "",
                    "amount": b.cr_amount,
                    "amount_display": format_in_indian(b.cr_amount),
                }
            )

    # ---------- SAVE PARTICULARS ----------
    if request.method == "POST" and selected_ledger:
        with transaction.atomic():
            # less: cash book not in pass book
            idx = 1
            for item in less_cb_not_pb:
                field_name = f"less_cb_not_pb_{idx}_particulars"
                text = (request.POST.get(field_name) or "").strip()
                if text:
                    TransactionEntry.objects.filter(
                        id=item["entry_id"],
                        ledger=selected_ledger,
                        transaction__ulb=current_ulb,
                    ).update(brs_cash_particulars=text)
                idx += 1

            # add: cash book not in pass book
            idx = 1
            for item in add_cb_not_pb:
                field_name = f"add_cb_not_pb_{idx}_particulars"
                text = (request.POST.get(field_name) or "").strip()
                if text:
                    TransactionEntry.objects.filter(
                        id=item["entry_id"],
                        ledger=selected_ledger,
                        transaction__ulb=current_ulb,
                    ).update(brs_cash_particulars=text)
                idx += 1

            # less: pass book not in cash book
            idx = 1
            for item in less_pb_not_cb:
                field_name = f"less_pb_not_cb_{idx}_particulars"
                text = (request.POST.get(field_name) or "").strip()
                if text:
                    BankReconciliationEntry.objects.filter(
                        id=item["entry_id"],
                        ulb=current_ulb,
                        ledger=selected_ledger,
                    ).update(brs_bank_particulars=text)
                idx += 1

            # add: pass book not in cash book
            idx = 1
            for item in add_pb_not_cb:
                field_name = f"add_pb_not_cb_{idx}_particulars"
                text = (request.POST.get(field_name) or "").strip()
                if text:
                    BankReconciliationEntry.objects.filter(
                        id=item["entry_id"],
                        ulb=current_ulb,
                        ledger=selected_ledger,
                    ).update(brs_bank_particulars=text)
                idx += 1

        return redirect(
            f"{request.path}?ledger_id={selected_ledger.id}&month={month}&year={year}"
        )

    context = {
        "current_ulb": current_ulb,
        "current_ulb_name": current_ulb.ulb_name,
        "allowed_codes": get_allowed_codes_for(request.user),

        "ledgers": ledgers,
        "selected_ledger": selected_ledger,

        "month": month,
        "year": year,
        "default_year": default_year,
        "month_name": month_name,

        "opening_cashbook_date": month_start,
        "closing_cashbook_date": month_end,
        "opening_passbook_date": month_start,
        "closing_passbook_date": month_end,

        "opening_cashbook_balance": opening_cashbook_balance,
        "closing_cashbook_balance": closing_cashbook_balance,
        "opening_passbook_balance": opening_passbook_balance,
        "closing_passbook_balance": closing_passbook_balance,

        "opening_cashbook_amount_display": opening_cashbook_amount_display,
        "closing_cashbook_amount_display": closing_cashbook_amount_display,
        "opening_passbook_amount_display": opening_passbook_amount_display,
        "closing_passbook_amount_display": closing_passbook_amount_display,

        "less_cb_not_pb": less_cb_not_pb,
        "add_cb_not_pb": add_cb_not_pb,
        "less_pb_not_cb": less_pb_not_cb,
        "add_pb_not_cb": add_pb_not_cb,

        "active_section": "brs",
    }

    # PDF export using same view and template (pdfkit)
    if request.GET.get("format") == "pdf":
        html_string = render_to_string(
            "accounts/brs_statement.html",
            {**context, "is_pdf": True},
            request=request,
        )

        options = {
            "page-size": "A4",
            "encoding": "UTF-8",
            "margin-top": "10mm",
            "margin-right": "10mm",
            "margin-bottom": "10mm",
            "margin-left": "10mm",
            "enable-local-file-access": "",
        }

        pdf = pdfkit.from_string(
            html_string,
            False,
            options=options,
            configuration=settings.PDFKIT_CONFIG,
        )

        response = HttpResponse(pdf, content_type="application/pdf")
        response[
            "Content-Disposition"
        ] = f'inline; filename="BRS_{current_ulb.ulb_name}_{month}_{year}.pdf"'
        return response

    return render(request, "accounts/brs_statement.html", context)

#----------------- Report Home -----------------

from .models import Ledger, FifteenthFinanceLedger  # adjust model names
@login_required
@role_required(["ROOT_DEV", "DEV", "ADMIN", "USER"])
def base_reports(request):
    """
    Reports main screen: checks MENU_ACCOUNTS_REPORT.
    Shows reports menu with items controlled by allowed_codes.
    """
    allowed_codes = get_allowed_codes_for(request.user)

    # only this code is checked here
    if "MENU_ACCOUNTS_REPORT" not in allowed_codes:
        return HttpResponseForbidden("You do not have permission for this action.")

    current_ulb_id = request.session.get("current_ulb_id")
    current_ulb = get_object_or_404(ULB, id=current_ulb_id) if current_ulb_id else None
    if not current_ulb:
        return redirect("ulb_select")

    sidebar_context = {
        "current_ulb": current_ulb,
        "current_ulb_name": current_ulb.ulb_name if current_ulb else None,
        "allowed_codes": allowed_codes,
        "active_section": "report",        # highlight Reports in accounts sidebar
    }
    context = {
        **sidebar_context,
        "active_report_section": None,      # nothing specific selected
    }
    return render(request, "accounts/base_report.html", context)

#----------------- 15th Finance Commission -----------------
@login_required
@role_required(["ROOT_DEV", "DEV", "ADMIN", "USER"])
def fifteenth_finance_commission(request):
    """
    15th Finance Commission main screen.
    Checks MENU_REPORTS_15TH_FINANCE_COMMISSION.
    """
    allowed_codes = get_allowed_codes_for(request.user)

    if "MENU_REPORTS_15TH_FINANCE_COMMISSION" not in allowed_codes:
        return HttpResponseForbidden("You do not have permission for this action.")

    current_ulb_id = request.session.get("current_ulb_id")
    current_ulb = get_object_or_404(ULB, id=current_ulb_id) if current_ulb_id else None
    if not current_ulb:
        return redirect("ulb_select")

    sidebar_context = {
        "current_ulb": current_ulb,
        "current_ulb_name": current_ulb.ulb_name if current_ulb else None,
        "allowed_codes": allowed_codes,
        "active_section": "report",
    }
    context = {
        **sidebar_context,
        "active_report_section": "15fc",    # to mark menu active
    }
    return render(request, "reports/base_15th_fc.html", context)


def get_current_fy_start(today=None):
    if today is None:
        today = date.today()
    fy_year = today.year
    april_first = date(fy_year, 4, 1)
    if today < april_first:
        fy_year -= 1
    return date(fy_year, 4, 1)
@login_required
@role_required(["ROOT_DEV", "DEV", "ADMIN", "USER"])
def fifteenth_finance_commission_define(request):
    allowed_codes = get_allowed_codes_for(request.user)

    if "MENU_15TH_FINANCE_COMMISSION_DEFINE" not in allowed_codes:
        return HttpResponseForbidden("You do not have permission for this action.")

    current_ulb_id = request.session.get("current_ulb_id")
    current_ulb = get_object_or_404(ULB, id=current_ulb_id) if current_ulb_id else None
    if not current_ulb:
        return redirect("ulb_select")

    # ledgers already defined for 15th FC
    selected_ledgers_15fc = Ledger.objects.filter(
        fifteenth_finance_mappings__ulb=current_ulb
    ).order_by("name")
    defined_ids = selected_ledgers_15fc.values_list("id", flat=True)

    # left dropdown: only 450* not yet defined
    ledgers_450_list = Ledger.objects.filter(
        ulb=current_ulb,
        name__istartswith="450",
    ).exclude(
        id__in=defined_ids
    ).order_by("name")

    # right table: all definitions
    defined_15fc_list = FifteenthFinanceLedger.objects.filter(
        ulb=current_ulb
    ).select_related("ledger").order_by("ledger__name")

    today = date.today()
    fy_start = get_current_fy_start(today)

    selected_ledger_id = None
    as_on_date = today.strftime("%Y-%m-%d")
    balance_as_on = ""
    ratio_untied = ""
    ratio_swm = ""
    ratio_rhwr = ""
    ratio_interest = ""

    if request.method == "POST":
        action = request.POST.get("action")
        sub_action = request.POST.get("sub_action")

        ledger_id_raw = request.POST.get("ledger_id") or ""
        selected_ledger_id = int(ledger_id_raw) if ledger_id_raw.isdigit() else None
        as_on_date = request.POST.get("as_on_date") or as_on_date
        balance_as_on = request.POST.get("balance_as_on") or balance_as_on
        ratio_untied = request.POST.get("ratio_untied") or ratio_untied
        ratio_swm = request.POST.get("ratio_swm") or ratio_swm
        ratio_rhwr = request.POST.get("ratio_rhwr") or ratio_rhwr
        ratio_interest = request.POST.get("ratio_interest") or ratio_interest

        def to_dec(val):
            if val is None or val == "":
                return None
            try:
                return Decimal(str(val))
            except Exception:
                return None

        # 1) Calculate balance: opening (from opening_type/opening_balance) + txns
        if action == "define" and sub_action == "calc_balance":
            if selected_ledger_id and as_on_date:
                try:
                    ledger = Ledger.objects.get(id=selected_ledger_id, ulb=current_ulb)

                    # opening balance sign based on DR/CR
                    op_bal = ledger.opening_balance or Decimal("0")
                    if ledger.opening_type == "CR":
                        op_bal = -op_bal  # credit opening = negative
                    # if DR or None, keep positive

                    tx_agg = TransactionEntry.objects.filter(
                        ledger_id=selected_ledger_id,
                        transaction__ulb=current_ulb,
                        transaction__voucher_date__lte=as_on_date,
                    ).aggregate(
                        total=Sum(F("dr_amount") - F("cr_amount"))
                    )
                    tx_total = tx_agg["total"] or Decimal("0")

                    bal = op_bal + tx_total
                    balance_as_on = f"{bal:.2f}"
                except Ledger.DoesNotExist:
                    balance_as_on = ""

        # 2) Save definition with as_on_date + balance_amount
        if action == "define" and sub_action == "define_ratio":
            if selected_ledger_id:
                try:
                    ledger = Ledger.objects.get(
                        id=selected_ledger_id,
                        ulb=current_ulb,
                        name__istartswith="450",
                    )

                    amt_untied = to_dec(ratio_untied)
                    amt_swm = to_dec(ratio_swm)
                    amt_rhwr = to_dec(ratio_rhwr)
                    amt_interest = to_dec(ratio_interest)
                    bal_dec = to_dec(balance_as_on)

                    obj, created = FifteenthFinanceLedger.objects.get_or_create(
                        ulb=current_ulb,
                        ledger=ledger,
                        defaults={
                            "defined_by": request.user,
                            "as_on_date": as_on_date or today,
                            "balance_amount": bal_dec,
                            "amount_untied": amt_untied,
                            "amount_swm": amt_swm,
                            "amount_rhwr": amt_rhwr,
                            "amount_interest": amt_interest,
                        },
                    )
                    if not created:
                        obj.as_on_date = as_on_date or obj.as_on_date or today
                        obj.balance_amount = bal_dec
                        obj.amount_untied = amt_untied
                        obj.amount_swm = amt_swm
                        obj.amount_rhwr = amt_rhwr
                        obj.amount_interest = amt_interest
                        obj.save()

                except Ledger.DoesNotExist:
                    pass

            return redirect("fifteenth_finance_commission_define")

        # 3) Undo
        if action == "undo":
            undo_id = request.POST.get("undo_ledger_id")
            if undo_id:
                FifteenthFinanceLedger.objects.filter(
                    ulb=current_ulb,
                    ledger_id=undo_id,
                ).delete()
            return redirect("fifteenth_finance_commission_define")

    sidebar_context = {
        "current_ulb": current_ulb,
        "current_ulb_name": current_ulb.ulb_name if current_ulb else None,
        "allowed_codes": allowed_codes,
        "active_section": "15fc_define",
    }

    context = {
        **sidebar_context,
        "active_report_section": "15fc",
        "active_15fc_tab": "define",
        "ledgers_450_list": ledgers_450_list,
        "selected_ledgers_15fc": selected_ledgers_15fc,
        "defined_15fc_list": defined_15fc_list,
        "selected_ledger_id": selected_ledger_id,
        "as_on_date": as_on_date,
        "balance_as_on": balance_as_on,
        "ratio_untied": ratio_untied,
        "ratio_swm": ratio_swm,
        "ratio_rhwr": ratio_rhwr,
        "ratio_interest": ratio_interest,
    }
    return render(request, "reports/fifteenth_finance_commission_define.html", context)

from .models import FifteenthFinanceLedger, FifteenthFinanceTxnRatio, FifteenthFinanceTxnAllocation
from openpyxl.utils import get_column_letter
def get_current_fy_start(today=None):
    if today is None:
        today = date.today()
    fy_year = today.year
    april_first = date(fy_year, 4, 1)
    if today < april_first:
        fy_year -= 1
    return date(fy_year, 4, 1)
@login_required
@role_required(["ROOT_DEV", "DEV", "ADMIN", "USER"])
def fifteenth_finance_transaction_define(request):
    allowed_codes = get_allowed_codes_for(request.user)
    if "MENU_15TH_FINANCE_COMMISSION_REPORT" not in allowed_codes:
        return HttpResponseForbidden("You do not have permission for this action.")

    current_ulb_id = request.session.get("current_ulb_id")
    current_ulb = get_object_or_404(ULB, id=current_ulb_id) if current_ulb_id else None
    if not current_ulb:
        return redirect("ulb_select")

    defined_ledgers = Ledger.objects.filter(
        fifteenth_finance_mappings__ulb=current_ulb
    ).order_by("name")

    today = date.today()

    first_def = FifteenthFinanceLedger.objects.filter(
        ulb=current_ulb
    ).aggregate(first_date=Min("as_on_date"))
    earliest_as_on = first_def["first_date"] or get_current_fy_start(today)

    # ---------------------- POST ----------------------
    if request.method == "POST":
        action = request.POST.get("action")

        redirect_from = request.POST.get("from_date") or earliest_as_on.strftime("%Y-%m-%d")
        redirect_to = request.POST.get("to_date") or today.strftime("%Y-%m-%d")
        redirect_ledger = request.POST.get("ledger_id") or ""
        redirect_q = request.POST.get("q") or ""
        redirect_url = (
            f"{request.path}?from_date={redirect_from}&to_date={redirect_to}"
            f"&ledger_id={redirect_ledger}&q={redirect_q}"
        )

        # unified save: full or split decided by "mode"
        if action == "save":
            tx_id_raw = request.POST.get("tx_id") or ""
            ledger_id_raw_post = request.POST.get("ledger_id") or ""
            ratio = request.POST.get("ratio") or ""
            mode = request.POST.get("mode") or "full"

            if not tx_id_raw.isdigit():
                return redirect(redirect_url)

            tx = get_object_or_404(
                Transaction.objects.select_related("ulb"),
                id=int(tx_id_raw),
            )

            if tx.ulb_id != current_ulb_id:
                return redirect(redirect_url)

            entries_qs = tx.entries.select_related("ledger")

            # if user is in a specific-ledger view, lock that ledger line;
            # else use ledger_id sent from the row (for "All" view)
            if ledger_id_raw_post.isdigit():
                selected_ledger_id_post = int(ledger_id_raw_post)
            else:
                selected_ledger_id_post = None

            if selected_ledger_id_post:
                source_entry = entries_qs.filter(
                    ledger_id=selected_ledger_id_post
                ).order_by("id").first()
            else:
                # fallback (should not really happen now that we always post ledger_id)
                if tx.voucher_type == "PYMT":
                    source_entry = entries_qs.filter(dr_amount__gt=0).order_by("id").first()
                elif tx.voucher_type in ("RECV", "CNTR"):
                    source_entry = entries_qs.filter(cr_amount__gt=0).order_by("id").first()
                elif tx.voucher_type == "JRNL":
                    source_entry = entries_qs.order_by("id").first()
                else:
                    source_entry = entries_qs.order_by("id").first()

            if not source_entry:
                return redirect(redirect_url)

            amt = (source_entry.dr_amount or Decimal("0")) + (source_entry.cr_amount or Decimal("0"))
            amt = amt.copy_abs()

            # clear existing allocations for this (tx, ledger)
            FifteenthFinanceTxnAllocation.objects.filter(
                ulb=current_ulb,
                transaction=tx,
                ledger=source_entry.ledger,
            ).delete()

            # ---------- FULL mode ----------
            if mode == "full":
                # require ratio in full mode
                if ratio not in ("40", "30_swm", "30_rhwr", "interest"):
                    return redirect(redirect_url)

                FifteenthFinanceTxnAllocation.objects.create(
                    ulb=current_ulb,
                    transaction=tx,
                    ledger=source_entry.ledger,
                    ratio_type=ratio,
                    amount=amt,
                )

            # ---------- SPLIT mode (only meaningful for RECV) ----------
            elif mode == "split":
                # auto split into 40/30/30, interest 0
                if tx.voucher_type != "RECV":
                    # for non-receipts, ignore split and just redirect
                    return redirect(redirect_url)

                amt_40 = (amt * Decimal("0.40")).quantize(Decimal("0.01"))
                amt_30_swm = (amt * Decimal("0.30")).quantize(Decimal("0.01"))
                amt_30_rhwr = (amt * Decimal("0.30")).quantize(Decimal("0.01"))

                # adjust rounding if needed
                split_sum = amt_40 + amt_30_swm + amt_30_rhwr
                diff = amt - split_sum
                if diff != Decimal("0"):
                    amt_40 += diff

                if amt_40:
                    FifteenthFinanceTxnAllocation.objects.create(
                        ulb=current_ulb,
                        transaction=tx,
                        ledger=source_entry.ledger,
                        ratio_type="40",
                        amount=amt_40,
                    )
                if amt_30_swm:
                    FifteenthFinanceTxnAllocation.objects.create(
                        ulb=current_ulb,
                        transaction=tx,
                        ledger=source_entry.ledger,
                        ratio_type="30_swm",
                        amount=amt_30_swm,
                    )
                if amt_30_rhwr:
                    FifteenthFinanceTxnAllocation.objects.create(
                        ulb=current_ulb,
                        transaction=tx,
                        ledger=source_entry.ledger,
                        ratio_type="30_rhwr",
                        amount=amt_30_rhwr,
                    )

            FifteenthFinanceTxnRatio.objects.update_or_create(
                ulb=current_ulb,
                transaction=tx,
                ledger=source_entry.ledger,
                defaults={"locked": True},
            )

            # IMPORTANT: Do NOT touch FifteenthFinanceLedger here
            return redirect(redirect_url)

    # ---------------------- GET LIST ----------------------
    ledger_id_raw = request.GET.get("ledger_id") or ""
    ledger_id = int(ledger_id_raw) if ledger_id_raw.isdigit() else None

    from_date = request.GET.get("from_date") or earliest_as_on.strftime("%Y-%m-%d")
    to_date = request.GET.get("to_date") or today.strftime("%Y-%m-%d")
    q = request.GET.get("q") or ""

    tx_qs = Transaction.objects.filter(ulb=current_ulb)

    if ledger_id:
        tx_qs = tx_qs.filter(entries__ledger_id=ledger_id).distinct()

    if from_date:
        tx_qs = tx_qs.filter(voucher_date__gte=from_date)
    if to_date:
        tx_qs = tx_qs.filter(voucher_date__lte=to_date)

    if q:
        tx_qs = tx_qs.filter(
            Q(voucher_no__icontains=q)
            | Q(entries__dr_amount__icontains=q)
            | Q(entries__cr_amount__icontains=q)
            | Q(paymentvendordetails__cheque_no__icontains=q)
        ).distinct()

    tx_qs = tx_qs.order_by("-voucher_date", "-id").prefetch_related("entries__ledger")

    # 15th-FC ledgers for this ULB
    fc_ledger_ids = set(
        FifteenthFinanceLedger.objects.filter(ulb=current_ulb).values_list(
            "ledger_id", flat=True
        )
    )

    transactions = []
    for tx in tx_qs:
        entries_qs = tx.entries.all().order_by("id")

        # only entries whose ledger is 15th-FC
        fc_entries = [e for e in entries_qs if e.ledger_id in fc_ledger_ids]
        if not fc_entries:
            continue

        # --------- pick the source entry (selected ledger line) ---------
        if ledger_id:
            # specific ledger: pick that ledger line among 15th-FC entries
            source_entry_for_display = None
            for e in fc_entries:
                if e.ledger_id == ledger_id:
                    source_entry_for_display = e
                    break
            if not source_entry_for_display:
                continue
        else:
            # All: show one row per voucher, but still tie row to that ledger explicitly
            source_entry_for_display = fc_entries[0]

        # --------- compute amount (always this ledger’s amount) ---------
        amt = (source_entry_for_display.dr_amount or Decimal("0")) + (
            source_entry_for_display.cr_amount or Decimal("0")
        )
        amount_agg = amt.copy_abs()

        # --------- opposite ledger name for particulars_first ---------
        if (source_entry_for_display.dr_amount or Decimal("0")) > 0:
            # current ledger is debit → opposite is first credit entry in voucher
            opp_candidates = [e for e in entries_qs if (e.cr_amount or Decimal("0")) > 0]
        else:
            # current ledger is credit → opposite is first debit entry
            opp_candidates = [e for e in entries_qs if (e.dr_amount or Decimal("0")) > 0]

        if opp_candidates:
            particulars_first = opp_candidates[0].ledger.name
        else:
            particulars_first = ""

        # --------- cheque, ratio, mode, locked ---------
        try:
            cheque_no = tx.paymentvendordetails.cheque_no or ""
        except Exception:
            cheque_no = ""

        default_ratio = ""
        locked = False
        has_split = False
        mode = "full"

        allocations_qs = FifteenthFinanceTxnAllocation.objects.filter(
            ulb=current_ulb,
            transaction=tx,
            ledger=source_entry_for_display.ledger,
        )

        if allocations_qs.exists():
            locked = True
            if allocations_qs.count() > 1:
                has_split = True
                mode = "split"
                default_ratio = ""
            else:
                alloc = allocations_qs.first()
                default_ratio = alloc.ratio_type or ""
                has_split = False
                mode = "full"

        # --------- NEW: compare allocated sum vs current amount ---------
        allocated_sum = allocations_qs.aggregate(
            total=Coalesce(Sum("amount"), Decimal("0.00"))
        )["total"]
        needs_update = (allocated_sum != amount_agg)

        transactions.append(
            {
                "id": tx.id,
                "date": tx.voucher_date,
                "voucher_type": tx.voucher_type,
                "voucher_no": tx.voucher_no,
                "cheque_no": cheque_no,
                "particulars_first": particulars_first,
                "narration": tx.narration,
                "amount": amount_agg,
                "ratio": default_ratio,
                "locked": locked,
                "has_split": has_split,
                "mode": mode,
                "ledger_id": source_entry_for_display.ledger_id,
                "needs_update": needs_update,  # highlight + enable ratio in template
            }
        )

    sidebar_context = {
        "current_ulb": current_ulb,
        "current_ulb_name": current_ulb.ulb_name if current_ulb else None,
        "allowed_codes": allowed_codes,
        "active_section": "report",
    }

    context = {
        **sidebar_context,
        "active_report_section": "15fc",
        "active_15fc_tab": "txn",
        "defined_ledgers": defined_ledgers,
        "transactions": transactions,
        "from_date": from_date,
        "to_date": to_date,
        "ledger_id": ledger_id,
        "q": q,
        "earliest_as_on": earliest_as_on.strftime("%Y-%m-%d"),
    }
    return render(
        request,
        "reports/fifteenth_finance_transaction_define.html",
        context,
    )

def get_current_fy_start(today=None):
    if today is None:
        today = date.today()
    fy_year = today.year
    april_first = date(fy_year, 4, 1)
    if today < april_first:
        fy_year -= 1
    return date(fy_year, 4, 1)
@login_required
@role_required(["ROOT_DEV", "DEV", "ADMIN", "USER"])
def fifteenth_fc_report(request):
    allowed_codes = get_allowed_codes_for(request.user)

    current_ulb_id = request.session.get("current_ulb_id")
    current_ulb = get_object_or_404(ULB, id=current_ulb_id) if current_ulb_id else None
    if not current_ulb:
        return redirect("ulb_select")

    defined_ledgers = Ledger.objects.filter(
        fifteenth_finance_mappings__ulb=current_ulb
    ).order_by("name")

    today = date.today()
    earliest_as_on = get_current_fy_start(today)

    # ------------- filters -------------
    ledger_id_raw = request.GET.get("ledger_id") or ""
    ledger_id = int(ledger_id_raw) if ledger_id_raw.isdigit() else None

    from_date = request.GET.get("from_date") or earliest_as_on.strftime("%Y-%m-%d")
    to_date = request.GET.get("to_date") or today.strftime("%Y-%m-%d")
    q = request.GET.get("q") or ""
    export = request.GET.get("export")  # if "excel", we return xlsx

    # base tx qs
    tx_qs = Transaction.objects.filter(
        ulb=current_ulb,
        voucher_date__gte=from_date,
        voucher_date__lte=to_date,
    ).select_related("ulb", "paymentvendordetails").prefetch_related("entries__ledger")

    # limit to transactions that touch 15th‑FC ledgers
    fc_ledger_ids = set(
        FifteenthFinanceLedger.objects.filter(ulb=current_ulb).values_list(
            "ledger_id", flat=True
        )
    )

    if ledger_id:
        tx_qs = tx_qs.filter(entries__ledger_id=ledger_id).distinct()
    else:
        tx_qs = tx_qs.filter(entries__ledger_id__in=fc_ledger_ids).distinct()

    if q:
        tx_qs = tx_qs.filter(
            Q(voucher_no__icontains=q)
            | Q(entries__dr_amount__icontains=q)
            | Q(entries__cr_amount__icontains=q)
            | Q(paymentvendordetails__cheque_no__icontains=q)
        ).distinct()

    # ---------------- opening balance from FifteenthFinanceLedger ----------------
    opening_40 = Decimal("0")
    opening_30_swm = Decimal("0")
    opening_30_rhwr = Decimal("0")
    opening_interest = Decimal("0")
    opening_total = Decimal("0")
    opening_as_on_date = None

    ff_ledger_qs = FifteenthFinanceLedger.objects.filter(ulb=current_ulb)
    if ledger_id:
        ff_ledger_qs = ff_ledger_qs.filter(ledger_id=ledger_id)

    ff_ledger = ff_ledger_qs.order_by("-as_on_date", "-id").first()
    if ff_ledger:
        opening_40 = ff_ledger.amount_untied or Decimal("0")
        opening_30_swm = ff_ledger.amount_swm or Decimal("0")
        opening_30_rhwr = ff_ledger.amount_rhwr or Decimal("0")
        opening_interest = ff_ledger.amount_interest or Decimal("0")
        opening_total = opening_40 + opening_30_swm + opening_30_rhwr + opening_interest
        opening_as_on_date = ff_ledger.as_on_date

    # ---------------- helper for opposite ledger name ----------------
    def opposite_ledger_name(tx, main_entry):
        entries = list(tx.entries.all().order_by("id"))
        if (main_entry.dr_amount or Decimal("0")) > 0:
            opp_candidates = [e for e in entries if (e.cr_amount or Decimal("0")) > 0]
        else:
            opp_candidates = [e for e in entries if (e.dr_amount or Decimal("0")) > 0]
        if opp_candidates:
            return opp_candidates[0].ledger.name
        return ""

    # ---------------- allocations (restricted ledger-wise) ----------------
    alloc_qs = FifteenthFinanceTxnAllocation.objects.filter(
        ulb=current_ulb,
        transaction__in=tx_qs,
    ).select_related("transaction", "ledger")

    if ledger_id:
        alloc_qs = alloc_qs.filter(ledger_id=ledger_id)
    else:
        alloc_qs = alloc_qs.filter(ledger_id__in=fc_ledger_ids)

    recv_rows = []
    recv_totals = {
        "40": Decimal("0"),
        "30_swm": Decimal("0"),
        "30_rhwr": Decimal("0"),
        "interest": Decimal("0"),
        "total": Decimal("0"),
    }

    pay_rows = []
    pay_totals = {
        "40": Decimal("0"),
        "30_swm": Decimal("0"),
        "30_rhwr": Decimal("0"),
        "interest": Decimal("0"),
        "total": Decimal("0"),
    }

    # Build allocations per transaction (for all voucher types)
    alloc_by_tx = {}
    for alloc in alloc_qs:
        alloc_by_tx.setdefault(alloc.transaction_id, []).append(alloc)

    # -------- build rows for ALL voucher types --------
    for tx in tx_qs.order_by("voucher_date", "id"):
        tx_allocs = alloc_by_tx.get(tx.id, [])
        if not tx_allocs:
            continue

        buckets = {
            "40": Decimal("0"),
            "30_swm": Decimal("0"),
            "30_rhwr": Decimal("0"),
            "interest": Decimal("0"),
        }
        main_alloc = None
        for alloc in tx_allocs:
            buckets[alloc.ratio_type] += alloc.amount
            if not main_alloc:
                main_alloc = alloc

        row_total = sum(buckets.values())

        entries = list(tx.entries.all().order_by("id"))
        main_entry = None
        for e in entries:
            if e.ledger_id == main_alloc.ledger_id:
                main_entry = e
                break
        if not main_entry:
            continue

        opp_name = opposite_ledger_name(tx, main_entry)

        is_receipt_side = False
        is_payment_side = False

        if tx.voucher_type == "RECV":
            is_receipt_side = True
        elif tx.voucher_type == "PYMT":
            is_payment_side = True
        elif tx.voucher_type == "CNTR":
            if (main_entry.dr_amount or Decimal("0")) > 0:
                is_receipt_side = True
            else:
                is_payment_side = True
        else:
            continue

        if is_receipt_side:
            recv_totals["40"] += buckets["40"]
            recv_totals["30_swm"] += buckets["30_swm"]
            recv_totals["30_rhwr"] += buckets["30_rhwr"]
            recv_totals["interest"] += buckets["interest"]
            recv_totals["total"] += row_total

            recv_rows.append(
                {
                    "date": tx.voucher_date,
                    "particulars": opp_name,
                    "b40": buckets["40"],
                    "b30_swm": buckets["30_swm"],
                    "b30_rhwr": buckets["30_rhwr"],
                    "interest": buckets["interest"],
                    "total": row_total,
                }
            )

        if is_payment_side:
            pay_totals["40"] += buckets["40"]
            pay_totals["30_swm"] += buckets["30_swm"]
            pay_totals["30_rhwr"] += buckets["30_rhwr"]
            pay_totals["interest"] += buckets["interest"]
            pay_totals["total"] += row_total

            pvd = getattr(tx, "paymentvendordetails", None)
            cheque_no = pvd.cheque_no if pvd and pvd.cheque_no else ""

            pay_rows.append(
                {
                    "date": tx.voucher_date,
                    "particulars": opp_name,
                    "voucher_no": tx.voucher_no,
                    "cheque_no": cheque_no,
                    "b40": buckets["40"],
                    "b30_swm": buckets["30_swm"],
                    "b30_rhwr": buckets["30_rhwr"],
                    "interest": buckets["interest"],
                    "total": row_total,
                }
            )

    # pair receipts and payments rows by index
    max_len = max(len(recv_rows), len(pay_rows))
    rows = []
    for i in range(max_len):
        recv = recv_rows[i] if i < len(recv_rows) else None
        pay = pay_rows[i] if i < len(pay_rows) else None
        rows.append(
            {
                "recv_date": recv["date"] if recv else None,
                "recv_particulars": recv["particulars"] if recv else "",
                "recv_40": recv["b40"] if recv else Decimal("0"),
                "recv_30_swm": recv["b30_swm"] if recv else Decimal("0"),
                "recv_30_rhwr": recv["b30_rhwr"] if recv else Decimal("0"),
                "recv_interest": recv["interest"] if recv else Decimal("0"),
                "recv_total": recv["total"] if recv else Decimal("0"),
                "pay_date": pay["date"] if pay else None,
                "pay_particulars": pay["particulars"] if pay else "",
                "pay_voucher_no": pay["voucher_no"] if pay else "",
                "pay_cheque_no": pay["cheque_no"] if pay else "",
                "pay_40": pay["b40"] if pay else Decimal("0"),
                "pay_30_swm": pay["b30_swm"] if pay else Decimal("0"),
                "pay_30_rhwr": pay["b30_rhwr"] if pay else Decimal("0"),
                "pay_interest": pay["interest"] if pay else Decimal("0"),
                "pay_total": pay["total"] if pay else Decimal("0"),
            }
        )

    total_receipts = recv_totals["total"]
    total_payments = pay_totals["total"]
    closing_balance = opening_total + recv_totals["total"] - pay_totals["total"]

    closing_40 = opening_40 + recv_totals["40"] - pay_totals["40"]
    closing_30_swm = opening_30_swm + recv_totals["30_swm"] - pay_totals["30_swm"]
    closing_30_rhwr = opening_30_rhwr + recv_totals["30_rhwr"] - pay_totals["30_rhwr"]
    closing_interest = opening_interest + recv_totals["interest"] - pay_totals["interest"]

    grand_recv_40 = opening_40 + recv_totals["40"]
    grand_recv_30_swm = opening_30_swm + recv_totals["30_swm"]
    grand_recv_30_rhwr = opening_30_rhwr + recv_totals["30_rhwr"]
    grand_recv_interest = opening_interest + recv_totals["interest"]

    grand_pay_40 = pay_totals["40"] + closing_40
    grand_pay_30_swm = pay_totals["30_swm"] + closing_30_swm
    grand_pay_30_rhwr = pay_totals["30_rhwr"] + closing_30_rhwr
    grand_pay_interest = pay_totals["interest"] + closing_interest

    grand_total_left = total_receipts + opening_total
    grand_total_right = total_payments + closing_balance

    # ---------- if export=excel: build and return workbook ----------
    if export == "excel":
        wb = Workbook()
        ws = wb.active
        ws.title = "15th FC"

        blue_fill = PatternFill(fill_type="solid", fgColor="1E90FF")
        bold_center_align = Alignment(horizontal="center", vertical="center")

        def fmt_date(d):
            if not d:
                return ""
            return d.strftime("%d-%m-%Y")

        # Row 1: Title A1:P1
        ws.merge_cells("A1:P1")
        ws["A1"].value = "15TH FINANCE REPORT"
        ws["A1"].font = Font(size=14, bold=True)
        ws["A1"].alignment = bold_center_align

        # Row 2: Current ULB name A2:P2
        ws.merge_cells("A2:P2")
        ws["A2"].value = current_ulb.ulb_name if current_ulb else ""
        ws["A2"].font = Font(size=12, bold=True)
        ws["A2"].alignment = bold_center_align

        # Row 3: Period A3:P3 (DD-MM-YYYY)
        ws.merge_cells("A3:P3")
        fd = datetime.fromisoformat(from_date).date()
        td = datetime.fromisoformat(to_date).date()
        ws["A3"].value = f"Period: {fmt_date(fd)} to {fmt_date(td)}"
        ws["A3"].font = Font(size=11, bold=False)
        ws["A3"].alignment = bold_center_align

        # Row 4: RECEIPTS / PAYMENT bands
        ws.merge_cells("A4:G4")
        ws["A4"].value = "RECEIPTS"
        ws["A4"].font = Font(bold=True, color="FFFFFF")
        ws["A4"].fill = blue_fill
        ws["A4"].alignment = bold_center_align

        ws.merge_cells("H4:P4")
        ws["H4"].value = "PAYMENT"
        ws["H4"].font = Font(bold=True, color="FFFFFF")
        ws["H4"].fill = blue_fill
        ws["H4"].alignment = bold_center_align

        # Row 5: column headers
        ws.append([
            "Date", "Particulars", "40% Untied", "30% SWM", "30% RHWR", "Interest Received", "Total",
            "Date", "Particulars", "Voucher No.", "Cheque No.", "40% Untied", "30% SWM", "30% RHWR", "Interest Received", "Total"
        ])

        # Opening balance row (Row 6)
        ws.append([
            fmt_date(opening_as_on_date),
            "Opening Balance",
            float(opening_40),
            float(opening_30_swm),
            float(opening_30_rhwr),
            float(opening_interest),
            float(opening_total),
            "", "", "", "", "", "", "", "", ""
        ])

        # Data rows (start at row 7)
        for r in rows:
            ws.append([
                fmt_date(r["recv_date"]),
                r["recv_particulars"] or "",
                float(r["recv_40"] or 0),
                float(r["recv_30_swm"] or 0),
                float(r["recv_30_rhwr"] or 0),
                float(r["recv_interest"] or 0),
                float(r["recv_total"] or 0),
                fmt_date(r["pay_date"]),
                r["pay_particulars"] or "",
                r["pay_voucher_no"] or "",
                r["pay_cheque_no"] or "",
                float(r["pay_40"] or 0),
                float(r["pay_30_swm"] or 0),
                float(r["pay_30_rhwr"] or 0),
                float(r["pay_interest"] or 0),
                float(r["pay_total"] or 0),
            ])

        # Blank row after data
        ws.append([])
        footer_start_row = ws.max_row + 1

        # FOOTER: Receipts side (A:G)
        ws.merge_cells(start_row=footer_start_row, start_column=1, end_row=footer_start_row, end_column=2)
        ws.cell(row=footer_start_row, column=1).value = "Total Receipts"
        ws.cell(row=footer_start_row, column=1).font = Font(bold=True)
        ws.cell(row=footer_start_row, column=3).value = float(recv_totals["40"])
        ws.cell(row=footer_start_row, column=4).value = float(recv_totals["30_swm"])
        ws.cell(row=footer_start_row, column=5).value = float(recv_totals["30_rhwr"])
        ws.cell(row=footer_start_row, column=6).value = float(recv_totals["interest"])
        ws.cell(row=footer_start_row, column=7).value = float(total_receipts)

        footer_row_cb = footer_start_row + 1

        ws.merge_cells(start_row=footer_row_cb, start_column=1, end_row=footer_row_cb, end_column=2)
        ws.cell(row=footer_row_cb, column=1).value = "Closing Balance"
        ws.cell(row=footer_row_cb, column=1).font = Font(bold=True)
        for col_idx in range(3, 8):
            ws.cell(row=footer_row_cb, column=col_idx).value = ""

        footer_row_gtl = footer_start_row + 2
        ws.merge_cells(start_row=footer_row_gtl, start_column=1, end_row=footer_row_gtl, end_column=2)
        ws.cell(row=footer_row_gtl, column=1).value = "Grand Total"
        ws.cell(row=footer_row_gtl, column=1).font = Font(bold=True)
        ws.cell(row=footer_row_gtl, column=3).value = float(grand_recv_40)
        ws.cell(row=footer_row_gtl, column=4).value = float(grand_recv_30_swm)
        ws.cell(row=footer_row_gtl, column=5).value = float(grand_recv_30_rhwr)
        ws.cell(row=footer_row_gtl, column=6).value = float(grand_recv_interest)
        ws.cell(row=footer_row_gtl, column=7).value = float(grand_total_left)

        # FOOTER: Payments side (H:P)
        ws.merge_cells(start_row=footer_start_row, start_column=8, end_row=footer_start_row, end_column=11)
        ws.cell(row=footer_start_row, column=8).value = "Total Payment"
        ws.cell(row=footer_start_row, column=8).font = Font(bold=True)
        ws.cell(row=footer_start_row, column=12).value = float(pay_totals["40"])
        ws.cell(row=footer_start_row, column=13).value = float(pay_totals["30_swm"])
        ws.cell(row=footer_start_row, column=14).value = float(pay_totals["30_rhwr"])
        ws.cell(row=footer_start_row, column=15).value = float(pay_totals["interest"])
        ws.cell(row=footer_start_row, column=16).value = float(total_payments)

        ws.merge_cells(start_row=footer_row_cb, start_column=8, end_row=footer_row_cb, end_column=11)
        ws.cell(row=footer_row_cb, column=8).value = "Closing Balance"
        ws.cell(row=footer_row_cb, column=8).font = Font(bold=True)
        ws.cell(row=footer_row_cb, column=12).value = float(closing_40)
        ws.cell(row=footer_row_cb, column=13).value = float(closing_30_swm)
        ws.cell(row=footer_row_cb, column=14).value = float(closing_30_rhwr)
        ws.cell(row=footer_row_cb, column=15).value = float(closing_interest)
        ws.cell(row=footer_row_cb, column=16).value = float(closing_balance)

        footer_row_gtr = footer_row_gtl
        ws.merge_cells(start_row=footer_row_gtr, start_column=8, end_row=footer_row_gtr, end_column=11)
        ws.cell(row=footer_row_gtr, column=8).value = "Grand Total"
        ws.cell(row=footer_row_gtr, column=8).font = Font(bold=True)
        ws.cell(row=footer_row_gtr, column=12).value = float(grand_pay_40)
        ws.cell(row=footer_row_gtr, column=13).value = float(grand_pay_30_swm)
        ws.cell(row=footer_row_gtr, column=14).value = float(grand_pay_30_rhwr)
        ws.cell(row=footer_row_gtr, column=15).value = float(grand_pay_interest)
        ws.cell(row=footer_row_gtr, column=16).value = float(grand_total_right)

        for col_idx in range(1, 17):
            col_letter = get_column_letter(col_idx)
            max_length = 0
            for row_idx in range(1, ws.max_row + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                if cell.value is not None:
                    length = len(str(cell.value))
                    if length > max_length:
                        max_length = length
            ws.column_dimensions[col_letter].width = (max_length + 2) * 1.2

        thin_side = Side(style="thin", color="000000")
        border = Border(top=thin_side, left=thin_side, right=thin_side, bottom=thin_side)
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=16):
            for cell in row:
                cell.border = border

        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = 'attachment; filename="15th_fc_report.xlsx"'
        wb.save(response)
        return response

    # ---------- normal HTML render ----------
    sidebar_context = {
        "current_ulb": current_ulb,
        "current_ulb_name": current_ulb.ulb_name if current_ulb else None,
        "allowed_codes": allowed_codes,
        "active_section": "report",
    }

    context = {
        **sidebar_context,
        "active_report_section": "15fc",
        "active_15fc_tab": "report",
        "defined_ledgers": defined_ledgers,
        "rows": rows,
        "from_date": from_date,
        "to_date": to_date,
        "ledger_id": ledger_id,
        "q": q,
        "earliest_as_on": earliest_as_on.strftime("%Y-%m-%d"),
        "opening_as_on_date": opening_as_on_date,
        "opening_40": opening_40,
        "opening_30_swm": opening_30_swm,
        "opening_30_rhwr": opening_30_rhwr,
        "opening_interest": opening_interest,
        "opening_total": opening_total,
        "recv_totals_40": recv_totals["40"],
        "recv_totals_30_swm": recv_totals["30_swm"],
        "recv_totals_30_rhwr": recv_totals["30_rhwr"],
        "recv_totals_interest": recv_totals["interest"],
        "pay_totals_40": pay_totals["40"],
        "pay_totals_30_swm": pay_totals["30_swm"],
        "pay_totals_30_rhwr": pay_totals["30_rhwr"],
        "pay_totals_interest": pay_totals["interest"],
        "closing_40": closing_40,
        "closing_30_swm": closing_30_swm,
        "closing_30_rhwr": closing_30_rhwr,
        "closing_interest": closing_interest,
        "grand_recv_40": grand_recv_40,
        "grand_recv_30_swm": grand_recv_30_swm,
        "grand_recv_30_rhwr": grand_recv_30_rhwr,
        "grand_recv_interest": grand_recv_interest,
        "grand_pay_40": grand_pay_40,
        "grand_pay_30_swm": grand_pay_30_swm,
        "grand_pay_30_rhwr": grand_pay_30_rhwr,
        "grand_pay_interest": grand_pay_interest,
        "total_receipts": total_receipts,
        "total_payments": total_payments,
        "closing_balance": closing_balance,
        "grand_total_left": grand_total_left,
        "grand_total_right": grand_total_right,
    }
    return render(request, "reports/15th_fc_report.html", context)

# ---------------------- 6th Finance Commission views ----------------------
@login_required
@role_required(["ROOT_DEV", "DEV", "ADMIN", "USER"])
def sixth_finance_commission(request):
    """
    6th Finance Commission main screen.
    Checks MENU_REPORTS_6TH_FINANCE_COMMISSION.
    """
    allowed_codes = get_allowed_codes_for(request.user)

    if "MENU_REPORTS_6TH_FINANCE_COMMISSION" not in allowed_codes:
        return HttpResponseForbidden("You do not have permission for this action.")

    current_ulb_id = request.session.get("current_ulb_id")
    current_ulb = get_object_or_404(ULB, id=current_ulb_id) if current_ulb_id else None
    if not current_ulb:
        return redirect("ulb_select")

    sidebar_context = {
        "current_ulb": current_ulb,
        "current_ulb_name": current_ulb.ulb_name if current_ulb else None,
        "allowed_codes": allowed_codes,
        "active_section": "report",
    }
    context = {
        **sidebar_context,
        "active_report_section": "6fc",   # to mark menu active in base_report
    }
    return render(request, "reports/base_6th_fc.html", context)

from .models import SixthFinanceTxnAllocation, SixthFinanceTxnRatio, SixthFinanceLedger
from django.utils.dateparse import parse_date
from django.db import transaction as db_transaction
from decimal import Decimal
def get_current_fy_start(today=None):
    if today is None:
        today = date.today()
    fy_year = today.year
    april_first = date(fy_year, 4, 1)
    if today < april_first:
        fy_year -= 1
    return date(fy_year, 4, 1)
@login_required
@role_required(["ROOT_DEV", "DEV", "ADMIN", "USER"])
def sixth_finance_commission_define(request):
    """
    6th Finance Commission - Ledger Define tab.
    Same pattern as fifteenth_finance_commission_define:
    - Left: choose 450* ledger, pick as-on date, calculate balance,
      enter / auto-split 6th FC ratios, save to SixthFinanceLedger.
    - Right: all 6th FC definitions for this ULB.
    """
    allowed_codes = get_allowed_codes_for(request.user)

    if "MENU_REPORTS_6TH_FINANCE_COMMISSION" not in allowed_codes:
        return HttpResponseForbidden("You do not have permission for this action.")

    current_ulb_id = request.session.get("current_ulb_id")
    current_ulb = get_object_or_404(ULB, id=current_ulb_id) if current_ulb_id else None
    if not current_ulb:
        return redirect("ulb_select")

    # Right table: all 6th FC definitions
    defined_6fc_list = (
        SixthFinanceLedger.objects.filter(ulb=current_ulb)
        .select_related("ledger")
        .order_by("ledger__name")
    )

    # Left dropdown: all 450* ledgers (you can also exclude already-defined if you want)
    ledgers_450_list = (
        Ledger.objects.filter(ulb=current_ulb, name__istartswith="450")
        .order_by("name")
    )

    today = date.today()
    fy_start = get_current_fy_start(today)  # not used now, but kept parallel to 15th FC

    selected_ledger_id = None
    as_on_date = today.strftime("%Y-%m-%d")
    balance_as_on = ""
    ratio_dev_total = ""
    ratio_maint = ""
    ratio_general = ""
    ratio_dev_tied_total = ""
    ratio_dev_tied_swm = ""
    ratio_dev_tied_others = ""
    ratio_dev_untied = ""

    if request.method == "POST":
        action = request.POST.get("action")
        sub_action = request.POST.get("sub_action")

        ledger_id_raw = request.POST.get("ledger_id") or ""
        selected_ledger_id = int(ledger_id_raw) if ledger_id_raw.isdigit() else None

        # keep posted values in strings (for redisplay on error or after calc)
        as_on_date = request.POST.get("as_on_date") or as_on_date
        balance_as_on = request.POST.get("balance_as_on") or balance_as_on
        ratio_dev_total = request.POST.get("ratio_dev_total") or ratio_dev_total
        ratio_maint = request.POST.get("ratio_maint") or ratio_maint
        ratio_general = request.POST.get("ratio_general") or ratio_general
        ratio_dev_tied_total = request.POST.get("ratio_dev_tied_total") or ratio_dev_tied_total
        ratio_dev_tied_swm = request.POST.get("ratio_dev_tied_swm") or ratio_dev_tied_swm
        ratio_dev_tied_others = request.POST.get("ratio_dev_tied_others") or ratio_dev_tied_others
        ratio_dev_untied = request.POST.get("ratio_dev_untied") or ratio_dev_untied

        def to_dec(val):
            if val is None or val == "":
                return None
            try:
                return Decimal(str(val))
            except Exception:
                return None

        # 1) Calculate balance: opening (opening_type/opening_balance) + txns, same as 15th FC
        if action == "define" and sub_action == "calc_balance":
            if selected_ledger_id and as_on_date:
                try:
                    ledger = Ledger.objects.get(id=selected_ledger_id, ulb=current_ulb)

                    # opening balance sign based on DR/CR
                    op_bal = ledger.opening_balance or Decimal("0")
                    if ledger.opening_type == "CR":
                        op_bal = -op_bal  # credit opening = negative

                    tx_agg = TransactionEntry.objects.filter(
                        ledger_id=selected_ledger_id,
                        transaction__ulb=current_ulb,
                        transaction__voucher_date__lte=as_on_date,
                    ).aggregate(
                        total=Sum(F("dr_amount") - F("cr_amount"))
                    )
                    tx_total = tx_agg["total"] or Decimal("0")

                    bal = op_bal + tx_total
                    balance_as_on = f"{bal:.2f}"
                except Ledger.DoesNotExist:
                    balance_as_on = ""

        # 2) Save definition with as_on_date + balance_amount
        if action == "define" and sub_action == "define_ratio":
            if selected_ledger_id:
                try:
                    ledger = Ledger.objects.get(
                        id=selected_ledger_id,
                        ulb=current_ulb,
                        name__istartswith="450",
                    )

                    bal_dec = to_dec(balance_as_on)
                    amt_dev_total = to_dec(ratio_dev_total)
                    amt_maint = to_dec(ratio_maint)
                    amt_general = to_dec(ratio_general)
                    amt_dev_tied_total = to_dec(ratio_dev_tied_total)
                    amt_dev_tied_swm = to_dec(ratio_dev_tied_swm)
                    amt_dev_tied_others = to_dec(ratio_dev_tied_others)
                    amt_dev_untied = to_dec(ratio_dev_untied)

                    with db_transaction.atomic():
                        obj, created = SixthFinanceLedger.objects.get_or_create(
                            ulb=current_ulb,
                            ledger=ledger,
                            defaults={
                                "defined_by": request.user,
                                "as_on_date": as_on_date or today,
                                "balance_amount": bal_dec,
                                "amount_dev_total": amt_dev_total,
                                "amount_maint": amt_maint,
                                "amount_general": amt_general,
                                "amount_dev_tied_total": amt_dev_tied_total,
                                "amount_dev_tied_swm": amt_dev_tied_swm,
                                "amount_dev_tied_others": amt_dev_tied_others,
                                "amount_dev_untied": amt_dev_untied,
                            },
                        )
                        if not created:
                            obj.as_on_date = as_on_date or obj.as_on_date or today
                            obj.balance_amount = bal_dec
                            obj.amount_dev_total = amt_dev_total
                            obj.amount_maint = amt_maint
                            obj.amount_general = amt_general
                            obj.amount_dev_tied_total = amt_dev_tied_total
                            obj.amount_dev_tied_swm = amt_dev_tied_swm
                            obj.amount_dev_tied_others = amt_dev_tied_others
                            obj.amount_dev_untied = amt_dev_untied
                            obj.save()

                except Ledger.DoesNotExist:
                    pass

            return redirect("sixth_finance_commission_define")

        # 3) Undo
        if action == "undo" and "BTN_6TH_FC_UNDO_LEDGER" in allowed_codes:
            undo_id = request.POST.get("undo_ledger_id")
            if undo_id:
                SixthFinanceLedger.objects.filter(
                    ulb=current_ulb,
                    ledger_id=undo_id,
                ).delete()
            return redirect("sixth_finance_commission_define")

    else:
        # GET: if you want, you can preload existing definition for a selected ledger
        ledger_id_raw = request.GET.get("ledger_id") or ""
        selected_ledger_id = int(ledger_id_raw) if ledger_id_raw.isdigit() else None

        if selected_ledger_id:
            existing = SixthFinanceLedger.objects.filter(
                ulb=current_ulb, ledger_id=selected_ledger_id
            ).select_related("ledger").first()
            if existing:
                as_on_date = (
                    existing.as_on_date.strftime("%Y-%m-%d")
                    if existing.as_on_date
                    else as_on_date
                )
                balance_as_on = (
                    f"{existing.balance_amount:.2f}"
                    if existing.balance_amount is not None
                    else ""
                )
                ratio_dev_total = existing.amount_dev_total or ""
                ratio_maint = existing.amount_maint or ""
                ratio_general = existing.amount_general or ""
                ratio_dev_tied_total = existing.amount_dev_tied_total or ""
                ratio_dev_tied_swm = existing.amount_dev_tied_swm or ""
                ratio_dev_tied_others = existing.amount_dev_tied_others or ""
                ratio_dev_untied = existing.amount_dev_untied or ""

    sidebar_context = {
        "current_ulb": current_ulb,
        "current_ulb_name": current_ulb.ulb_name if current_ulb else None,
        "allowed_codes": allowed_codes,
        "active_section": "report",
    }

    context = {
        **sidebar_context,
        "active_report_section": "6fc",
        "active_6fc_tab": "define",
        "ledgers_450_list": ledgers_450_list,
        "defined_6fc_list": defined_6fc_list,
        "selected_ledger_id": selected_ledger_id,
        "as_on_date": as_on_date,
        "balance_as_on": balance_as_on,
        "ratio_dev_total": ratio_dev_total,
        "ratio_maint": ratio_maint,
        "ratio_general": ratio_general,
        "ratio_dev_tied_total": ratio_dev_tied_total,
        "ratio_dev_tied_swm": ratio_dev_tied_swm,
        "ratio_dev_tied_others": ratio_dev_tied_others,
        "ratio_dev_untied": ratio_dev_untied,
    }
    return render(request, "reports/sixth_fc_define.html", context)

def get_current_fy_start(today=None):
    if today is None:
        today = date.today()
    fy_year = today.year
    april_first = date(fy_year, 4, 1)
    if today < april_first:
        fy_year -= 1
    return date(fy_year, 4, 1)
@login_required
@role_required(["ROOT_DEV", "DEV", "ADMIN", "USER"])
def sixth_finance_transaction_define(request):
    allowed_codes = get_allowed_codes_for(request.user)
    if "MENU_REPORTS_6TH_FINANCE_COMMISSION" not in allowed_codes:
        return HttpResponseForbidden("You do not have permission for this action.")

    current_ulb_id = request.session.get("current_ulb_id")
    current_ulb = get_object_or_404(ULB, id=current_ulb_id) if current_ulb_id else None
    if not current_ulb:
        return redirect("ulb_select")

    today = date.today()

    # 6th‑FC ledgers defined for this ULB (for dropdown)
    defined_ledgers = Ledger.objects.filter(
        sixth_finance_mappings__ulb=current_ulb
    ).order_by("name")

    # earliest as_on_date from SixthFinanceLedger (for default from_date)
    first_def = SixthFinanceLedger.objects.filter(
        ulb=current_ulb
    ).aggregate(first_date=Min("as_on_date"))
    earliest_as_on = first_def["first_date"] or get_current_fy_start(today)

    # ---------------------- POST ----------------------
    if request.method == "POST":
        action = request.POST.get("action")

        redirect_from = request.POST.get("from_date") or earliest_as_on.strftime("%Y-%m-%d")
        redirect_to = request.POST.get("to_date") or today.strftime("%Y-%m-%d")
        redirect_ledger = request.POST.get("ledger_id") or ""
        redirect_q = request.POST.get("q") or ""
        redirect_url = (
            f"{request.path}?from_date={redirect_from}&to_date={redirect_to}"
            f"&ledger_id={redirect_ledger}&q={redirect_q}"
        )

        # unified save: full or split decided by "mode"
        if action == "save":
            tx_id_raw = request.POST.get("tx_id") or ""
            ledger_id_raw_post = request.POST.get("ledger_id") or ""
            ratio = request.POST.get("ratio") or ""    # e.g. DEV_TOTAL, MAINT, GENERAL, ...
            mode = request.POST.get("mode") or "full"  # "full" or "split"

            if not tx_id_raw.isdigit():
                return redirect(redirect_url)

            tx = get_object_or_404(
                Transaction.objects.select_related("ulb"),
                id=int(tx_id_raw),
            )

            if tx.ulb_id != current_ulb_id:
                return redirect(redirect_url)

            entries_qs = tx.entries.select_related("ledger")

            # which ledger line are we allocating for this tx?
            if ledger_id_raw_post.isdigit():
                selected_ledger_id_post = int(ledger_id_raw_post)
            else:
                selected_ledger_id_post = None

            if selected_ledger_id_post:
                source_entry = entries_qs.filter(
                    ledger_id=selected_ledger_id_post
                ).order_by("id").first()
            else:
                # fallback (should rarely happen if we always send ledger_id)
                if tx.voucher_type == "PYMT":
                    source_entry = entries_qs.filter(dr_amount__gt=0).order_by("id").first()
                elif tx.voucher_type in ("RECV", "CNTR"):
                    source_entry = entries_qs.filter(cr_amount__gt=0).order_by("id").first()
                elif tx.voucher_type == "JRNL":
                    source_entry = entries_qs.order_by("id").first()
                else:
                    source_entry = entries_qs.order_by("id").first()

            if not source_entry:
                return redirect(redirect_url)

            # absolute amount for this ledger line
            amt = (source_entry.dr_amount or Decimal("0")) + (source_entry.cr_amount or Decimal("0"))
            amt = amt.copy_abs()

            # clear existing allocations for this (tx, ledger)
            SixthFinanceTxnAllocation.objects.filter(
                ulb=current_ulb,
                transaction=tx,
                ledger=source_entry.ledger,
            ).delete()

            # ---------- FULL mode ----------
            if mode == "full":
                # require a valid bucket
                if ratio not in (
                    "DEV_TOTAL",
                    "MAINT",
                    "GENERAL",
                    "DEV_TIED_TOTAL",
                    "DEV_TIED_SWM",
                    "DEV_TIED_OTHERS",
                    "DEV_UNTIED",
                ):
                    return redirect(redirect_url)

                SixthFinanceTxnAllocation.objects.create(
                    ulb=current_ulb,
                    transaction=tx,
                    ledger=source_entry.ledger,
                    ratio_type=ratio,
                    amount=amt,
                )

            # ---------- SPLIT mode ----------
            elif mode == "split":
                # Example split: full amount into Dev/Maint/General 30/20/50
                # (you can change logic if you want Dev sub‑buckets instead)
                amt_dev = (amt * Decimal("0.30")).quantize(Decimal("0.01"))
                amt_maint = (amt * Decimal("0.20")).quantize(Decimal("0.01"))
                amt_general = (amt * Decimal("0.50")).quantize(Decimal("0.01"))

                split_sum = amt_dev + amt_maint + amt_general
                diff = amt - split_sum
                if diff != Decimal("0"):
                    # adjust Dev with rounding difference
                    amt_dev += diff

                if amt_dev:
                    SixthFinanceTxnAllocation.objects.create(
                        ulb=current_ulb,
                        transaction=tx,
                        ledger=source_entry.ledger,
                        ratio_type="DEV_TOTAL",
                        amount=amt_dev,
                    )
                if amt_maint:
                    SixthFinanceTxnAllocation.objects.create(
                        ulb=current_ulb,
                        transaction=tx,
                        ledger=source_entry.ledger,
                        ratio_type="MAINT",
                        amount=amt_maint,
                    )
                if amt_general:
                    SixthFinanceTxnAllocation.objects.create(
                        ulb=current_ulb,
                        transaction=tx,
                        ledger=source_entry.ledger,
                        ratio_type="GENERAL",
                        amount=amt_general,
                    )

            # lock this (tx, ledger) as defined
            SixthFinanceTxnRatio.objects.update_or_create(
                ulb=current_ulb,
                transaction=tx,
                ledger=source_entry.ledger,
                defaults={"locked": True},
            )

            return redirect(redirect_url)

    # ---------------------- GET LIST ----------------------
    ledger_id_raw = request.GET.get("ledger_id") or ""
    ledger_id = int(ledger_id_raw) if ledger_id_raw.isdigit() else None

    from_date = request.GET.get("from_date") or earliest_as_on.strftime("%Y-%m-%d")
    to_date = request.GET.get("to_date") or today.strftime("%Y-%m-%d")
    q = request.GET.get("q") or ""

    tx_qs = Transaction.objects.filter(ulb=current_ulb)

    if ledger_id:
        tx_qs = tx_qs.filter(entries__ledger_id=ledger_id).distinct()

    if from_date:
        tx_qs = tx_qs.filter(voucher_date__gte=from_date)
    if to_date:
        tx_qs = tx_qs.filter(voucher_date__lte=to_date)

    if q:
        tx_qs = tx_qs.filter(
            Q(voucher_no__icontains=q)
            | Q(entries__dr_amount__icontains=q)
            | Q(entries__cr_amount__icontains=q)
            | Q(paymentvendordetails__cheque_no__icontains=q)
        ).distinct()

    tx_qs = tx_qs.order_by("-voucher_date", "-id").prefetch_related("entries__ledger")

    # 6th‑FC ledgers for this ULB (for filtering fc_entries)
    fc_ledger_ids = set(
        SixthFinanceLedger.objects.filter(ulb=current_ulb).values_list(
            "ledger_id", flat=True
        )
    )

    transactions = []
    for tx in tx_qs:
        entries_qs = tx.entries.all().order_by("id")

        # only entries whose ledger is 6th‑FC
        fc_entries = [e for e in entries_qs if e.ledger_id in fc_ledger_ids]
        if not fc_entries:
            continue

        # pick the source entry for display row
        if ledger_id:
            source_entry_for_display = None
            for e in fc_entries:
                if e.ledger_id == ledger_id:
                    source_entry_for_display = e
                    break
            if not source_entry_for_display:
                continue
        else:
            source_entry_for_display = fc_entries[0]

        # amount for that ledger line
        amt = (source_entry_for_display.dr_amount or Decimal("0")) + (
            source_entry_for_display.cr_amount or Decimal("0")
        )
        amount_agg = amt.copy_abs()

        # opposite ledger name for particulars_first
        if (source_entry_for_display.dr_amount or Decimal("0")) > 0:
            opp_candidates = [e for e in entries_qs if (e.cr_amount or Decimal("0")) > 0]
        else:
            opp_candidates = [e for e in entries_qs if (e.dr_amount or Decimal("0")) > 0]

        if opp_candidates:
            particulars_first = opp_candidates[0].ledger.name
        else:
            particulars_first = ""

        # cheque, ratio, mode, locked, split
        try:
            cheque_no = tx.paymentvendordetails.cheque_no or ""
        except Exception:
            cheque_no = ""

        default_ratio = ""
        locked = False
        has_split = False
        mode = "full"

        allocations_qs = SixthFinanceTxnAllocation.objects.filter(
            ulb=current_ulb,
            transaction=tx,
            ledger=source_entry_for_display.ledger,
        )

        if allocations_qs.exists():
            locked = True
            if allocations_qs.count() > 1:
                has_split = True
                mode = "split"
                default_ratio = ""
            else:
                alloc = allocations_qs.first()
                default_ratio = alloc.ratio_type or ""
                has_split = False
                mode = "full"

        # compare allocated sum vs current amount
        allocated_sum = allocations_qs.aggregate(
            total=Coalesce(Sum("amount"), Decimal("0.00"))
        )["total"]
        needs_update = (allocated_sum != amount_agg)

        transactions.append(
            {
                "id": tx.id,
                "date": tx.voucher_date,
                "voucher_type": tx.voucher_type,
                "voucher_no": tx.voucher_no,
                "cheque_no": cheque_no,
                "particulars_first": particulars_first,
                "narration": tx.narration,
                "amount": amount_agg,
                "ratio": default_ratio,          # DEV_TOTAL / MAINT / GENERAL /...
                "locked": locked,
                "has_split": has_split,
                "mode": mode,                    # "full" or "split"
                "ledger_id": source_entry_for_display.ledger_id,
                "needs_update": needs_update,
            }
        )

    sidebar_context = {
        "current_ulb": current_ulb,
        "current_ulb_name": current_ulb.ulb_name if current_ulb else None,
        "allowed_codes": allowed_codes,
        "active_section": "report",
    }

    context = {
        **sidebar_context,
        "active_report_section": "6fc",
        "active_6fc_tab": "txn",
        "defined_ledgers": defined_ledgers,
        "transactions": transactions,
        "from_date": from_date,
        "to_date": to_date,
        "ledger_id": ledger_id,
        "q": q,
        "earliest_as_on": earliest_as_on.strftime("%Y-%m-%d"),
    }
    return render(
        request,
        "reports/sixth_finance_transaction_define.html",
        context,
    )

def get_current_fy_start(today=None):
    if today is None:
        today = date.today()
    fy_year = today.year
    april_first = date(fy_year, 4, 1)
    if today < april_first:
        fy_year -= 1
    return date(fy_year, 4, 1)


@login_required
@role_required(["ROOT_DEV", "DEV", "ADMIN", "USER"])
def sixth_fc_report(request):
    allowed_codes = get_allowed_codes_for(request.user)

    current_ulb_id = request.session.get("current_ulb_id")
    current_ulb = get_object_or_404(ULB, id=current_ulb_id) if current_ulb_id else None
    if not current_ulb:
        return redirect("ulb_select")

    # Ledgers that are mapped to 6th FC
    defined_ledgers = Ledger.objects.filter(
        sixth_finance_mappings__ulb=current_ulb
    ).order_by("name")

    today = date.today()
    earliest_as_on = get_current_fy_start(today)

    # ------------- filters -------------
    ledger_id_raw = request.GET.get("ledger_id") or ""
    ledger_id = int(ledger_id_raw) if ledger_id_raw.isdigit() else None

    from_date = request.GET.get("from_date") or earliest_as_on.strftime("%Y-%m-%d")
    to_date = request.GET.get("to_date") or today.strftime("%Y-%m-%d")
    q = request.GET.get("q") or ""
    export = request.GET.get("export")  # if "excel", we return xlsx

    # base tx qs
    tx_qs = Transaction.objects.filter(
        ulb=current_ulb,
        voucher_date__gte=from_date,
        voucher_date__lte=to_date,
    ).select_related("ulb", "paymentvendordetails").prefetch_related("entries__ledger")

    # limit to transactions that touch 6th‑FC ledgers
    fc_ledger_ids = set(
        SixthFinanceLedger.objects.filter(ulb=current_ulb).values_list(
            "ledger_id", flat=True
        )
    )

    if ledger_id:
        tx_qs = tx_qs.filter(entries__ledger_id=ledger_id).distinct()
    else:
        tx_qs = tx_qs.filter(entries__ledger_id__in=fc_ledger_ids).distinct()

    if q:
        tx_qs = tx_qs.filter(
            Q(voucher_no__icontains=q)
            | Q(entries__dr_amount__icontains=q)
            | Q(entries__cr_amount__icontains=q)
            | Q(paymentvendordetails__cheque_no__icontains=q)
        ).distinct()

    # ---------------- opening balance from SixthFinanceLedger ----------------
    opening_dev_tied_swm = Decimal("0")
    opening_dev_tied_others = Decimal("0")
    opening_dev_untied = Decimal("0")
    opening_maint = Decimal("0")
    opening_general = Decimal("0")
    opening_total = Decimal("0")
    opening_as_on_date = None

    sf_ledger_qs = SixthFinanceLedger.objects.filter(ulb=current_ulb)
    if ledger_id:
        sf_ledger_qs = sf_ledger_qs.filter(ledger_id=ledger_id)

    sf_ledger = sf_ledger_qs.order_by("-as_on_date", "-id").first()
    if sf_ledger:
        opening_dev_tied_swm = sf_ledger.amount_dev_tied_swm or Decimal("0")
        opening_dev_tied_others = sf_ledger.amount_dev_tied_others or Decimal("0")
        opening_dev_untied = sf_ledger.amount_dev_untied or Decimal("0")
        opening_maint = sf_ledger.amount_maint or Decimal("0")
        opening_general = sf_ledger.amount_general or Decimal("0")
        opening_total = (
            opening_dev_tied_swm
            + opening_dev_tied_others
            + opening_dev_untied
            + opening_maint
            + opening_general
        )
        opening_as_on_date = sf_ledger.as_on_date

    # ---------------- helper for opposite ledger name ----------------
    def opposite_ledger_name(tx, main_entry):
        entries = list(tx.entries.all().order_by("id"))
        if (main_entry.dr_amount or Decimal("0")) > 0:
            opp_candidates = [e for e in entries if (e.cr_amount or Decimal("0")) > 0]
        else:
            opp_candidates = [e for e in entries if (e.dr_amount or Decimal("0")) > 0]
        if opp_candidates:
            return opp_candidates[0].ledger.name
        return ""

    # ---------------- allocations (restricted ledger-wise) ----------------
    alloc_qs = SixthFinanceTxnAllocation.objects.filter(
        ulb=current_ulb,
        transaction__in=tx_qs,
    ).select_related("transaction", "ledger")

    if ledger_id:
        alloc_qs = alloc_qs.filter(ledger_id=ledger_id)
    else:
        alloc_qs = alloc_qs.filter(ledger_id__in=fc_ledger_ids)

    recv_rows = []
    recv_totals = {
        "DEV_TIED_SWM": Decimal("0"),
        "DEV_TIED_OTHERS": Decimal("0"),
        "DEV_UNTIED": Decimal("0"),
        "MAINT": Decimal("0"),
        "GENERAL": Decimal("0"),
        "TOTAL": Decimal("0"),
    }

    pay_rows = []
    pay_totals = {
        "DEV_TIED_SWM": Decimal("0"),
        "DEV_TIED_OTHERS": Decimal("0"),
        "DEV_UNTIED": Decimal("0"),
        "MAINT": Decimal("0"),
        "GENERAL": Decimal("0"),
        "TOTAL": Decimal("0"),
    }

    # Build allocations per transaction (for all voucher types)
    alloc_by_tx = {}
    for alloc in alloc_qs:
        alloc_by_tx.setdefault(alloc.transaction_id, []).append(alloc)

    # -------- build rows for ALL voucher types --------
    for tx in tx_qs.order_by("voucher_date", "id"):
        tx_allocs = alloc_by_tx.get(tx.id, [])
        if not tx_allocs:
            continue

        buckets = {
            "DEV_TIED_SWM": Decimal("0"),
            "DEV_TIED_OTHERS": Decimal("0"),
            "DEV_UNTIED": Decimal("0"),
            "MAINT": Decimal("0"),
            "GENERAL": Decimal("0"),
        }
        main_alloc = None
        for alloc in tx_allocs:
            if alloc.ratio_type not in buckets:
                continue
            buckets[alloc.ratio_type] += alloc.amount
            if not main_alloc:
                main_alloc = alloc

        row_total = sum(buckets.values())

        entries = list(tx.entries.all().order_by("id"))
        main_entry = None
        if main_alloc:
            for e in entries:
                if e.ledger_id == main_alloc.ledger_id:
                    main_entry = e
                    break
        if not main_entry:
            continue

        opp_name = opposite_ledger_name(tx, main_entry)

        is_receipt_side = False
        is_payment_side = False

        if tx.voucher_type == "RECV":
            is_receipt_side = True
        elif tx.voucher_type == "PYMT":
            is_payment_side = True
        elif tx.voucher_type == "CNTR":
            if (main_entry.dr_amount or Decimal("0")) > 0:
                is_receipt_side = True
            else:
                is_payment_side = True
        else:
            continue

        if is_receipt_side:
            for k in ("DEV_TIED_SWM", "DEV_TIED_OTHERS", "DEV_UNTIED", "MAINT", "GENERAL"):
                recv_totals[k] += buckets[k]
            recv_totals["TOTAL"] += row_total

            recv_rows.append(
                {
                    "date": tx.voucher_date,
                    "particulars": opp_name,
                    "dev_tied_swm": buckets["DEV_TIED_SWM"],
                    "dev_tied_others": buckets["DEV_TIED_OTHERS"],
                    "dev_untied": buckets["DEV_UNTIED"],
                    "maint": buckets["MAINT"],
                    "general": buckets["GENERAL"],
                    "total": row_total,
                }
            )

        if is_payment_side:
            for k in ("DEV_TIED_SWM", "DEV_TIED_OTHERS", "DEV_UNTIED", "MAINT", "GENERAL"):
                pay_totals[k] += buckets[k]
            pay_totals["TOTAL"] += row_total

            # cheque number from PaymentVendorDetails one-to-one
            pvd = getattr(tx, "paymentvendordetails", None)
            cheque_no = pvd.cheque_no if pvd and pvd.cheque_no else ""

            pay_rows.append(
                {
                    "date": tx.voucher_date,
                    "particulars": opp_name,
                    "voucher_no": tx.voucher_no,
                    "cheque_no": cheque_no,
                    "dev_tied_swm": buckets["DEV_TIED_SWM"],
                    "dev_tied_others": buckets["DEV_TIED_OTHERS"],
                    "dev_untied": buckets["DEV_UNTIED"],
                    "maint": buckets["MAINT"],
                    "general": buckets["GENERAL"],
                    "total": row_total,
                }
            )

    # pair receipts and payments rows by index
    max_len = max(len(recv_rows), len(pay_rows))
    rows = []
    for i in range(max_len):
        recv = recv_rows[i] if i < len(recv_rows) else None
        pay = pay_rows[i] if i < len(pay_rows) else None
        rows.append(
            {
                "recv_date": recv["date"] if recv else None,
                "recv_particulars": recv["particulars"] if recv else "",
                "recv_dev_tied_swm": recv["dev_tied_swm"] if recv else Decimal("0"),
                "recv_dev_tied_others": recv["dev_tied_others"] if recv else Decimal("0"),
                "recv_dev_untied": recv["dev_untied"] if recv else Decimal("0"),
                "recv_maint": recv["maint"] if recv else Decimal("0"),
                "recv_general": recv["general"] if recv else Decimal("0"),
                "recv_total": recv["total"] if recv else Decimal("0"),
                "pay_date": pay["date"] if pay else None,
                "pay_particulars": pay["particulars"] if pay else "",
                "pay_voucher_no": pay["voucher_no"] if pay else "",
                "pay_cheque_no": pay["cheque_no"] if pay else "",
                "pay_dev_tied_swm": pay["dev_tied_swm"] if pay else Decimal("0"),
                "pay_dev_tied_others": pay["dev_tied_others"] if pay else Decimal("0"),
                "pay_dev_untied": pay["dev_untied"] if pay else Decimal("0"),
                "pay_maint": pay["maint"] if pay else Decimal("0"),
                "pay_general": pay["general"] if pay else Decimal("0"),
                "pay_total": pay["total"] if pay else Decimal("0"),
            }
        )

    total_receipts = recv_totals["TOTAL"]
    total_payments = pay_totals["TOTAL"]
    closing_balance = opening_total + total_receipts - total_payments

    closing_dev_tied_swm = (
        opening_dev_tied_swm + recv_totals["DEV_TIED_SWM"] - pay_totals["DEV_TIED_SWM"]
    )
    closing_dev_tied_others = (
        opening_dev_tied_others
        + recv_totals["DEV_TIED_OTHERS"]
        - pay_totals["DEV_TIED_OTHERS"]
    )
    closing_dev_untied = (
        opening_dev_untied + recv_totals["DEV_UNTIED"] - pay_totals["DEV_UNTIED"]
    )
    closing_maint = opening_maint + recv_totals["MAINT"] - pay_totals["MAINT"]
    closing_general = opening_general + recv_totals["GENERAL"] - pay_totals["GENERAL"]

    grand_recv_dev_tied_swm = opening_dev_tied_swm + recv_totals["DEV_TIED_SWM"]
    grand_recv_dev_tied_others = opening_dev_tied_others + recv_totals["DEV_TIED_OTHERS"]
    grand_recv_dev_untied = opening_dev_untied + recv_totals["DEV_UNTIED"]
    grand_recv_maint = opening_maint + recv_totals["MAINT"]
    grand_recv_general = opening_general + recv_totals["GENERAL"]

    grand_pay_dev_tied_swm = pay_totals["DEV_TIED_SWM"] + closing_dev_tied_swm
    grand_pay_dev_tied_others = (
        pay_totals["DEV_TIED_OTHERS"] + closing_dev_tied_others
    )
    grand_pay_dev_untied = pay_totals["DEV_UNTIED"] + closing_dev_untied
    grand_pay_maint = pay_totals["MAINT"] + closing_maint
    grand_pay_general = pay_totals["GENERAL"] + closing_general

    grand_total_left = total_receipts + opening_total
    grand_total_right = total_payments + closing_balance

    # ---------- if export=excel: export structured like web page ----------
    if export == "excel":
        wb = Workbook()
        ws = wb.active
        ws.title = "6th FC"

        blue_fill = PatternFill(fill_type="solid", fgColor="1E90FF")
        bold_center_align = Alignment(horizontal="center", vertical="center")

        def fmt_date(d):
            if not d:
                return ""
            return d.strftime("%d-%m-%Y")

        # Row 1: Title
        ws.merge_cells("A1:R1")
        ws["A1"].value = "6TH FINANCE COMMISSION REPORT"
        ws["A1"].font = Font(size=14, bold=True)
        ws["A1"].alignment = bold_center_align

        # Row 2: ULB
        ws.merge_cells("A2:R2")
        ws["A2"].value = current_ulb.ulb_name if current_ulb else ""
        ws["A2"].font = Font(size=12, bold=True)
        ws["A2"].alignment = bold_center_align

        # Row 3: Period
        ws.merge_cells("A3:R3")
        fd = datetime.fromisoformat(from_date).date()
        td = datetime.fromisoformat(to_date).date()
        ws["A3"].value = f"Period: {fmt_date(fd)} to {fmt_date(td)}"
        ws["A3"].font = Font(size=11)
        ws["A3"].alignment = bold_center_align

        # Row 4: Receipts / Payments band
        ws.merge_cells("A4:H4")
        ws["A4"].value = "RECEIPTS"
        ws["A4"].font = Font(bold=True, color="FFFFFF")
        ws["A4"].fill = blue_fill
        ws["A4"].alignment = bold_center_align

        ws.merge_cells("I4:R4")
        ws["I4"].value = "PAYMENTS"
        ws["I4"].font = Font(bold=True, color="FFFFFF")
        ws["I4"].fill = blue_fill
        ws["I4"].alignment = bold_center_align

        # Header rows to mimic HTML (3 rows)
        # Row 5
        ws.append([
            "Date", "Particulars",  # A,B
            "Development Fund 30%", "", "",  # C,D,E
            "Maintenance 20%",      # F
            "General 50%",          # G
            "Total",                # H
            "Date", "Particulars", "Voucher No.", "Cheque No.",  # I,J,K,L
            "Development Fund 30%", "", "",  # M,N,O
            "Maintenance 20%",      # P
            "General 50%",          # Q
            "Total",                # R
        ])
        ws.merge_cells("A5:A7")
        ws.merge_cells("B5:B7")
        ws.merge_cells("C5:E5")
        ws.merge_cells("F5:F7")
        ws.merge_cells("G5:G7")
        ws.merge_cells("H5:H7")
        ws.merge_cells("I5:I7")
        ws.merge_cells("J5:J7")
        ws.merge_cells("K5:K7")
        ws.merge_cells("L5:L7")
        ws.merge_cells("M5:O5")
        ws.merge_cells("P5:P7")
        ws.merge_cells("Q5:Q7")
        ws.merge_cells("R5:R7")

        # Row 6
        ws.append([
            "", "",                # A,B
            "Tied Fund 60%", "",   # C,D
            "Untied 40%",          # E
            "", "", "",            # F,G,H
            "", "", "", "",        # I,J,K,L
            "Tied Fund 60%", "",   # M,N
            "Untied 40%",          # O
            "", "", "",            # P,Q,R
        ])
        ws.merge_cells("C6:D6")
        ws.merge_cells("M6:N6")
        ws.merge_cells("E6:E7")
        ws.merge_cells("O6:O7")

        # Row 7
        ws.append([
            "", "",                # A,B
            "SWM 44%", "Others 16%", "",  # C,D,E
            "", "", "",            # F,G,H
            "", "", "", "",        # I,J,K,L
            "SWM 44%", "Others 16%", "",  # M,N,O
            "", "", "",            # P,Q,R
        ])

        # Opening row (Row 8)
        ws.append([
            fmt_date(opening_as_on_date),
            "Opening Balance",
            float(opening_dev_tied_swm),
            float(opening_dev_tied_others),
            float(opening_dev_untied),
            float(opening_maint),
            float(opening_general),
            float(opening_total),
            "", "", "", "",
            "", "", "", "", "", "",
        ])

        # Data rows (starting row 9)
        for r in rows:
            ws.append([
                fmt_date(r["recv_date"]),
                r["recv_particulars"] or "",
                float(r["recv_dev_tied_swm"] or 0),
                float(r["recv_dev_tied_others"] or 0),
                float(r["recv_dev_untied"] or 0),
                float(r["recv_maint"] or 0),
                float(r["recv_general"] or 0),
                float(r["recv_total"] or 0),
                fmt_date(r["pay_date"]),
                r["pay_particulars"] or "",
                r["pay_voucher_no"] or "",
                r["pay_cheque_no"] or "",
                float(r["pay_dev_tied_swm"] or 0),
                float(r["pay_dev_tied_others"] or 0),
                float(r["pay_dev_untied"] or 0),
                float(r["pay_maint"] or 0),
                float(r["pay_general"] or 0),
                float(r["pay_total"] or 0),
            ])

        # Blank row
        ws.append([])
        footer_start_row = ws.max_row + 1

        # ----- Left footer (Receipts) -----
        tr_row = footer_start_row + 1
        ws.merge_cells(start_row=tr_row, start_column=1, end_row=tr_row, end_column=2)
        ws.cell(row=tr_row, column=1).value = "Total Receipts"
        ws.cell(row=tr_row, column=1).font = Font(bold=True)
        ws.cell(row=tr_row, column=3).value = float(recv_totals["DEV_TIED_SWM"])
        ws.cell(row=tr_row, column=4).value = float(recv_totals["DEV_TIED_OTHERS"])
        ws.cell(row=tr_row, column=5).value = float(recv_totals["DEV_UNTIED"])
        ws.cell(row=tr_row, column=6).value = float(recv_totals["MAINT"])
        ws.cell(row=tr_row, column=7).value = float(recv_totals["GENERAL"])
        ws.cell(row=tr_row, column=8).value = float(total_receipts)

        br_row = tr_row + 1
        ws.merge_cells(start_row=br_row, start_column=1, end_row=br_row, end_column=8)

        gr_row = tr_row + 2
        ws.merge_cells(start_row=gr_row, start_column=1, end_row=gr_row, end_column=2)
        ws.cell(row=gr_row, column=1).value = "Grand Total"
        ws.cell(row=gr_row, column=1).font = Font(bold=True)
        ws.cell(row=gr_row, column=3).value = float(grand_recv_dev_tied_swm)
        ws.cell(row=gr_row, column=4).value = float(grand_recv_dev_tied_others)
        ws.cell(row=gr_row, column=5).value = float(grand_recv_dev_untied)
        ws.cell(row=gr_row, column=6).value = float(grand_recv_maint)
        ws.cell(row=gr_row, column=7).value = float(grand_recv_general)
        ws.cell(row=gr_row, column=8).value = float(grand_total_left)

        # ----- Right footer (Payments) -----
        tp_row = tr_row
        ws.merge_cells(start_row=tp_row, start_column=9, end_row=tp_row, end_column=12)
        ws.cell(row=tp_row, column=9).value = "Total Payment"
        ws.cell(row=tp_row, column=9).font = Font(bold=True)
        ws.cell(row=tp_row, column=13).value = float(pay_totals["DEV_TIED_SWM"])
        ws.cell(row=tp_row, column=14).value = float(pay_totals["DEV_TIED_OTHERS"])
        ws.cell(row=tp_row, column=15).value = float(pay_totals["DEV_UNTIED"])
        ws.cell(row=tp_row, column=16).value = float(pay_totals["MAINT"])
        ws.cell(row=tp_row, column=17).value = float(pay_totals["GENERAL"])
        ws.cell(row=tp_row, column=18).value = float(total_payments)

        cb_row = tp_row + 1
        ws.merge_cells(start_row=cb_row, start_column=9, end_row=cb_row, end_column=12)
        ws.cell(row=cb_row, column=9).value = "Closing Balance"
        ws.cell(row=cb_row, column=9).font = Font(bold=True)
        ws.cell(row=cb_row, column=13).value = float(closing_dev_tied_swm)
        ws.cell(row=cb_row, column=14).value = float(closing_dev_tied_others)
        ws.cell(row=cb_row, column=15).value = float(closing_dev_untied)
        ws.cell(row=cb_row, column=16).value = float(closing_maint)
        ws.cell(row=cb_row, column=17).value = float(closing_general)
        ws.cell(row=cb_row, column=18).value = float(closing_balance)

        gtp_row = gr_row
        ws.merge_cells(start_row=gtp_row, start_column=9, end_row=gtp_row, end_column=12)
        ws.cell(row=gtp_row, column=9).value = "Grand Total"
        ws.cell(row=gtp_row, column=9).font = Font(bold=True)
        ws.cell(row=gtp_row, column=13).value = float(grand_pay_dev_tied_swm)
        ws.cell(row=gtp_row, column=14).value = float(grand_pay_dev_tied_others)
        ws.cell(row=gtp_row, column=15).value = float(grand_pay_dev_untied)
        ws.cell(row=gtp_row, column=16).value = float(grand_pay_maint)
        ws.cell(row=gtp_row, column=17).value = float(grand_pay_general)
        ws.cell(row=gtp_row, column=18).value = float(grand_total_right)

        thin_side = Side(style="thin", color="000000")
        border = Border(top=thin_side, left=thin_side, right=thin_side, bottom=thin_side)
        for col_idx in range(1, ws.max_column + 1):
            col_letter = get_column_letter(col_idx)
            max_length = 0
            for row_idx in range(1, ws.max_row + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                if cell.value is not None:
                    length = len(str(cell.value))
                    if length > max_length:
                        max_length = length
                cell.border = border
            ws.column_dimensions[col_letter].width = (max_length + 2) * 1.2

        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = 'attachment; filename="6th_fc_report.xlsx"'
        wb.save(response)
        return response

    # ---------- normal HTML render ----------
    sidebar_context = {
        "current_ulb": current_ulb,
        "current_ulb_name": current_ulb.ulb_name if current_ulb else None,
        "allowed_codes": allowed_codes,
        "active_section": "report",
    }

    context = {
        **sidebar_context,
        "active_report_section": "6fc",
        "active_6fc_tab": "report",
        "defined_ledgers": defined_ledgers,
        "rows": rows,
        "from_date": from_date,
        "to_date": to_date,
        "ledger_id": ledger_id,
        "q": q,
        "earliest_as_on": earliest_as_on.strftime("%Y-%m-%d"),
        "opening_as_on_date": opening_as_on_date,
        "opening_dev_tied_swm": opening_dev_tied_swm,
        "opening_dev_tied_others": opening_dev_tied_others,
        "opening_dev_untied": opening_dev_untied,
        "opening_maint": opening_maint,
        "opening_general": opening_general,
        "opening_total": opening_total,
        "recv_totals_dev_tied_swm": recv_totals["DEV_TIED_SWM"],
        "recv_totals_dev_tied_others": recv_totals["DEV_TIED_OTHERS"],
        "recv_totals_dev_untied": recv_totals["DEV_UNTIED"],
        "recv_totals_maint": recv_totals["MAINT"],
        "recv_totals_general": recv_totals["GENERAL"],
        "pay_totals_dev_tied_swm": pay_totals["DEV_TIED_SWM"],
        "pay_totals_dev_tied_others": pay_totals["DEV_TIED_OTHERS"],
        "pay_totals_dev_untied": pay_totals["DEV_UNTIED"],
        "pay_totals_maint": pay_totals["MAINT"],
        "pay_totals_general": pay_totals["GENERAL"],
        "closing_dev_tied_swm": closing_dev_tied_swm,
        "closing_dev_tied_others": closing_dev_tied_others,
        "closing_dev_untied": closing_dev_untied,
        "closing_maint": closing_maint,
        "closing_general": closing_general,
        "grand_recv_dev_tied_swm": grand_recv_dev_tied_swm,
        "grand_recv_dev_tied_others": grand_recv_dev_tied_others,
        "grand_recv_dev_untied": grand_recv_dev_untied,
        "grand_recv_maint": grand_recv_maint,
        "grand_recv_general": grand_recv_general,
        "grand_pay_dev_tied_swm": grand_pay_dev_tied_swm,
        "grand_pay_dev_tied_others": grand_pay_dev_tied_others,
        "grand_pay_dev_untied": grand_pay_dev_untied,
        "grand_pay_maint": grand_pay_maint,
        "grand_pay_general": grand_pay_general,
        "total_receipts": total_receipts,
        "total_payments": total_payments,
        "closing_balance": closing_balance,
        "grand_total_left": grand_total_left,
        "grand_total_right": grand_total_right,
    }
    return render(request, "reports/6th_fc_report.html", context)

#-------Utilization Certificate-------
@login_required
@role_required(["ROOT_DEV", "DEV", "ADMIN", "USER"])
def utilization_certificate(request):
    """
    Utilization Certificate main screen.
    Checks MENU_REPORTS_UTILIZATION_CERTIFICATE.
    """
    allowed_codes = get_allowed_codes_for(request.user)

    if "MENU_REPORTS_UTILIZATION_CERTIFICATE" not in allowed_codes:
        return HttpResponseForbidden("You do not have permission for this action.")

    current_ulb_id = request.session.get("current_ulb_id")
    current_ulb = get_object_or_404(ULB, id=current_ulb_id) if current_ulb_id else None
    if not current_ulb:
        return redirect("ulb_select")

    # only set tab when explicitly provided in URL
    tab = request.GET.get("tab")  # None if not passed

    sidebar_context = {
        "current_ulb": current_ulb,
        "current_ulb_name": current_ulb.ulb_name if current_ulb else None,
        "allowed_codes": allowed_codes,
        "active_section": "report",
    }

    context = {
        **sidebar_context,
        "active_report_section": "uc",        
    }
    return render(request, "uc/base_uc.html", context)

from accounts.models import (ReceiptUCDetails, ReceiptUCUtilization, ReceiptUCUtilizationLine, PaymentUCUsage, )
@login_required
@role_required(["ROOT_DEV", "DEV", "ADMIN", "USER"])
def uc_prepared(request):
    allowed_codes = get_allowed_codes_for(request.user)

    if "MENU_REPORTS_UTILIZATION_CERTIFICATE" not in allowed_codes:
        return HttpResponseForbidden("You do not have permission for this action.")

    current_ulb_id = request.session.get("current_ulb_id")
    current_ulb = get_object_or_404(ULB, id=current_ulb_id) if current_ulb_id else None
    if not current_ulb:
        return redirect("ulb_select")

    today = date.today()
    fy_start = date(today.year if today.month >= 4 else today.year - 1, 4, 1)

    ledger_id_raw = request.GET.get("ledger_id") or ""
    ledger_id = int(ledger_id_raw) if ledger_id_raw.isdigit() else None

    from_date = request.GET.get("from_date") or fy_start.strftime("%Y-%m-%d")
    to_date = request.GET.get("to_date") or today.strftime("%Y-%m-%d")
    q = request.GET.get("q") or ""

    tx_qs = (
        Transaction.objects.filter(
            ulb=current_ulb,
            voucher_type=VoucherType.RECEIPT,
            voucher_date__gte=from_date,
            voucher_date__lte=to_date,
            receiptucdetails__uc_applicable=True,
        )
        .select_related("receiptucdetails")
        .prefetch_related("entries__ledger")
    )

    if ledger_id:
        tx_qs = tx_qs.filter(entries__ledger_id=ledger_id).distinct()

    if q:
        tx_qs = tx_qs.filter(
            Q(receiptucdetails__letter_no__icontains=q)
            | Q(voucher_no__icontains=q)
            | Q(entries__dr_amount__icontains=q)
            | Q(entries__cr_amount__icontains=q)
            | Q(entries__ledger__name__icontains=q)
        ).distinct()

    defined_ledgers = (
        Ledger.objects.filter(
            ulb=current_ulb,
            name__istartswith="450",
        ).order_by("name")
    )

    # ----------------- HANDLE SAVE (UTILIZATION) -----------------
    if request.method == "POST" and request.POST.get("action") == "save_uc_payment":
        uc_id_raw = request.POST.get("uc_id") or ""
        uc_id = int(uc_id_raw) if uc_id_raw.isdigit() else None
        if not uc_id:
            return redirect("uc_prepared")

        receipt_tx = get_object_or_404(
            Transaction,
            id=uc_id,
            ulb=current_ulb,
            voucher_type=VoucherType.RECEIPT,
        )
        receipt_uc = get_object_or_404(ReceiptUCDetails, transaction=receipt_tx)

        def parse_decimal(val):
            try:
                return Decimal(str(val).replace(",", "").strip() or "0")
            except Exception:
                return Decimal("0")

        today_str = request.POST.get("today_date") or ""
        uc_identifier = request.POST.get("uc_identifier") or ""
        try:
            today_date = (
                datetime.strptime(today_str, "%Y-%m-%d").date()
                if today_str
                else today
            )
        except ValueError:
            today_date = today

        # ALWAYS create a new header for each save
        util_obj = ReceiptUCUtilization.objects.create(
            receipt_uc=receipt_uc,
            utilized_amount=Decimal("0"),
            uc_date=today_date,
        )
        total_utilized = Decimal("0")

        voucher_ids = request.POST.getlist("voucher_ids[]")
        voucher_amounts = request.POST.getlist("voucher_amounts[]")

        pair_count = len(voucher_ids)
        excess_index = pair_count - 1 if pair_count > 0 else None

        for idx, (vid_raw, amt_raw) in enumerate(zip(voucher_ids, voucher_amounts)):
            if not vid_raw:
                continue
            amount = parse_decimal(amt_raw)
            if amount <= 0:
                continue

            pay_tx = get_object_or_404(
                Transaction,
                id=int(vid_raw),
                ulb=current_ulb,
                voucher_type=VoucherType.PAYMENT,
            )

            # NEW ROW: always create new line
            ReceiptUCUtilizationLine.objects.create(
                utilization=util_obj,
                payment_txn=pay_tx,
                amount=amount,
                created_on=today_date,
                uc_identifier=uc_identifier,
            )

            if idx != excess_index:
                total_utilized += amount

            # Track usage per payment voucher
            pay_usage, _ = PaymentUCUsage.objects.get_or_create(payment_txn=pay_tx)
            pay_usage.used_amount = (pay_usage.used_amount or Decimal("0")) + amount

            # full amount of this payment voucher (for lock logic)
            pay_total = (
                pay_tx.entries.aggregate(
                    total=Sum("cr_amount", output_field=DecimalField())
                )["total"]
                or Decimal("0")
            )
            if pay_usage.used_amount >= pay_total:
                if hasattr(pay_usage, "is_locked"):
                    pay_usage.is_locked = True
            pay_usage.save()

        util_obj.utilized_amount = total_utilized
        util_obj.uc_date = today_date
        util_obj.save()

        return redirect(
            f"{request.path}?ledger_id={ledger_id_raw}&from_date={from_date}&to_date={to_date}&q={q}"
        )

    # ----------------- UC ROWS (WITH UTILIZED + BALANCE) -----------------
    uc_rows = []
    for tx in tx_qs:
        amount = (
            tx.entries.aggregate(
                total=Sum("cr_amount", output_field=DecimalField())
            )["total"]
            or Decimal("0")
        )

        # Sum all utilization headers for this UC
        util_total = (
            ReceiptUCUtilization.objects.filter(
                receipt_uc__transaction=tx
            ).aggregate(
                total=Sum("utilized_amount", output_field=DecimalField())
            )["total"]
            or Decimal("0")
        )

        balance = amount - util_total

        first_entry = tx.entries.all().order_by("id").first()
        ledger_name = first_entry.ledger.name if first_entry and first_entry.ledger else ""

        uc = tx.receiptucdetails

        uc_rows.append(
            {
                "id": tx.id,
                "date": tx.voucher_date,
                "letter_no": uc.letter_no,
                "letter_date": uc.letter_date,
                "ledger_name": ledger_name,
                "narration": tx.narration,
                "amount": amount,
                "utilized_amount": util_total,
                "balance": balance,
            }
        )

    # ----------------- PREPARE MODE (LOAD PAYMENT VOUCHERS) -----------------
    prepare_mode = False
    selected_uc = None
    unutilized_grant = None
    payment_vouchers = []
    existing_lines = []  # unused now; always fresh

    if (
        (request.method == "POST" and request.POST.get("action") == "prepare_uc")
        or (request.method == "GET" and request.GET.get("action") == "prepare_uc")
    ):
        prepare_mode = True
        uc_id_raw = request.POST.get("uc_id") or request.GET.get("uc_id") or ""
        uc_id = int(uc_id_raw) if uc_id_raw.isdigit() else None

        selected_uc = next((row for row in uc_rows if row["id"] == uc_id), None)
        if selected_uc:
            unutilized_grant = selected_uc["balance"]
            existing_lines = []

            effective_ledger_id = ledger_id
            if not effective_ledger_id:
                tx = Transaction.objects.get(id=uc_id, ulb=current_ulb)
                first_entry = tx.entries.all().order_by("id").first()
                effective_ledger_id = (
                    first_entry.ledger_id if first_entry and first_entry.ledger_id else None
                )

            if effective_ledger_id:
                payment_qs = (
                    Transaction.objects.filter(
                        ulb=current_ulb,
                        voucher_type=VoucherType.PAYMENT,
                        voucher_date__gte=from_date,
                        voucher_date__lte=to_date,
                        entries__ledger_id=effective_ledger_id,
                    )
                    .select_related("paymentvendordetails")
                    .distinct()
                    .order_by("voucher_date", "voucher_no")
                )

                for pv in payment_qs:
                    pay_amount = (
                        pv.entries.filter(ledger_id=effective_ledger_id).aggregate(
                            total=Sum("cr_amount", output_field=DecimalField())
                        )["total"]
                        or Decimal("0")
                    )

                    pay_usage, _ = PaymentUCUsage.objects.get_or_create(payment_txn=pv)
                    used_amount = pay_usage.used_amount or Decimal("0")
                    excess_amount = pay_amount - used_amount

                    # skip fully used or locked vouchers
                    if excess_amount <= 0:
                        continue
                    if hasattr(pay_usage, "is_locked") and pay_usage.is_locked:
                        continue

                    cheque_no = ""
                    if hasattr(pv, "paymentvendordetails") and pv.paymentvendordetails:
                        cheque_no = pv.paymentvendordetails.cheque_no or ""

                    payment_vouchers.append(
                        {
                            "id": pv.id,
                            "date": pv.voucher_date.strftime("%d-%m-%Y"),
                            "voucher_no": pv.voucher_no,
                            "cheque_no": cheque_no,
                            "amount": float(excess_amount),
                            "narration": pv.narration or "",
                        }
                    )

    sidebar_context = {
        "current_ulb": current_ulb,
        "current_ulb_name": current_ulb.ulb_name if current_ulb else None,
        "allowed_codes": allowed_codes,
        "active_section": "report",
    }

    context = {
        **sidebar_context,
        "active_report_section": "uc",
        "active_uc_tab": "prepared",
        "defined_ledgers": defined_ledgers,
        "ledger_id": ledger_id,
        "from_date": from_date,
        "to_date": to_date,
        "earliest_as_on": fy_start.strftime("%Y-%m-%d"),
        "q": q,
        "uc_rows": uc_rows,
        "prepare_mode": prepare_mode,
        "selected_uc": selected_uc,
        "unutilized_grant": unutilized_grant,
        "payment_vouchers": payment_vouchers,
        "existing_lines": existing_lines,
    }
    return render(request, "uc/uc_prepared.html", context)


@login_required
@role_required(["ROOT_DEV", "DEV", "ADMIN", "USER"])
def uc_report(request):
    allowed_codes = get_allowed_codes_for(request.user)
    if "MENU_UC_REPORTS" not in allowed_codes:
        return HttpResponseForbidden("You do not have permission for this action.")

    current_ulb_id = request.session.get("current_ulb_id")
    current_ulb = get_object_or_404(ULB, id=current_ulb_id) if current_ulb_id else None
    if not current_ulb:
        return redirect("ulb_select")

    today = date.today()
    fy_start = date(today.year if today.month >= 4 else today.year - 1, 4, 1)

    from_date = request.GET.get("from_date") or fy_start.strftime("%Y-%m-%d")
    to_date = request.GET.get("to_date") or today.strftime("%Y-%m-%d")
    ledger_q = (request.GET.get("ledger_q") or "").strip()
    uc_q = (request.GET.get("uc_q") or "").strip()
    amount_q = (request.GET.get("amount_q") or "").strip()

    qs = (
        ReceiptUCUtilization.objects
        .select_related(
            "receipt_uc",
            "receipt_uc__transaction",
            "receipt_uc__transaction__receiptucdetails",
        )
        .filter(
            receipt_uc__transaction__ulb=current_ulb,
            receipt_uc__transaction__voucher_type=VoucherType.RECEIPT,
            receipt_uc__transaction__voucher_date__gte=from_date,
            receipt_uc__transaction__voucher_date__lte=to_date,
            receipt_uc__uc_applicable=True,
        )
        .order_by("receipt_uc__transaction_id", "uc_date", "id")
    )

    if ledger_q:
        qs = qs.filter(
            receipt_uc__transaction__entries__ledger__name__icontains=ledger_q
        ).distinct()

    if uc_q:
        qs = qs.filter(
            Q(receipt_uc__letter_no__icontains=uc_q)
            | Q(receipt_uc__letter_date__icontains=uc_q)
            | Q(receipt_uc__transaction__voucher_no__icontains=uc_q)
        ).distinct()

    if amount_q:
        qs = qs.filter(
            Q(receipt_uc__transaction__entries__dr_amount__icontains=amount_q)
            | Q(receipt_uc__transaction__entries__cr_amount__icontains=amount_q)
        ).distinct()

    uc_reports = []

    # track cumulative utilized per UC to compute previous balance
    cumulative_by_uc = {}

    for util_obj in qs:
        tx = util_obj.receipt_uc.transaction
        uc_key = util_obj.receipt_uc_id

        # full grant amount for this receipt
        amount = (
            tx.entries.aggregate(
                total=Sum("cr_amount", output_field=DecimalField())
            )["total"]
            or Decimal("0")
        )

        # sum of all previous headers for this UC
        prev_used = cumulative_by_uc.get(uc_key, Decimal("0"))

        # unutilized grant *before* this save
        unutilized_grant = amount - prev_used

        # this save's utilization
        this_used = util_obj.utilized_amount or Decimal("0")

        balance = unutilized_grant - this_used

        # update cumulative for next rows of same UC
        cumulative_by_uc[uc_key] = prev_used + this_used

        uc_date = util_obj.uc_date or tx.voucher_date

        first_entry = tx.entries.all().order_by("id").first()
        ledger_name = first_entry.ledger.name if first_entry and first_entry.ledger else ""

        uc = tx.receiptucdetails
        letter_no = uc.letter_no or ""
        letter_date_str = uc.letter_date.strftime("%d-%m-%Y") if uc.letter_date else ""

        if letter_no and letter_date_str:
            uc_id_display = f"{letter_no} / {letter_date_str}"
        elif letter_no:
            uc_id_display = letter_no
        elif letter_date_str:
            uc_id_display = letter_date_str
        else:
            uc_id_display = tx.voucher_no

        uc_reports.append(
            {
                "id": util_obj.id,
                "receipt_uc_id": util_obj.receipt_uc_id,
                "uc_date": uc_date,
                "uc_id": uc_id_display,
                "ledger_name": ledger_name,
                "unutilized_grant": unutilized_grant,
                "total_payment": this_used,
                "balance": balance,
            }
        )

    sidebar_context = {
        "current_ulb": current_ulb,
        "current_ulb_name": current_ulb.ulb_name if current_ulb else None,
        "allowed_codes": allowed_codes,
        "active_section": "report",
    }

    context = {
        **sidebar_context,
        "active_report_section": "uc",
        "active_uc_tab": "reports",
        "from_date": from_date,
        "to_date": to_date,
        "ledger_q": ledger_q,
        "uc_q": uc_q,
        "amount_q": amount_q,
        "uc_reports": uc_reports,
    }
    return render(request, "uc/uc_report.html", context)

@login_required
@role_required(["ROOT_DEV", "DEV", "ADMIN", "USER"])
def uc_undo_last(request, util_id):
    if request.method != "POST":
        return HttpResponseForbidden("Invalid method.")

    allowed_codes = get_allowed_codes_for(request.user)
    if "MENU_UC_REPORTS" not in allowed_codes:
        return HttpResponseForbidden("You do not have permission for this action.")

    current_ulb_id = request.session.get("current_ulb_id")
    current_ulb = get_object_or_404(ULB, id=current_ulb_id) if current_ulb_id else None
    if not current_ulb:
        return redirect("ulb_select")

    util_obj = get_object_or_404(
        ReceiptUCUtilization,
        id=util_id,
        receipt_uc__transaction__ulb=current_ulb,
        receipt_uc__transaction__voucher_type=VoucherType.RECEIPT,
    )

    lines = list(
        ReceiptUCUtilizationLine.objects.filter(utilization=util_obj)
    )
    for line in lines:
        pay_usage, _ = PaymentUCUsage.objects.get_or_create(
            payment_txn=line.payment_txn
        )
        pay_usage.used_amount = (pay_usage.used_amount or Decimal("0")) - (
            line.amount or Decimal("0")
        )
        if pay_usage.used_amount < 0:
            pay_usage.used_amount = Decimal("0")
        if hasattr(pay_usage, "is_locked") and pay_usage.used_amount <= 0:
            pay_usage.is_locked = False
        pay_usage.save()

    ReceiptUCUtilizationLine.objects.filter(utilization=util_obj).delete()
    util_obj.delete()

    return redirect("uc_report")


@login_required
@role_required(["ROOT_DEV", "DEV", "ADMIN", "USER"])
def uc_report_export_excel(request, util_id):

    allowed_codes = get_allowed_codes_for(request.user)
    if "MENU_REPORTS_UTILIZATION_CERTIFICATE" not in allowed_codes:
        return HttpResponseForbidden("You do not have permission for this action.")

    current_ulb_id = request.session.get("current_ulb_id")
    current_ulb = get_object_or_404(ULB, id=current_ulb_id)

    util_obj = get_object_or_404(
        ReceiptUCUtilization,
        id=util_id,
        receipt_uc__transaction__ulb=current_ulb,
    )

    receipt_uc: ReceiptUCDetails = util_obj.receipt_uc
    receipt_tx = receipt_uc.transaction

    grant_amount = receipt_uc.grant_amount or Decimal("0")

    previous_utilized = (
        ReceiptUCUtilization.objects.filter(
            receipt_uc=receipt_uc,
            uc_date__lt=util_obj.uc_date,
        ).aggregate(total=Sum("utilized_amount"))["total"]
        or Decimal("0")
    )

    unutilized_grant = grant_amount - previous_utilized

    lines = list(
        ReceiptUCUtilizationLine.objects.filter(utilization=util_obj)
        .select_related("payment_txn", "payment_txn__paymentvendordetails")
        .order_by("created_on", "id")
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "UC Report"

    bold = Font(bold=True)
    center = Alignment(horizontal="center")
    right = Alignment(horizontal="right")
    left = Alignment(horizontal="left")
    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_fill = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")

    def fmt_inr(n: Decimal) -> str:
        s = f"{n:.2f}"
        whole, frac = s.split(".")
        neg = whole.startswith("-")
        if neg:
            whole = whole[1:]
        if len(whole) > 3:
            head = whole[:-3]
            tail = whole[-3:]
            parts = []
            while len(head) > 2:
                parts.insert(0, head[-2:])
                head = head[:-2]
            if head:
                parts.insert(0, head)
            whole = ",".join(parts) + "," + tail
        return ("-" if neg else "") + whole + "." + frac

    row = 1

    # ULB name
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
    cell = ws.cell(row=row, column=1, value=current_ulb.ulb_name)
    cell.font = Font(bold=True, size=12)
    cell.alignment = center
    row += 1

    # Title: UC Payment details
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
    cell = ws.cell(row=row, column=1, value="UC Payment details")
    cell.font = Font(bold=True, size=11)
    cell.alignment = center
    row += 1

    # Ledger (without code), centered A:D
    first_entry = receipt_tx.entries.first()
    ledger_name = ""
    if first_entry and first_entry.ledger:
        parts = first_entry.ledger.name.split(" ", 1)
        ledger_name = parts[1] if len(parts) == 2 else parts[0]

    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
    cell = ws.cell(row=row, column=1, value=ledger_name)
    cell.font = bold
    cell.alignment = center
    row += 1

    # UC Date (dd-mm-yyyy)
    uc_date = util_obj.uc_date.date()
    uc_date_str = uc_date.strftime("%d-%m-%Y")
    ws.cell(row=row, column=1, value="UC Date").font = bold
    ws.cell(row=row, column=2, value=uc_date_str).alignment = left
    row += 1

    # Header row
    headers = ["Voucher Date", "Voucher Number", "Cheque Number", "Amount"]
    header_row = row
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=header_row, column=col, value=h)
        c.font = bold
        c.alignment = center
        c.border = border
        c.fill = header_fill
    row += 1

    # First data row: UC ID + Unutilized grant amount
    uc_id = f"{receipt_uc.letter_no} / {receipt_uc.letter_date}"
    ws.cell(row=row, column=1, value=uc_id).alignment = left
    ws.cell(row=row, column=2, value="Unutilized grant amount").alignment = left
    c = ws.cell(row=row, column=4, value=fmt_inr(unutilized_grant))
    c.alignment = right
    for col in range(1, 5):
        ws.cell(row=row, column=col).border = border
    row += 1

    remaining = unutilized_grant
    total_payment = Decimal("0")
    excess_amount = Decimal("0")
    excess_tx = None

    # Payment vouchers
    for line in lines:
        tx = line.payment_txn
        amt = line.amount or Decimal("0")

        cheque = ""
        if hasattr(tx, "paymentvendordetails"):
            cheque = tx.paymentvendordetails.cheque_no or ""

        usable = min(amt, remaining)
        excess = amt - usable

        if usable <= 0:
            if excess > 0:
                excess_amount += excess
                excess_tx = tx
            continue

        voucher_date_str = tx.voucher_date.strftime("%d-%m-%Y")

        ws.cell(row=row, column=1, value=voucher_date_str).alignment = left
        ws.cell(row=row, column=2, value=tx.voucher_no).alignment = left
        ws.cell(row=row, column=3, value=cheque).alignment = left
        c = ws.cell(row=row, column=4, value=fmt_inr(usable))
        c.alignment = right

        for col in range(1, 5):
            ws.cell(row=row, column=col).border = border

        total_payment += usable
        remaining -= usable

        if excess > 0:
            excess_amount += excess
            excess_tx = tx

        row += 1

    # Total payment
    c = ws.cell(row=row, column=2, value="Total payment")
    c.font = bold
    c.alignment = left
    c = ws.cell(row=row, column=4, value=fmt_inr(total_payment))
    c.font = bold
    c.alignment = right
    for col in range(1, 5):
        ws.cell(row=row, column=col).border = border
    row += 1

    # Balance unutilised
    balance = unutilized_grant - total_payment
    c = ws.cell(row=row, column=2, value="Balance unutilised")
    c.font = bold
    c.alignment = left
    c = ws.cell(row=row, column=4, value=fmt_inr(balance))
    c.font = bold
    c.alignment = right
    for col in range(1, 5):
        ws.cell(row=row, column=col).border = border
    row += 1

    # Excess voucher amount row only if exists
    if excess_amount > 0 and excess_tx:
        cheque = ""
        if hasattr(excess_tx, "paymentvendordetails"):
            cheque = excess_tx.paymentvendordetails.cheque_no or ""

        voucher_date_str = excess_tx.voucher_date.strftime("%d-%m-%Y")
        ws.cell(row=row, column=1, value=voucher_date_str).alignment = left
        ws.cell(row=row, column=2, value=excess_tx.voucher_no).alignment = left
        ws.cell(row=row, column=3, value="Excess voucher amount").alignment = left
        c = ws.cell(row=row, column=4, value=fmt_inr(excess_amount))
        c.font = bold
        c.alignment = right
        for col in range(1, 5):
            ws.cell(row=row, column=col).border = border

    # Auto-fit widths
    for col_idx in range(1, 5):
        max_len = 0
        col_letter = get_column_letter(col_idx)
        for cell in ws[col_letter]:
            if cell.value is None:
                continue
            val_str = str(cell.value)
            if len(val_str) > max_len:
                max_len = len(val_str)
        ws.column_dimensions[col_letter].width = max_len + 2

    ws.freeze_panes = ws[f"A{header_row + 1}"]

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename="uc_report_{util_id}.xlsx"'
    wb.save(response)
    return response


from django.http import Http404, HttpResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.db.models import Sum, DecimalField, Sum as Sum2
from django.template.loader import get_template
from django.conf import settings
# ... your other imports ...

def format_indian(num):
    if num is None:
        return ""
    try:
        n = float(num)
    except (TypeError, ValueError):
        return num
    s = f"{n:.2f}"
    before, after = s.split(".")
    last3 = before[-3:]
    rest = before[:-3]
    if rest:
        parts = []
        while len(rest) > 2:
            parts.insert(0, rest[-2:])
            rest = rest[:-2]
        if rest:
            parts.insert(0, rest)
        before = ",".join(parts) + "," + last3
    else:
        before = last3
    return before + "." + after

@login_required
@role_required(["ROOT_DEV", "DEV", "ADMIN", "USER"])
def uc_btc42a_form(request):
    # permissions
    allowed_codes = get_allowed_codes_for(request.user)
    if "BTN_UC_OPEN_FORMS" not in allowed_codes and "MENU_UC_REPORTS" not in allowed_codes:
        raise Http404("You do not have permission for this action.")

    current_ulb_id = request.session.get("current_ulb_id")
    current_ulb = get_object_or_404(ULB, id=current_ulb_id) if current_ulb_id else None
    if not current_ulb:
        return redirect("ulb_select")

    uc_id = request.GET.get("uc_id")
    if not uc_id:
        raise Http404("UC id is required")

    uc = get_object_or_404(ReceiptUCDetails, id=uc_id)
    tx = uc.transaction

    major_head = uc.major_head or ""
    sub_major_head = uc.sub_major_head or ""
    minor_head = uc.minor_head or ""
    sub_head = uc.sub_head or ""

    treasury_code = uc.treasury_code or ""
    ddo_code = uc.ddo_code or ""
    bill_code = uc.bill_code or ""
    head_code = ""

    uc_number = uc.letter_no or uc_id
    uc_date = uc.uc_bill_date or uc.letter_date or tx.voucher_date

    if uc_date:
        month_name = uc_date.strftime("%B")
        year_text = uc_date.strftime("%Y")
    else:
        month_name = ""
        year_text = ""

    office_name = current_ulb.ulb_name
    sanctioned_amount = uc.grant_amount or 0

    total_credit = tx.entries.aggregate(
        total=Sum("cr_amount", output_field=DecimalField())
    )["total"] or 0

    prev_unspent = 0
    utilized_amount = (
        uc.utilizations.aggregate(total=Sum2("utilized_amount"))["total"]
        if hasattr(uc, "utilizations")
        else 0
    ) or 0

    unutilised_balance = (sanctioned_amount or 0) - utilized_amount
    surrendered_amount = 0
    surrendered_letter_no = ""
    surrendered_letter_date = None

    # formatted (Indian) display values
    sanctioned_amount_disp = format_indian(sanctioned_amount)
    utilized_amount_disp = format_indian(utilized_amount)
    unutilised_balance_disp = format_indian(unutilised_balance)
    surrendered_amount_disp = format_indian(surrendered_amount)

    letter_no = uc.letter_no or ""
    letter_date = uc.letter_date

    grant_rows = [
        {
            "sanction_letter": f"{letter_no} / {letter_date.strftime('%d-%m-%Y')}" if letter_no and letter_date else letter_no,
            "grantee_name": office_name,
            "purpose": "",
            "bill_no_date": f"{uc.uc_bill_no} / {uc.uc_bill_date.strftime('%d-%m-%Y')}" if uc.uc_bill_no and uc.uc_bill_date else uc.uc_bill_no,
            "grant_drawn": sanctioned_amount_disp,
            "tv_no_date": f"{uc.tv_no} / {uc.tv_date.strftime('%d-%m-%Y')}" if uc.tv_no and uc.tv_date else uc.tv_no,
            "amount_uc": utilized_amount_disp,
            "balance": unutilised_balance_disp,
            "amount_surrendered": surrendered_amount_disp,
        }
    ]

    sidebar_context = {
        "current_ulb": current_ulb,
        "current_ulb_name": current_ulb.ulb_name if current_ulb else None,
        "allowed_codes": allowed_codes,
        "active_section": "report",
    }

    btc_context = {
        "major_head": major_head,
        "sub_major_head": sub_major_head,
        "minor_head": minor_head,
        "sub_head": sub_head,
        "treasury_code": treasury_code,
        "ddo_code": ddo_code,
        "bill_code": bill_code,
        "head_code": head_code,
        "uc_number": uc_number,
        "uc_date": uc_date,
        "month_name": month_name,
        "year_text": year_text,
        "office_name": office_name,
        "letter_no": letter_no,
        "letter_date": letter_date,
        "sanctioned_amount": sanctioned_amount_disp,
        "grantee_name": office_name,
        "prev_unspent": prev_unspent,
        "utilized_amount": utilized_amount_disp,
        "unutilised_balance": unutilised_balance_disp,
        "surrendered_amount": surrendered_amount_disp,
        "surrendered_letter_no": surrendered_letter_no,
        "surrendered_letter_date": surrendered_letter_date,
        "grant_rows": grant_rows,
        "active_report_section": "uc",
        "active_uc_tab": "reports",
        "uc_id": uc.id,
    }

    context = {**sidebar_context, **btc_context}

    export = request.GET.get("export")
    if export == "pdf":
        # On Railway we disable WeasyPrint via USE_WEASYPRINT
        if not getattr(settings, "USE_WEASYPRINT", True):
            return HttpResponse(
                "PDF generation is not available on this server.",
                status=503,
            )

        from weasyprint import HTML  # local import so it doesn't run at startup

        template = get_template("uc/uc_btc42a_pdf.html")
        html_string = template.render(context, request)
        base_url = request.build_absolute_uri("/")
        pdf_bytes = HTML(string=html_string, base_url=base_url).write_pdf()
        response = HttpResponse(pdf_bytes, content_type="application/pdf")
        response["Content-Disposition"] = f'attachment; filename="btc42a_uc_{uc.id}.pdf"'
        return response

    return render(request, "uc/uc_btc42a_form.html", context)

